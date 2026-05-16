from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from app.excel_writer import safe_write_cell
from app.logger import get_logger


logger = get_logger(__name__)


def _is_empty(value):
    return value is None or value == ""


def _style_written_cell(ws, cell_ref, value):
    cell = ws[cell_ref]
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    try:
        column_letter, row = coordinate_from_string(cell_ref)
        column_index = column_index_from_string(column_letter)
        width = 18 if len(str(value or "")) <= 20 else 48
        current_width = ws.column_dimensions[column_letter].width or 0
        ws.column_dimensions[column_letter].width = max(current_width, width)
        ws.row_dimensions[row].height = 48 if "\n" in str(value or "") or len(str(value or "")) > 60 else 24

        if column_index == 1:
            ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width or 0, 18)
        else:
            ws.column_dimensions[get_column_letter(column_index)].width = max(
                ws.column_dimensions[get_column_letter(column_index)].width or 0,
                width,
            )
    except Exception:
        logger.debug("V4 rule executor cell style skipped: cell_ref=%s", cell_ref)


def execute_excel_rule_preview_to_workbook(
    operations: list,
    output_path: str,
) -> dict:
    warnings = []
    operations_written = 0

    try:
        if not isinstance(operations, list):
            raise ValueError("operations must be a list")
        if _is_empty(output_path):
            raise ValueError("output_path must not be empty")

        logger.info("V4 Excel rule executor started: output_path=%s", output_path)

        wb = Workbook()
        ws = wb.active
        ws.title = "V4 Rule Executor"

        for index, operation in enumerate(operations):
            if not isinstance(operation, dict):
                warnings.append(f"operation[{index}] 不是 object，已跳过")
                continue

            rule_id = operation.get("rule_id") or f"operation[{index}]"
            target_cell = operation.get("target_cell")
            value = operation.get("value")

            if _is_empty(target_cell):
                warnings.append(f"规则 {rule_id} 无 target_cell，已跳过")
                continue

            if _is_empty(value):
                warnings.append(f"规则 {rule_id} 无 value，已跳过")
                continue

            written_cell = safe_write_cell(ws, target_cell, value)
            if not written_cell:
                warnings.append(f"规则 {rule_id} 写入 cell 为空，已跳过")
                continue

            _style_written_cell(ws, written_cell, value)
            operations_written += 1

        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

        logger.info(
            "V4 Excel rule executor succeeded: output_path=%s operations_written=%s warnings=%s",
            path,
            operations_written,
            len(warnings),
        )
        return {
            "success": True,
            "output_path": str(path),
            "operations_written": operations_written,
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception("V4 Excel rule executor failed: output_path=%s", output_path)
        return {
            "success": False,
            "error": str(exc),
            "operations_written": 0,
            "warnings": warnings,
        }

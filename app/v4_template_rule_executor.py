from pathlib import Path
import re
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_from_string

from app.excel_writer import safe_write_cell
from app.logger import get_logger
from app.runtime_paths import get_base_dir
from app.v4_excel_rule_preview import build_excel_rule_preview
from app.v4_excel_rules import load_excel_render_rules
from app.v4_examples import load_example
from app.v4_renderer import render_example_to_description_fields
from app.v4_schema import load_product_schema


logger = get_logger(__name__)


def _is_empty(value):
    return value is None or value == ""


def _safe_output_filename_part(value: str) -> str:
    text = str(value or "").strip()
    if text.endswith(".json"):
        text = text[:-5]
    text = re.sub(r"[^\w.-]+", "_", text, flags=re.UNICODE)
    return text.strip("._") or "output"


def _style_written_cell(ws, cell_ref, value):
    try:
        cell = ws[cell_ref]
        cell.alignment = Alignment(wrap_text=True, vertical="top")

        column_letter, row = coordinate_from_string(cell_ref)
        text_value = str(value or "")
        current_width = ws.column_dimensions[column_letter].width or 0
        if len(text_value) > 40:
            ws.column_dimensions[column_letter].width = max(current_width, 48)

        if "\n" in text_value or len(text_value) > 60:
            line_count = max(1, text_value.count("\n") + 1)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 0, min(120, 24 * line_count))
    except Exception:
        logger.debug("V4 template rule executor cell style skipped: cell_ref=%s", cell_ref)


def execute_rules_to_template_excel(
    template_path: str,
    operations: list,
    output_path: str,
) -> dict:
    warnings = []
    operations_written = 0

    try:
        if _is_empty(template_path):
            raise ValueError("template_path must not be empty")
        if not isinstance(operations, list):
            raise ValueError("operations must be a list")
        if _is_empty(output_path):
            raise ValueError("output_path must not be empty")

        template_file = Path(template_path)
        if not template_file.is_file():
            raise FileNotFoundError(f"template file not found: {template_path}")

        logger.info(
            "V4 template rule executor started: template_path=%s output_path=%s",
            template_path,
            output_path,
        )

        wb = load_workbook(template_file)
        ws = wb.active

        for index, operation in enumerate(operations):
            if not isinstance(operation, dict):
                warnings.append(f"operation[{index}] 不是 object，已跳过")
                continue

            rule_id = operation.get("rule_id") or f"operation[{index}]"
            target_cell = operation.get("target_cell")
            value = operation.get("value")

            if _is_empty(target_cell):
                warnings.append(f"规则 {rule_id} 无 target_cell，已跳过")
                continue
            if _is_empty(value):
                warnings.append(f"规则 {rule_id} 无 value，已跳过")
                continue

            try:
                written_cell = safe_write_cell(ws, target_cell, value)
            except Exception as exc:
                warnings.append(f"规则 {rule_id} 写入 {target_cell} 失败：{exc}")
                continue

            if not written_cell:
                warnings.append(f"规则 {rule_id} 写入 cell 为空，已跳过")
                continue

            _style_written_cell(ws, written_cell, value)
            operations_written += 1

        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)

        logger.info(
            "V4 template rule executor succeeded: output_path=%s operations_written=%s warnings=%s",
            output_file,
            operations_written,
            len(warnings),
        )
        return {
            "success": True,
            "output_path": str(output_file),
            "operations_written": operations_written,
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception(
            "V4 template rule executor failed: template_path=%s output_path=%s",
            template_path,
            output_path,
        )
        return {
            "success": False,
            "error": str(exc),
            "operations_written": 0,
            "warnings": warnings,
        }


def execute_rules_to_template_excel_with_preview(
    example_name: str,
    template_key: str,
    template_path: str = "",
) -> dict:
    warnings = []

    try:
        if _is_empty(example_name):
            raise ValueError("example_name must not be empty")
        if _is_empty(template_key):
            raise ValueError("template_key must not be empty")
        if _is_empty(template_path):
            raise ValueError("template_path must not be empty")

        example = load_example(example_name)
        if not example:
            return {
                "success": False,
                "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
            }

        logger.info(
            "V4 template rule export with preview started: example_name=%s template_key=%s template_path=%s",
            example_name,
            template_key,
            template_path,
        )

        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        rules_config = load_excel_render_rules()
        templates = rules_config.get("templates") if isinstance(rules_config, dict) else None
        if not isinstance(templates, dict) or template_key not in templates:
            return {
                "success": False,
                "error": "template_key \u4e0d\u5b58\u5728",
            }

        preview_result = build_excel_rule_preview(
            example,
            render_result.get("description_fields", {}),
            rules_config,
            template_key,
        )
        if not preview_result.get("success"):
            return {
                "success": False,
                "error": preview_result.get("error", "V4 Excel rule preview failed"),
            }

        operations = preview_result.get("operations", [])
        if not isinstance(operations, list):
            operations = []
        warnings.extend(preview_result.get("warnings", []))

        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_example_name = _safe_output_filename_part(example_name)
        safe_template_key = _safe_output_filename_part(template_key)
        filename = f"{safe_example_name}_{safe_template_key}_template_rule_{timestamp}.xlsx"
        output_path = output_dir / filename

        executor_result = execute_rules_to_template_excel(
            template_path,
            operations,
            str(output_path),
        )
        if not executor_result.get("success"):
            return {
                "success": False,
                "error": executor_result.get("error", "V4 template rule executor failed"),
                "operations": operations,
                "warnings": warnings + executor_result.get("warnings", []),
            }

        warnings.extend(executor_result.get("warnings", []))

        logger.info(
            "V4 template rule export with preview succeeded: output_path=%s operations_written=%s warnings=%s",
            output_path,
            executor_result.get("operations_written", 0),
            len(warnings),
        )
        return {
            "success": True,
            "filename": filename,
            "output_path": str(output_path),
            "operations_written": executor_result.get("operations_written", 0),
            "operations": operations,
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception(
            "V4 template rule export with preview failed: example_name=%s template_key=%s template_path=%s",
            example_name,
            template_key,
            template_path,
        )
        return {
            "success": False,
            "error": str(exc),
            "operations_written": 0,
            "operations": [],
            "warnings": warnings,
        }

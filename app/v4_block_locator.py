import json
from json import JSONDecodeError
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


def _get_block_rules_path():
    return get_base_dir() / "v4" / "rules" / "block_rules.json"


def load_block_rules():
    rules_path = _get_block_rules_path()
    try:
        with rules_path.open("r", encoding="utf-8") as f:
            rules = json.load(f)
    except JSONDecodeError as exc:
        logger.error("[BlockLocator] Block rules JSON parse failed: path=%s error=%s", rules_path, exc)
        return {}
    except OSError as exc:
        logger.error("[BlockLocator] Block rules read failed: path=%s error=%s", rules_path, exc)
        return {}

    return rules if isinstance(rules, dict) else {}


def _normalize_block_rules(block_rules):
    rules = block_rules if isinstance(block_rules, dict) else {}
    blocks = rules.get("blocks", [])
    if not isinstance(blocks, list):
        return []

    normalized = []
    for item in blocks:
        if not isinstance(item, dict):
            continue

        block_name = str(item.get("block_name") or "").strip()
        labels = item.get("anchor_labels", [])
        if not isinstance(labels, list):
            labels = []
        anchor_labels = []
        seen = set()
        for label in labels:
            text = str(label or "").strip()
            if text and text not in seen:
                anchor_labels.append(text)
                seen.add(text)

        if block_name and anchor_labels:
            normalized.append({
                "block_name": block_name,
                "anchor_labels": anchor_labels,
            })

    return normalized


def _scan_sheet_labels(ws, anchor_labels):
    anchor_set = set(anchor_labels)
    matches = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            value = str(cell.value).strip()
            if value in anchor_set:
                matches.append({
                    "label": value,
                    "cell": cell.coordinate,
                    "row": cell.row,
                    "col": cell.column,
                })
    return matches


def _select_best_window(matches):
    if len(matches) < 2:
        return []

    sorted_matches = sorted(matches, key=lambda item: (item["row"], item["col"]))
    best_window = []
    best_labels = set()
    for index, start in enumerate(sorted_matches):
        window = [
            item
            for item in sorted_matches[index:]
            if item["row"] - start["row"] <= 5
        ]
        labels = {item["label"] for item in window}
        if len(labels) < 2:
            continue
        if len(labels) > len(best_labels) or (
            len(labels) == len(best_labels)
            and window
            and (not best_window or window[0]["row"] < best_window[0]["row"])
        ):
            best_window = window
            best_labels = labels

    if not best_window:
        return []

    selected = []
    seen = set()
    for item in best_window:
        if item["label"] in seen:
            continue
        selected.append(item)
        seen.add(item["label"])
    return selected


def _range_from_matches(matches):
    min_row = min(item["row"] for item in matches)
    max_row = max(item["row"] for item in matches)
    min_col = min(item["col"] for item in matches)
    max_col = max(item["col"] for item in matches)

    start_row = min_row
    end_row = max_row + 3
    start_col = min_col
    end_col = max_col + 3
    return {
        "start_row": start_row,
        "end_row": end_row,
        "start_col": start_col,
        "end_col": end_col,
        "range": f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
    }


def scan_template_blocks(template_path, block_rules):
    warnings = []
    rules = _normalize_block_rules(block_rules)
    if not rules:
        return {
            "success": False,
            "blocks": [],
            "warnings": ["没有可用的区块规则"],
        }

    template = Path(template_path) if template_path else None
    if not template or not template.is_file():
        return {
            "success": False,
            "blocks": [],
            "warnings": ["请先上传 Excel 模板"],
        }

    try:
        wb = load_workbook(template, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        logger.exception("[BlockLocator] Template block scan failed: path=%s", template)
        return {
            "success": False,
            "blocks": [],
            "warnings": [str(exc) or "模板区块扫描失败"],
        }

    blocks = []
    for rule in rules:
        best_block = None
        for ws in wb.worksheets:
            sheet_matches = _scan_sheet_labels(ws, rule["anchor_labels"])
            selected_matches = _select_best_window(sheet_matches)
            if not selected_matches:
                continue

            block_range = _range_from_matches(selected_matches)
            candidate = {
                "block_name": rule["block_name"],
                "sheet": ws.title,
                "matched_labels": [
                    {
                        "label": item["label"],
                        "cell": item["cell"],
                    }
                    for item in selected_matches
                ],
                **block_range,
            }
            if best_block is None or len(candidate["matched_labels"]) > len(best_block["matched_labels"]):
                best_block = candidate

        if best_block:
            blocks.append(best_block)
        else:
            warnings.append(f"未识别到区块：{rule['block_name']}")

    logger.info(
        "[BlockLocator] Template blocks scanned: rules=%s blocks=%s warnings=%s",
        len(rules),
        len(blocks),
        len(warnings),
    )
    return {
        "success": True,
        "blocks": blocks,
        "warnings": warnings,
    }

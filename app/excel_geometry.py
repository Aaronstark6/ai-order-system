from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.runtime_paths import get_base_dir
from app.template_manager import get_profile


BASE_DIR = get_base_dir()
TEMPLATE_UPLOAD_DIR = BASE_DIR / "templates" / "uploads"
DEFAULT_COLUMN_WIDTH_PX = 64
DEFAULT_ROW_HEIGHT_PX = 24
MIN_COLUMNS = 12
MIN_ROWS = 40
MAX_COLUMNS = 20
MAX_ROWS = 80


def _column_width_to_px(width):
    try:
        number = float(width)
    except (TypeError, ValueError):
        return DEFAULT_COLUMN_WIDTH_PX
    if number <= 0:
        return DEFAULT_COLUMN_WIDTH_PX
    return max(1, int(number * 7 + 5))


def _row_height_to_px(height):
    try:
        number = float(height)
    except (TypeError, ValueError):
        return DEFAULT_ROW_HEIGHT_PX
    if number <= 0:
        return DEFAULT_ROW_HEIGHT_PX
    return max(1, int(number * 96 / 72))


def _error(message):
    return {"success": False, "error": message}


def get_template_geometry(profile_id: str):
    try:
        profile = get_profile(profile_id)
        if not profile:
            return _error("映射不存在")

        template_file = Path(str(profile.get("template_file") or "").strip()).name
        if not template_file:
            return _error("当前映射未上传模板文件")

        template_path = (TEMPLATE_UPLOAD_DIR / template_file).resolve()
        upload_root = TEMPLATE_UPLOAD_DIR.resolve()
        try:
            template_path.relative_to(upload_root)
        except ValueError:
            return _error("模板文件路径非法")

        if not template_path.exists() or not template_path.is_file():
            return _error("模板文件不存在")

        workbook = load_workbook(template_path, read_only=False, data_only=True)
        sheet = workbook.active
        max_column = min(max(MIN_COLUMNS, int(sheet.max_column or MIN_COLUMNS)), MAX_COLUMNS)
        max_row = min(max(MIN_ROWS, int(sheet.max_row or MIN_ROWS)), MAX_ROWS)

        columns = []
        for col_index in range(1, max_column + 1):
            letter = get_column_letter(col_index)
            dimension = sheet.column_dimensions.get(letter)
            width = _column_width_to_px(getattr(dimension, "width", None) if dimension else None)
            columns.append({
                "index": col_index,
                "letter": letter,
                "width": width,
            })

        rows = []
        for row_index in range(1, max_row + 1):
            dimension = sheet.row_dimensions.get(row_index)
            height = _row_height_to_px(getattr(dimension, "height", None) if dimension else None)
            rows.append({
                "index": row_index,
                "height": height,
            })

        merged_cells = []
        for merged_range in sheet.merged_cells.ranges:
            min_col = int(merged_range.min_col)
            min_row = int(merged_range.min_row)
            max_col = int(merged_range.max_col)
            end_row = int(merged_range.max_row)
            if min_col > max_column or min_row > max_row:
                continue
            merged_cells.append({
                "range": str(merged_range),
                "min_col": min_col,
                "min_row": min_row,
                "max_col": max_col,
                "max_row": end_row,
            })

        workbook.close()
        return {
            "success": True,
            "sheet_name": sheet.title,
            "columns": columns,
            "rows": rows,
            "merged_cells": merged_cells,
            "max_column": max_column,
            "max_row": max_row,
        }
    except Exception as e:
        return _error(f"读取模板几何失败：{e}")

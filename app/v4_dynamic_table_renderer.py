import json
from json import JSONDecodeError

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


TABLE_NAME_BY_KEY = {
    "formula_items": "配方表",
    "test_items": "检测项目表",
    "packaging_items": "包装明细表",
}


def _get_dynamic_table_mapping_path():
    return get_base_dir() / "v4" / "rules" / "dynamic_table_mapping.json"


def load_dynamic_table_mapping():
    mapping_path = _get_dynamic_table_mapping_path()
    try:
        with mapping_path.open("r", encoding="utf-8") as f:
            mapping = json.load(f)
    except JSONDecodeError as exc:
        logger.error("[DynamicTableRenderer] Mapping JSON parse failed: path=%s error=%s", mapping_path, exc)
        return {}
    except OSError as exc:
        logger.error("[DynamicTableRenderer] Mapping read failed: path=%s error=%s", mapping_path, exc)
        return {}

    return mapping if isinstance(mapping, dict) else {}


def _get_product_tables(order_object):
    product = order_object.get("product", {}) if isinstance(order_object, dict) else {}
    if not isinstance(product, dict):
        return {}
    tables = product.get("tables", {})
    return tables if isinstance(tables, dict) else {}


def _get_table_config(mapping, table_key):
    config = mapping.get(table_key, {}) if isinstance(mapping, dict) else {}
    return config if isinstance(config, dict) else {}


def _parse_start_cell(start_cell):
    try:
        row, col = coordinate_to_tuple(str(start_cell or "").strip())
        return row, col
    except ValueError:
        return None, None


def _table_display_name(table_key, config):
    name = str(config.get("table_name") or "").strip()
    return name or TABLE_NAME_BY_KEY.get(table_key, table_key)


def _extract_columns(rows):
    if not rows or not isinstance(rows[0], dict):
        return []
    return [str(key) for key in rows[0].keys()]


def order_object_to_dynamic_table_operations(order_object, mapping):
    warnings = []
    operations = []
    table_summaries = []

    if not isinstance(order_object, dict):
        return {
            "success": False,
            "operations": [],
            "tables": [],
            "warnings": ["Order Object 必须是对象"],
        }

    product_tables = _get_product_tables(order_object)
    if not product_tables:
        return {
            "success": True,
            "operations": [],
            "tables": [],
            "warnings": [],
        }

    for table_key, rows in product_tables.items():
        table_key = str(table_key or "").strip()
        config = _get_table_config(mapping, table_key)
        table_name = _table_display_name(table_key, config)
        start_cell = str(config.get("start_cell") or "").strip().upper()
        start_row, start_col = _parse_start_cell(start_cell)

        if start_row is None or start_col is None:
            warnings.append(f"表格缺少有效 start_cell：{table_key}")
            continue
        if not isinstance(rows, list):
            warnings.append(f"表格数据不是数组：product.tables.{table_key}")
            continue
        if not rows:
            warnings.append(f"表格为空：{table_key}")
            continue

        columns = _extract_columns(rows)
        if not columns:
            warnings.append(f"表格首行无可用字段：{table_key}")
            continue

        valid_row_count = 0
        for row_index, row_data in enumerate(rows):
            if not isinstance(row_data, dict):
                warnings.append(f"{table_name} 第 {row_index + 1} 行不是对象，已跳过。")
                continue

            excel_row = start_row + row_index
            valid_row_count += 1
            for column_index, field in enumerate(columns):
                target_col = get_column_letter(start_col + column_index)
                value = row_data.get(field, "")
                operations.append({
                    "operation": "write_text",
                    "target_cell": f"{target_col}{excel_row}",
                    "value": "" if value is None else str(value),
                    "table_key": table_key,
                    "table_name": table_name,
                    "row_index": row_index,
                    "field": field,
                })

        if valid_row_count:
            end_col = get_column_letter(start_col + len(columns) - 1)
            end_row = start_row + len(rows) - 1
            table_summaries.append({
                "table_key": table_key,
                "table_name": table_name,
                "start_cell": start_cell,
                "range": f"{get_column_letter(start_col)}{start_row}:{end_col}{end_row}",
                "rows_count": len(rows),
                "columns_count": len(columns),
                "fields": columns,
            })

    logger.info(
        "[DynamicTableRenderer] Operations generated: tables=%s operations=%s warnings=%s",
        len(table_summaries),
        len(operations),
        len(warnings),
    )
    return {
        "success": True,
        "operations": operations,
        "tables": table_summaries,
        "warnings": warnings,
    }

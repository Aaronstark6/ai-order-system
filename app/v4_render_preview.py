from copy import deepcopy
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


def _as_list(value):
    return value if isinstance(value, list) else []


def _text(value):
    return "" if value is None else str(value)


def _parse_cell(cell_ref):
    match = re.fullmatch(r"\s*([A-Za-z]+)\s*(\d+)\s*", str(cell_ref or ""))
    if not match:
        return "", None
    return match.group(1).upper(), int(match.group(2))


def _load_merged_ranges(template_path):
    if not template_path:
        return []
    try:
        path = Path(template_path)
        if not path.is_file():
            return []
        workbook = load_workbook(path, read_only=False, data_only=False)
        ranges = []
        for worksheet in workbook.worksheets:
            for merged_range in worksheet.merged_cells.ranges:
                ranges.append(
                    {
                        "sheet": worksheet.title,
                        "range": str(merged_range),
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col,
                    }
                )
        workbook.close()
        return ranges
    except Exception:
        return []


def _merged_display_info(cell_ref, merged_ranges):
    column, row = _parse_cell(cell_ref)
    if not column or row is None:
        return {
            "display_cell": cell_ref,
            "merged_range": "",
            "display_note": "",
        }

    try:
        cell_row, cell_col = coordinate_to_tuple(cell_ref)
    except Exception:
        return {
            "display_cell": cell_ref,
            "merged_range": "",
            "display_note": "",
        }

    for merged_range in merged_ranges:
        if (
            merged_range["min_row"] <= cell_row <= merged_range["max_row"]
            and merged_range["min_col"] <= cell_col <= merged_range["max_col"]
        ):
            display_cell = f"{column_from_index(merged_range['min_col'])}{merged_range['min_row']}"
            range_text = merged_range["range"]
            if display_cell == cell_ref:
                note = f"{cell_ref} 位于合并区域 {range_text} 的左上角"
            else:
                note = f"{cell_ref} 显示于 {display_cell}（合并区域 {range_text}）"
            return {
                "display_cell": display_cell,
                "merged_range": range_text,
                "display_note": note,
            }

    return {
        "display_cell": cell_ref,
        "merged_range": "",
        "display_note": "",
    }


def column_from_index(index):
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _normalize_operation(operation):
    if not isinstance(operation, dict):
        return None
    item = deepcopy(operation)
    item["source"] = str(item.get("source") or item.get("type") or "").strip().lower()
    item["op_type"] = str(item.get("op_type") or item.get("operation") or "").strip().lower()
    item["target_cell"] = str(item.get("target_cell") or item.get("start_cell") or "").strip().upper()
    item["value"] = _text(item.get("value", item.get("content", "")))
    return item


def _table_name(operation):
    return str(operation.get("table_name") or "未命名表格").strip() or "未命名表格"


def _block_name(operation):
    return str(operation.get("block_name") or "未命名区块").strip() or "未命名区块"


def _status(item, conflict_cells, overwrite_cells):
    target_cell = str(item.get("target_cell") or "").strip().upper()
    skipped = bool(item.get("skipped"))
    return {
        "skipped": skipped,
        "conflict": bool(item.get("conflict")) or target_cell in conflict_cells,
        "overwrite_warning": bool(item.get("overwrite_warning")) or target_cell in overwrite_cells,
        "skip_reason": str(item.get("skip_reason") or ""),
        "safety_status": str(item.get("safety_status") or ("skipped" if skipped else "ok")),
    }


def _conflict_cells(mapping_safety):
    cells = set()
    for item in _as_list((mapping_safety or {}).get("conflicts")):
        if isinstance(item, dict):
            cell = str(item.get("target_cell") or "").strip().upper()
            if cell:
                cells.add(cell)
    return cells


def _overwrite_cells(mapping_safety):
    cells = set()
    for item in _as_list((mapping_safety or {}).get("overwrite_warnings")):
        if isinstance(item, dict):
            cell = str(item.get("target_cell") or item.get("cell") or "").strip().upper()
            if cell:
                cells.add(cell)
    return cells


def build_render_preview(processed_operations, mapping_safety=None, template_path=None):
    mapping_safety = mapping_safety if isinstance(mapping_safety, dict) else {}
    conflict_cells = _conflict_cells(mapping_safety)
    overwrite_cells = _overwrite_cells(mapping_safety)
    warnings = []
    cells = []
    cell_preview = []
    table_groups = {}
    block_preview = []
    merged_ranges = _load_merged_ranges(template_path)

    for index, operation in enumerate(_as_list(processed_operations), start=1):
        item = _normalize_operation(operation)
        if not item:
            warnings.append(f"第 {index} 条 operation 格式无效，已跳过。")
            continue

        source = item.get("source")
        target_cell = item.get("target_cell")
        value = item.get("value")
        status = _status(item, conflict_cells, overwrite_cells)
        display_info = _merged_display_info(target_cell, merged_ranges)
        if target_cell:
            cells.append(
                {
                    "cell": target_cell,
                    "display_cell": display_info.get("display_cell", target_cell),
                    "merged_range": display_info.get("merged_range", ""),
                    "display_note": display_info.get("display_note", ""),
                    "value": value,
                    "operation_type": item.get("op_type"),
                    "source": source,
                    **status,
                }
            )

        if source == "structured":
            cell_preview.append(
                {
                    "cell": target_cell,
                    "display_cell": display_info.get("display_cell", target_cell),
                    "merged_range": display_info.get("merged_range", ""),
                    "display_note": display_info.get("display_note", ""),
                    "source": source,
                    "op_type": item.get("op_type"),
                    "value": value,
                    **status,
                }
            )
            continue

        if source == "table":
            column, row_number = _parse_cell(target_cell)
            if not column or row_number is None:
                warnings.append(f"第 {index} 条 table operation 缺少有效 target_cell，已跳过。")
                continue
            table = table_groups.setdefault(_table_name(item), {})
            row = table.setdefault(row_number, {})
            row[column] = {
                "value": value,
                "display_cell": display_info.get("display_cell", target_cell),
                "merged_range": display_info.get("merged_range", ""),
                "display_note": display_info.get("display_note", ""),
                **status,
            }
            continue

        if source == "block":
            block_preview.append(
                {
                    "block_name": _block_name(item),
                    "target_cell": target_cell,
                    "display_cell": display_info.get("display_cell", target_cell),
                    "merged_range": display_info.get("merged_range", ""),
                    "display_note": display_info.get("display_note", ""),
                    "value": value,
                    **status,
                }
            )
            continue

    skipped_preview = []
    for skipped in _as_list(mapping_safety.get("skipped_operations")):
        if not isinstance(skipped, dict):
            continue
        item = _normalize_operation(skipped) or deepcopy(skipped)
        skipped_item = {
            "cell": str(item.get("target_cell") or item.get("start_cell") or "").strip().upper(),
            "value": _text(item.get("value", item.get("content", ""))),
            "operation_type": str(item.get("op_type") or item.get("operation") or ""),
            "source": str(item.get("source") or item.get("type") or ""),
            "skipped": True,
            "conflict": bool(item.get("conflict")),
            "overwrite_warning": bool(item.get("overwrite_warning")),
            "skip_reason": str(item.get("skip_reason") or ""),
            "safety_status": "skipped",
        }
        skipped_preview.append(skipped_item)
        cells.append(skipped_item)

    table_preview = []
    for table_name in sorted(table_groups):
        rows = []
        for row_number in sorted(table_groups[table_name]):
            rows.append(
                {
                    "row_number": row_number,
                    "cells": dict(sorted(table_groups[table_name][row_number].items())),
                }
            )
        table_preview.append(
            {
                "table_name": table_name,
                "rows": rows,
            }
        )

    return {
        "success": True,
        "cells": cells,
        "cell_preview": cell_preview,
        "table_preview": table_preview,
        "block_preview": block_preview,
        "skipped_preview": skipped_preview,
        "mapping_safety": {
            "conflicts": deepcopy(mapping_safety.get("conflicts", []))
            if isinstance(mapping_safety.get("conflicts", []), list)
            else [],
            "warnings": deepcopy(mapping_safety.get("warnings", []))
            if isinstance(mapping_safety.get("warnings", []), list)
            else [],
            "skipped_operations": deepcopy(mapping_safety.get("skipped_operations", []))
            if isinstance(mapping_safety.get("skipped_operations", []), list)
            else [],
        },
        "warnings": warnings,
    }

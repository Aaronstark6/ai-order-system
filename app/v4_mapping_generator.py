from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string


STRUCTURED_SEMANTICS = {"customer_info", "document_info", "product_detail"}
TABLE_SEMANTICS = {"formula", "packaging", "production_batch", "label_requirement", "order_items", "product_detail", "table"}

EXACT_SOURCE_PATHS = {
    "客户名称": "customer.name",
    "客户": "customer.name",
    "客户公司": "customer.name",
    "公司名称": "customer.name",
    "客户性质": "customer.type",
    "国家": "customer.country",
    "出口国家": "customer.country",
    "目的国": "customer.country",
    "联系人": "customer.contact_person",
    "联系人姓名": "customer.contact_person",
    "数量": "order.quantity",
    "订购数量": "order.quantity",
    "采购数量": "order.quantity",
    "负责人": "order.owner",
    "日期": "order.order_date",
    "文档编号": "document.doc_no",
    "产品名称": "product.product_name",
    "品名": "product.product_name",
    "产品": "product.product_name",
    "产品类型": "product.product_type",
    "规格": "product.spec",
    "产品规格": "product.spec",
    "净含量": "product.spec",
    "包装": "product.packaging",
    "包装方式": "product.packaging",
    "单价": "order.unit_price",
    "价格": "order.unit_price",
    "交期": "order.delivery_time",
    "交货期": "order.delivery_time",
    "Logo": "document.logo",
    "LOGO": "document.logo",
    "图片": "document.image",
    "产品图片": "product.image",
}

AMBIGUOUS_REJECT_LABELS = {
    "客户设计制作",
    "我司设计制作",
    "不贴标签",
}


def _empty_mapping():
    return {
        "structured": [],
        "tables": [],
        "blocks": [],
        "needs_review": [],
        "rejected_candidates": [],
    }


def _clean_label(value):
    text = str(value or "").strip()
    while text and (text.endswith(":") or text.endswith("：")):
        text = text.rstrip(":：").rstrip()
    return text


def _line_count(value):
    text = str(value or "")
    return len(text.splitlines()) if text else 0


def _is_long_label(value):
    text = str(value or "").strip()
    return len(text) > 40 or _line_count(text) > 2


def _cell_parts(cell):
    try:
        column_letter, row_number = coordinate_from_string(str(cell or "").strip())
        return column_index_from_string(column_letter), int(row_number)
    except Exception:
        return None, None


def _right_cell(cell):
    col_index, row_number = _cell_parts(cell)
    if not col_index or not row_number:
        return ""
    return f"{get_column_letter(col_index + 1)}{row_number}"


def _target_col_from_number(col_number):
    try:
        return get_column_letter(int(col_number))
    except Exception:
        return ""


def _in_bounds(label, bounds):
    if not isinstance(label, dict) or not isinstance(bounds, dict):
        return False

    col_index, row_number = _cell_parts(label.get("cell"))
    if not col_index or not row_number:
        return False

    try:
        start_row = int(bounds.get("start_row"))
        end_row = int(bounds.get("end_row"))
        start_col = int(bounds.get("start_col"))
        end_col = int(bounds.get("end_col"))
    except (TypeError, ValueError):
        return False

    return start_row <= row_number <= end_row and start_col <= col_index <= end_col


def infer_source_path(label):
    result = classify_label_mapping(label)
    return result.get("source_path", "")


def classify_label_mapping(label):
    text = _clean_label(label)
    if not text:
        return {
            "source_path": "",
            "confidence": "low",
            "status": "rejected",
            "reason": "空标签",
        }

    if _is_long_label(text):
        return {
            "source_path": "",
            "confidence": "low",
            "status": "rejected",
            "reason": "标签过长或包含过多换行，疑似说明文本",
        }

    if text in EXACT_SOURCE_PATHS:
        return {
            "source_path": EXACT_SOURCE_PATHS[text],
            "confidence": "high",
            "status": "recommended",
            "reason": "核心字段精确匹配",
        }

    if text in AMBIGUOUS_REJECT_LABELS:
        return {
            "source_path": "",
            "confidence": "low",
            "status": "rejected",
            "reason": "标签语义属于制作/标签要求，不是结构化订单字段",
        }

    if "客户" in text:
        return {
            "source_path": "customer.name",
            "confidence": "low",
            "status": "needs_review",
            "reason": "仅包含客户关键词，需人工确认具体字段",
        }

    if "产品" in text:
        return {
            "source_path": f"product.fields.{text}",
            "confidence": "low",
            "status": "needs_review",
            "reason": "仅包含产品关键词，需人工确认字段含义",
        }

    if "数量" in text:
        return {
            "source_path": "order.quantity",
            "confidence": "medium",
            "status": "needs_review",
            "reason": "包含数量关键词，但不是精确字段",
        }

    return {
        "source_path": "",
        "confidence": "low",
        "status": "rejected",
        "reason": "未匹配到可信结构化字段",
    }


def _candidate_payload(label, classification, source_region="", semantic_type=""):
    clean_label = _clean_label(label.get("value") if isinstance(label, dict) else label)
    cell = label.get("cell", "") if isinstance(label, dict) else ""
    sheet = label.get("sheet", "") if isinstance(label, dict) else ""
    return {
        "label": clean_label,
        "source_path": classification.get("source_path", ""),
        "target_cell": _right_cell(cell),
        "operation": "write_text",
        "sheet": sheet,
        "semantic_type": semantic_type,
        "source_region": source_region,
        "confidence": classification.get("confidence", "low"),
        "status": classification.get("status", "rejected"),
        "reason": classification.get("reason", ""),
        "auto_generated": True,
    }


def _labels_for_region(labels, region):
    bounds = region.get("bounds", {}) if isinstance(region, dict) else {}
    return [label for label in labels if _in_bounds(label, bounds)]


def _add_structured_candidate(label, region, result, seen):
    clean_label = _clean_label(label.get("value"))
    key = (label.get("sheet", ""), clean_label, label.get("cell", ""))
    if key in seen:
        return
    seen.add(key)

    classification = classify_label_mapping(clean_label)
    payload = _candidate_payload(
        label,
        classification,
        source_region=region.get("suggested_name") or region.get("name", "") if region else "Labels Preview",
        semantic_type=region.get("semantic_type", "inferred_label") if region else "inferred_label",
    )

    if payload["status"] == "recommended":
        result["structured"].append(payload)
    elif payload["status"] == "needs_review":
        result["needs_review"].append(payload)
    else:
        result["rejected_candidates"].append(payload)


def _build_structured_mapping(labels, regions, result):
    seen = set()
    for region in regions:
        if region.get("semantic_type") not in STRUCTURED_SEMANTICS:
            continue
        for label in _labels_for_region(labels, region):
            _add_structured_candidate(label, region, result, seen)

    for label in labels:
        _add_structured_candidate(label, None, result, seen)


def _region_candidate(region, reason, status="rejected"):
    return {
        "label": region.get("suggested_name") or region.get("name", ""),
        "source_path": "",
        "target_cell": "",
        "confidence": region.get("confidence", "low"),
        "status": status,
        "reason": reason,
        "semantic_type": region.get("semantic_type", ""),
        "type": region.get("type", ""),
        "sheet": region.get("sheet", ""),
        "auto_generated": True,
    }


def _build_table_mapping(regions, result):
    for region in regions:
        if region.get("type") != "table":
            continue
        if region.get("semantic_type") not in TABLE_SEMANTICS:
            result["rejected_candidates"].append(
                _region_candidate(region, "表格语义不在允许自动生成范围")
            )
            continue
        if region.get("confidence") == "low":
            result["needs_review"].append(
                _region_candidate(region, "表格区域置信度较低，需人工确认", status="needs_review")
            )
            continue

        bounds = region.get("bounds", {}) if isinstance(region.get("bounds", {}), dict) else {}
        start_col = bounds.get("start_col")
        label_names = region.get("label_names", [])
        if not isinstance(label_names, list):
            label_names = []

        try:
            header_row = int(bounds.get("start_row"))
            data_start_row = header_row + 1
        except (TypeError, ValueError):
            data_start_row = bounds.get("start_row")

        columns = []
        for offset, label in enumerate(label_names):
            columns.append(
                {
                    "label": _clean_label(label),
                    "target_col": _target_col_from_number((start_col or 1) + offset),
                }
            )

        result["tables"].append(
            {
                "table_name": region.get("suggested_name") or region.get("name", ""),
                "start_row": data_start_row,
                "sheet": region.get("sheet", ""),
                "columns": columns,
                "semantic_type": region.get("semantic_type", ""),
                "confidence": region.get("confidence", "medium"),
                "status": "recommended",
                "reason": "语义和置信度满足自动生成 table mapping 条件",
                "auto_generated": True,
            }
        )


def _build_block_mapping(regions, result):
    for region in regions:
        if region.get("type") != "block":
            continue
        if region.get("semantic_type") == "unknown":
            result["rejected_candidates"].append(
                _region_candidate(region, "未知语义 block 不进入推荐 block mapping")
            )
            continue

        bounds = region.get("bounds", {}) if isinstance(region.get("bounds", {}), dict) else {}
        target_col = _target_col_from_number(bounds.get("start_col"))
        target_cell = f"{target_col}{bounds.get('start_row')}" if target_col and bounds.get("start_row") else ""
        result["blocks"].append(
            {
                "block_name": region.get("suggested_name") or region.get("name", ""),
                "target_cell": target_cell,
                "operation": "write_block",
                "sheet": region.get("sheet", ""),
                "semantic_type": region.get("semantic_type", ""),
                "confidence": region.get("confidence", "medium"),
                "status": "recommended",
                "reason": "明确语义 block，可生成 write_block 预览",
                "auto_generated": True,
            }
        )


def generate_auto_mapping(template_analysis):
    analysis = template_analysis if isinstance(template_analysis, dict) else {}
    labels = analysis.get("labels", [])
    if not isinstance(labels, list):
        labels = []

    template_structure = analysis.get("template_structure", {})
    if not isinstance(template_structure, dict):
        template_structure = {}
    regions = template_structure.get("recommended_regions", [])
    if not isinstance(regions, list):
        regions = []

    result = _empty_mapping()
    _build_structured_mapping(labels, regions, result)
    _build_table_mapping(regions, result)
    _build_block_mapping(regions, result)
    return result

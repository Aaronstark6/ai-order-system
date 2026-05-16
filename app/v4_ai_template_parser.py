import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from app.logger import get_logger


logger = get_logger(__name__)


DEFAULT_TARGET_CELLS = ("B10", "B14", "B18", "B22", "B26", "B30")


def _normalize_text(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "", text)
    return re.sub(r"[:：/／\\|_\-（）()【】\[\]{}]+", "", text)


def _cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _field_entries(field_library):
    entries = []

    description_fields = []
    if isinstance(field_library, dict):
        raw_description_fields = field_library.get("description_fields")
        if isinstance(raw_description_fields, dict):
            description_fields = list(raw_description_fields.keys())
        elif isinstance(raw_description_fields, list):
            description_fields = raw_description_fields

    for key in description_fields:
        key_text = str(key or "").strip()
        if not key_text:
            continue
        entries.append({
            "key": key_text,
            "label": key_text,
            "source": "description_field",
            "terms": [_normalize_text(key_text)],
        })

    raw_fields = field_library.get("fields") if isinstance(field_library, dict) else None
    if isinstance(raw_fields, dict):
        raw_fields = list(raw_fields.values())
    if isinstance(raw_fields, list):
        for field in raw_fields:
            if not isinstance(field, dict):
                continue
            key = str(field.get("key") or "").strip()
            label = str(field.get("label") or key).strip()
            description = str(field.get("description") or "").strip()
            terms = [_normalize_text(part) for part in [key, label, description] if str(part or "").strip()]
            entries.append({
                "key": key,
                "label": label,
                "source": "field_library",
                "terms": terms,
            })

    return entries


def _iter_text_cells(ws):
    merged_anchors = {}
    for merged_range in ws.merged_cells.ranges:
        merged_anchors[str(merged_range.start_cell.coordinate)] = str(merged_range)

    max_row = min(ws.max_row or 1, 200)
    max_column = min(ws.max_column or 1, 80)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            text = _cell_text(cell.value)
            if not text:
                continue
            yield {
                "coordinate": cell.coordinate,
                "row": cell.row,
                "column": cell.column,
                "text": text,
                "normalized_text": _normalize_text(text),
                "merged_range": merged_anchors.get(cell.coordinate, ""),
            }


def _find_target_cell(ws, row, column):
    max_column = min(ws.max_column or 1, 80)

    for offset in range(1, 6):
        target_column = column + offset
        if target_column > max_column:
            break
        cell = ws.cell(row=row, column=target_column)
        if _cell_text(cell.value) == "":
            return cell.coordinate

    for offset in range(1, 4):
        target_row = row + offset
        if target_row > (ws.max_row or row):
            break
        cell = ws.cell(row=target_row, column=column)
        if _cell_text(cell.value) == "":
            return cell.coordinate

    return f"{get_column_letter(column + 1)}{row}"


def _match_entry(cell, entries):
    text = cell.get("normalized_text", "")
    if not text:
        return None

    best = None
    best_score = 0
    for entry in entries:
        for term in entry.get("terms", []):
            if not term:
                continue
            if term == text:
                score = 100 + len(term)
            elif term in text or text in term:
                score = 60 + min(len(term), len(text))
            else:
                score = 0
            if score > best_score:
                best_score = score
                best = entry

    return best if best_score >= 62 else None


def _build_text_rule(entry, target_cell, index):
    source_key = entry.get("key") or f"field_{index}"
    return {
        "id": f"ai_{source_key}_{index}",
        "type": "text",
        "source": {
            "description_field": source_key,
        },
        "target": {
            "sheet": "active",
            "cell": target_cell,
        },
    }


def _build_checkbox_rule(cell, index):
    return {
        "id": f"ai_checkbox_{index}",
        "type": "checkbox",
        "condition": {
            "path": "product.form",
            "equals": "soft_capsule",
        },
        "target": {
            "sheet": "active",
            "cell": cell["coordinate"],
        },
        "checked_text": "☑",
        "unchecked_text": "☐",
    }


def _fallback_rules(entries, warnings):
    description_entries = [entry for entry in entries if entry.get("source") == "description_field"]
    if not description_entries:
        warnings.append("未找到可用于生成规则的 V4 description_fields")
        return []

    warnings.append("模板文本未匹配到明确字段，已按 description_fields 顺序生成回退规则")
    rules = []
    for index, entry in enumerate(description_entries[:len(DEFAULT_TARGET_CELLS)]):
        rules.append(_build_text_rule(entry, DEFAULT_TARGET_CELLS[index], index + 1))
    return rules


def analyze_template(template_path: str, field_library: dict) -> dict:
    warnings = []

    try:
        template_file = Path(str(template_path or "")).expanduser()
        if not template_file.is_file():
            return {
                "success": False,
                "error": f"template file not found: {template_path}",
            }

        entries = _field_entries(field_library if isinstance(field_library, dict) else {})
        if not entries:
            warnings.append("字段库为空，无法进行字段匹配")

        wb = load_workbook(template_file, data_only=False)
        ws = wb.active
        rules = []
        matched_keys = set()

        for cell in _iter_text_cells(ws):
            text = cell["text"]
            if "☐" in text or "☑" in text or "□" in text:
                rules.append(_build_checkbox_rule(cell, len(rules) + 1))
                continue

            entry = _match_entry(cell, entries)
            if not entry:
                continue

            if entry["source"] != "description_field":
                warnings.append(f"字段 {entry.get('label') or entry.get('key')} 暂未映射到 V4 description_fields，已跳过")
                continue

            if entry["key"] in matched_keys:
                continue

            matched_keys.add(entry["key"])
            target_cell = _find_target_cell(ws, cell["row"], cell["column"])
            rules.append(_build_text_rule(entry, target_cell, len(rules) + 1))

        if not rules:
            rules = _fallback_rules(entries, warnings)

        template_key = f"ai_{template_file.stem}_template"
        rules_config = {
            "version": "v4.15-ai-template",
            "description": "AI template parser generated Excel render rules",
            "templates": {
                template_key: {
                    "label": f"AI 自动解析模板：{template_file.name}",
                    "rules": rules,
                }
            },
        }

        logger.info(
            "V4 AI template parser finished: template_path=%s template_key=%s rules=%s warnings=%s",
            template_file,
            template_key,
            len(rules),
            len(warnings),
        )
        return {
            "success": True,
            "template_key": template_key,
            "rules_config": rules_config,
            "rules_count": len(rules),
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception("V4 AI template parser failed: template_path=%s", template_path)
        return {
            "success": False,
            "error": str(exc),
            "warnings": warnings,
        }


def _unique_template_key(template_key, used_keys):
    base_key = str(template_key or "ai_template").strip() or "ai_template"
    if base_key not in used_keys:
        used_keys.add(base_key)
        return base_key

    index = 2
    while f"{base_key}_{index}" in used_keys:
        index += 1

    unique_key = f"{base_key}_{index}"
    used_keys.add(unique_key)
    return unique_key


def analyze_templates(template_paths: list, field_library: dict) -> dict:
    warnings = []
    parsed_templates = []
    used_keys = set()
    rules_config = {
        "version": "v4.16-ai-batch-template",
        "description": "AI batch template parser generated Excel render rules",
        "templates": {},
    }

    if not isinstance(template_paths, list) or not template_paths:
        return {
            "success": False,
            "error": "template_paths 不能为空",
            "rules_config": rules_config,
            "templates": [],
            "warnings": warnings,
        }

    for template_path in template_paths:
        template_label = Path(str(template_path or "")).name or str(template_path or "")
        result = analyze_template(str(template_path or ""), field_library)
        if not result.get("success"):
            warning = f"模板 {template_label} 解析失败：{result.get('error', '未知错误')}"
            warnings.append(warning)
            parsed_templates.append({
                "success": False,
                "template_path": str(template_path or ""),
                "template_key": "",
                "rules_count": 0,
                "warnings": result.get("warnings", []),
                "error": result.get("error", "未知错误"),
            })
            continue

        source_config = result.get("rules_config", {})
        source_templates = source_config.get("templates") if isinstance(source_config, dict) else {}
        source_key = result.get("template_key", "")
        source_template = source_templates.get(source_key, {}) if isinstance(source_templates, dict) else {}
        template_key = _unique_template_key(source_key, used_keys)
        if template_key != source_key:
            warnings.append(f"模板 {template_label} 的规则键重复，已自动调整为 {template_key}")

        rules_config["templates"][template_key] = source_template
        template_warnings = result.get("warnings", [])
        warnings.extend([f"{template_label}：{warning}" for warning in template_warnings])
        parsed_templates.append({
            "success": True,
            "template_path": str(template_path or ""),
            "template_key": template_key,
            "rules_count": result.get("rules_count", 0),
            "warnings": template_warnings,
        })

    return {
        "success": bool(rules_config["templates"]),
        "rules_config": rules_config,
        "templates": parsed_templates,
        "warnings": warnings,
    }


def _localized_rule(rule):
    if not isinstance(rule, dict):
        return {}

    rule_type = rule.get("type") or ""
    target = rule.get("target") if isinstance(rule.get("target"), dict) else {}
    source = rule.get("source") if isinstance(rule.get("source"), dict) else {}
    condition = rule.get("condition") if isinstance(rule.get("condition"), dict) else {}

    return {
        "规则 ID": rule.get("id", ""),
        "类型": "勾选框" if rule_type == "checkbox" else "文本" if rule_type == "text" else rule_type,
        "目标单元格": target.get("cell", ""),
        "来源字段": source.get("description_field", ""),
        "条件": condition,
        "勾选状态文本": rule.get("checked_text", ""),
        "未勾状态文本": rule.get("unchecked_text", ""),
    }


def 分析模板(template_path: str, 字段库: dict) -> dict:
    result = analyze_template(template_path, 字段库)
    rules_config = result.get("rules_config", {}) if isinstance(result, dict) else {}
    template_key = result.get("template_key", "") if isinstance(result, dict) else ""
    template = {}
    if isinstance(rules_config, dict):
        templates = rules_config.get("templates")
        if isinstance(templates, dict):
            template = templates.get(template_key, {}) if template_key else {}
    rules = template.get("rules", []) if isinstance(template, dict) else []

    localized = {
        "成功": bool(result.get("success")) if isinstance(result, dict) else False,
        "模板键": template_key,
        "规则配置": rules_config,
        "规则数量": result.get("rules_count", len(rules)) if isinstance(result, dict) else 0,
        "规则列表": [_localized_rule(rule) for rule in rules if isinstance(rule, dict)],
        "警告": result.get("warnings", []) if isinstance(result, dict) else [],
    }
    if isinstance(result, dict) and result.get("error"):
        localized["错误"] = result.get("error")

    return localized


def 批量分析模板(template_paths: list, 字段库: dict) -> dict:
    result = analyze_templates(template_paths, 字段库)
    return {
        "成功": bool(result.get("success")) if isinstance(result, dict) else False,
        "规则配置": result.get("rules_config", {}) if isinstance(result, dict) else {},
        "模板列表": result.get("templates", []) if isinstance(result, dict) else [],
        "警告": result.get("warnings", []) if isinstance(result, dict) else [],
        "错误": result.get("error", "") if isinstance(result, dict) else "",
    }

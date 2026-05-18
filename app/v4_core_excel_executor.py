from copy import copy
from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


def _safe_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r"[^\w.-]+", "_", text, flags=re.UNICODE)
    return text.strip("._") or "template"


def _resolve_merged_target(ws, cell_ref, warnings):
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            start_cell = merged_range.start_cell.coordinate
            if start_cell != cell_ref:
                warnings.append(f"{cell_ref} 位于合并单元格内，已写入左上角 {start_cell}。")
            return start_cell
    return cell_ref


def _set_wrap_text(cell):
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.vertical = alignment.vertical or "top"
    cell.alignment = alignment


def _validate_template_and_operations(template_path, operations):
    if not template_path:
        return None, {
            "success": False,
            "error": "尚未上传模板",
            "operations_count": 0,
            "warnings": [],
        }

    template = Path(template_path)
    if not template.is_file():
        return None, {
            "success": False,
            "error": "尚未上传模板",
            "operations_count": 0,
            "warnings": [],
        }

    if not isinstance(operations, list):
        return None, {
            "success": False,
            "error": "尚未生成 operations",
            "operations_count": 0,
            "warnings": [],
        }

    return template, None


def _load_workbook(template, warnings):
    try:
        return load_workbook(template), None
    except (InvalidFileException, OSError, ValueError) as exc:
        logger.exception("[CoreExcelExecutor] Template load failed: path=%s", template)
        return None, {
            "success": False,
            "error": "Excel 写入失败",
            "operations_count": 0,
            "warnings": warnings + [str(exc)],
        }


def _write_value(ws, target_cell, value, warnings):
    write_cell_ref = _resolve_merged_target(ws, target_cell, warnings)
    cell = ws[write_cell_ref]
    cell.value = "" if value is None else str(value)
    _set_wrap_text(cell)


def _save_workbook(wb, template, operations_written, warnings, log_context):
    output_dir = get_base_dir() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    template_name = _safe_filename_part(template.stem)
    filename = f"v4_core_real_excel_{template_name}_{timestamp}.xlsx"
    output_path = output_dir / filename
    wb.save(output_path)

    logger.info(
        "[CoreExcelExecutor] Real Excel generated: context=%s path=%s operations=%s warnings=%s",
        log_context,
        output_path,
        operations_written,
        len(warnings),
    )
    return {
        "success": True,
        "filename": filename,
        "download_url": f"/api/download/{filename}",
        "operations_count": operations_written,
        "warnings": warnings,
        "output_path": str(output_path),
    }


def execute_unified_operations(template_path, operations):
    warnings = []
    operations_written = 0
    template, error_result = _validate_template_and_operations(template_path, operations)
    if error_result:
        return error_result

    wb, error_result = _load_workbook(template, warnings)
    if error_result:
        return error_result

    try:
        ws = wb.active
        for index, operation in enumerate(operations, start=1):
            if not isinstance(operation, dict):
                warnings.append(f"第 {index} 条 unified operation 无效，已跳过。")
                continue

            op_type = str(operation.get("op_type") or "").strip()
            if op_type not in {"write_text", "write_table_cell", "write_block"}:
                warnings.append(f"暂不支持 op_type={op_type or '空'}，已跳过。")
                continue

            target_cell = str(operation.get("target_cell") or "").strip()
            if not target_cell:
                warnings.append(f"第 {index} 条 unified operation 缺少 target_cell，已跳过。")
                continue

            try:
                _write_value(ws, target_cell, operation.get("value"), warnings)
                operations_written += 1
            except Exception as exc:
                logger.warning(
                    "[CoreExcelExecutor] Unified operation failed: index=%s target=%s error=%s",
                    index,
                    target_cell,
                    exc,
                    exc_info=True,
                )
                warnings.append(f"{target_cell} 写入失败：{exc}")

        return _save_workbook(wb, template, operations_written, warnings, "unified")
    except Exception as exc:
        logger.exception("[CoreExcelExecutor] Excel save failed")
        return {
            "success": False,
            "error": "无法保存 Excel",
            "operations_count": operations_written,
            "warnings": warnings + [str(exc)],
        }


def execute_operations_to_excel(template_path, operations):
    warnings = []
    operations_written = 0
    template, error_result = _validate_template_and_operations(template_path, operations)
    if error_result:
        return error_result

    wb, error_result = _load_workbook(template, warnings)
    if error_result:
        return error_result

    table_operations_count = sum(
        1
        for operation in operations
        if isinstance(operation, dict) and str(operation.get("table_name") or "").strip()
    )

    try:
        ws = wb.active
        for index, operation in enumerate(operations, start=1):
            if not isinstance(operation, dict):
                warnings.append(f"第 {index} 条 operation 无效，已跳过。")
                continue

            operation_type = str(operation.get("operation") or "").strip()
            if operation_type != "write_text":
                warnings.append(f"暂不支持 operation={operation_type or '空'}，已跳过。")
                continue

            target_cell = str(operation.get("target_cell") or "").strip()
            if not target_cell:
                warnings.append(f"第 {index} 条 operation 缺少 target_cell，已跳过。")
                continue

            try:
                _write_value(ws, target_cell, operation.get("value"), warnings)
                operations_written += 1
            except Exception as exc:
                logger.warning(
                    "[CoreExcelExecutor] Operation failed: index=%s target=%s error=%s",
                    index,
                    target_cell,
                    exc,
                    exc_info=True,
                )
                warnings.append(f"{target_cell} 写入失败：{exc}")

        return _save_workbook(
            wb,
            template,
            operations_written,
            warnings,
            f"legacy table_operations={table_operations_count}",
        )
    except Exception as exc:
        logger.exception("[CoreExcelExecutor] Excel save failed")
        return {
            "success": False,
            "error": "无法保存 Excel",
            "operations_count": operations_written,
            "warnings": warnings + [str(exc)],
        }

"""V4 Excel template fingerprint extraction."""

import hashlib
import json
import posixpath
import re
import zipfile
from datetime import date, datetime, time
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


MAX_KEYWORD_CELLS = 1000
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}
WORKBOOK_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _normalize_text_value(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        text = value.isoformat()
    else:
        text = str(value)

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = "\n".join(re.sub(r"[ \t\f\v]+", " ", line).strip() for line in text.split("\n"))
    return text.strip()


def _normalize_sheet_name(sheet_name):
    return str(sheet_name or "")


def _cell_sort_key(cell):
    cell_text = str(cell or "")
    try:
        row, column = coordinate_to_tuple(cell_text)
        return row, column, cell_text
    except Exception:
        return 0, 0, cell_text


def _validate_excel_path(excel_path):
    path = Path(excel_path)
    if not path.is_file():
        raise FileNotFoundError("Excel 模板文件不存在")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError("仅支持 .xlsx、.xlsm、.xltx、.xltm 格式的 Excel 模板")
    return path


def _read_workbook_sheet_paths(archive):
    rels = {}
    try:
        with archive.open("xl/_rels/workbook.xml.rels") as rels_file:
            rels_root = ElementTree.parse(rels_file).getroot()
    except KeyError:
        return []

    for rel in rels_root.findall(f"{{{PACKAGE_REL_NS}}}Relationship"):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rels[rel_id] = target

    try:
        with archive.open("xl/workbook.xml") as workbook_file:
            workbook_root = ElementTree.parse(workbook_file).getroot()
    except KeyError:
        return []

    sheet_paths = []
    for sheet in workbook_root.findall(f".//{{{WORKBOOK_NS}}}sheet"):
        name = sheet.attrib.get("name", "")
        rel_id = sheet.attrib.get(f"{{{REL_NS}}}id")
        target = rels.get(rel_id or "")
        if not name or not target:
            continue

        if target.startswith("/"):
            zip_path = target.lstrip("/")
        else:
            zip_path = posixpath.normpath(posixpath.join("xl", target))
        sheet_paths.append((name, zip_path))

    return sheet_paths


def _read_merged_ranges_from_sheet(archive, sheet_name, zip_path):
    merged_ranges = []
    try:
        with archive.open(zip_path) as sheet_file:
            for event, elem in ElementTree.iterparse(sheet_file, events=("end",)):
                if elem.tag == f"{{{WORKBOOK_NS}}}mergeCell":
                    cell_range = elem.attrib.get("ref")
                    if cell_range:
                        merged_ranges.append(f"{_normalize_sheet_name(sheet_name)}!{str(cell_range)}")
                elem.clear()
    except KeyError:
        return []
    return merged_ranges


def _read_merged_ranges(excel_path):
    merged_ranges = []
    try:
        with zipfile.ZipFile(excel_path) as archive:
            for sheet_name, zip_path in _read_workbook_sheet_paths(archive):
                merged_ranges.extend(_read_merged_ranges_from_sheet(archive, sheet_name, zip_path))
    except (zipfile.BadZipFile, ElementTree.ParseError):
        return []
    return sorted(set(str(item) for item in merged_ranges))


def _extract_keyword_cells(workbook, max_cells=MAX_KEYWORD_CELLS):
    keyword_cells = []
    truncated = False

    worksheets = sorted(workbook.worksheets, key=lambda item: _normalize_sheet_name(item.title))
    for worksheet in worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                keyword_cells.append(
                    {
                        "sheet": worksheet.title,
                        "cell": str(cell.coordinate),
                        "value": _normalize_text_value(cell.value),
                    }
                )
                if len(keyword_cells) >= max_cells:
                    truncated = True
                    return _normalize_keyword_cells(keyword_cells), truncated

    return _normalize_keyword_cells(keyword_cells), truncated


def _normalize_keyword_cells(keyword_cells):
    normalized_cells = []
    for item in keyword_cells:
        if not isinstance(item, dict):
            continue
        normalized_cells.append(
            {
                "sheet": _normalize_sheet_name(item.get("sheet", "")),
                "cell": str(item.get("cell", "")),
                "value": _normalize_text_value(item.get("value", "")),
            }
        )

    return sorted(
        normalized_cells,
        key=lambda item: (
            item.get("sheet", ""),
            *_cell_sort_key(item.get("cell", "")),
        ),
    )


def _normalize_fingerprint_parts(sheet_names, merged_ranges, keyword_cells):
    normalized_sheet_names = sorted(_normalize_sheet_name(sheet_name) for sheet_name in sheet_names)
    normalized_merged_ranges = sorted(str(item) for item in merged_ranges)
    normalized_keyword_cells = _normalize_keyword_cells(keyword_cells)
    return {
        "normalized_sheet_names": normalized_sheet_names,
        "normalized_merged_ranges": normalized_merged_ranges,
        "normalized_keyword_cells": normalized_keyword_cells,
    }


def _build_layout_hash(fingerprint_debug):
    hash_payload = {
        "sheet_names": fingerprint_debug.get("normalized_sheet_names", []),
        "merged_ranges": fingerprint_debug.get("normalized_merged_ranges", []),
        "keyword_cells": fingerprint_debug.get("normalized_keyword_cells", []),
    }
    hash_text = json.dumps(hash_payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(hash_text.encode("utf-8")).hexdigest()


def build_template_fingerprint(excel_path, max_cells=MAX_KEYWORD_CELLS):
    path = _validate_excel_path(excel_path)

    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        sheet_names = sorted(_normalize_sheet_name(sheet_name) for sheet_name in workbook.sheetnames)
        merged_ranges = _read_merged_ranges(path)
        keyword_cells, truncated = _extract_keyword_cells(workbook, max_cells=max_cells)
        fingerprint_debug = _normalize_fingerprint_parts(sheet_names, merged_ranges, keyword_cells)
        layout_hash = _build_layout_hash(fingerprint_debug)

        return {
            "sheet_names": sheet_names,
            "merged_ranges": merged_ranges,
            "non_empty_cells_count": len(keyword_cells),
            "non_empty_cells_limit": max_cells,
            "non_empty_cells_truncated": truncated,
            "keyword_cells": keyword_cells,
            "layout_hash": layout_hash,
            "fingerprint_debug": fingerprint_debug,
        }
    finally:
        if workbook is not None:
            workbook.close()

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


TABLE_PATTERNS = [
    {
        "table_name": "配方表",
        "table_key": "formula_items",
        "headers": [
            {"label": "原料名", "field": "name"},
            {"label": "含量", "field": "amount"},
            {"label": "百分比", "field": "percentage"},
        ],
    },
    {
        "table_name": "检测项目表",
        "table_key": "test_items",
        "headers": [
            {"label": "项目名称", "field": "name"},
            {"label": "标准", "field": "standard"},
            {"label": "结果", "field": "result"},
        ],
    },
]

BLOCK_KEYWORDS = [
    "产品说明",
    "产品特点",
    "包装要求",
    "其他要求",
    "备注",
    "说明",
]


def _clean_text(value):
    text = str(value or "").strip()
    return text if text else ""


def _is_numeric_text(text):
    normalized = text.replace(",", "").replace(".", "", 1).replace("-", "", 1)
    return normalized.isdigit()


def _iter_workbook_sheets(template_path):
    workbook = load_workbook(template_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            yield sheet
    finally:
        workbook.close()


def scan_excel_labels(template_path):
    labels = []
    for sheet in _iter_workbook_sheets(template_path):
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = _clean_text(cell.value)
                if not value or _is_numeric_text(value):
                    continue
                labels.append(
                    {
                        "sheet": sheet.title,
                        "cell": cell.coordinate,
                        "value": value,
                    }
                )
    return labels


def scan_table_regions(template_path):
    table_regions = []
    for sheet in _iter_workbook_sheets(template_path):
        for row in sheet.iter_rows():
            string_cells = []
            row_values = []
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                value = _clean_text(cell.value)
                if not value:
                    continue
                string_cells.append((cell, value))
                row_values.append(value)

            for pattern in TABLE_PATTERNS:
                labels = [item["label"] for item in pattern["headers"]]
                if not all(label in row_values for label in labels):
                    continue

                columns = []
                for header in pattern["headers"]:
                    match_cell = next((cell for cell, value in string_cells if value == header["label"]), None)
                    if match_cell is None:
                        continue
                    columns.append(
                        {
                            "label": header["label"],
                            "field": header["field"],
                            "target_col": get_column_letter(match_cell.column),
                            "header_cell": match_cell.coordinate,
                        }
                    )

                table_regions.append(
                    {
                        "sheet": sheet.title,
                        "table_key": pattern["table_key"],
                        "table_name": pattern["table_name"],
                        "header_row": row[0].row if row else None,
                        "start_cell": columns[0]["header_cell"] if columns else "",
                        "columns": columns,
                    }
                )
    return table_regions


def scan_block_regions(template_path):
    block_regions = []
    for label in scan_excel_labels(template_path):
        value = str(label.get("value") or "").strip().rstrip(":：")
        if value not in BLOCK_KEYWORDS:
            continue
        block_regions.append(
            {
                "sheet": label.get("sheet", ""),
                "block_name": value,
                "source_cell": label.get("cell", ""),
                "target_cell": label.get("cell", ""),
            }
        )
    return block_regions

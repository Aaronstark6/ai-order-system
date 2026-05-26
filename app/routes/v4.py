import base64
import json
import re
import shutil
import uuid
from copy import deepcopy
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Optional
from zipfile import BadZipFile

from fastapi import APIRouter, Body, File, Form, UploadFile
from fastapi.responses import FileResponse
from openpyxl.utils.exceptions import InvalidFileException

from app.logger import get_logger
from app.ai_parser import parse_message
from app.chat_preprocessor import preprocess_chat_text
from app.runtime_paths import get_base_dir
from app.v4_batch_template_executor import execute_batch_template_to_excel
from app.v4_excel_rule_executor import execute_excel_rule_preview_to_workbook
from app.v4_excel_renderer import export_description_fields_to_debug_excel
from app.v4_excel_rule_preview import build_excel_rule_preview
from app.v4_excel_rules import get_template_rules, load_excel_render_rules, save_excel_render_rules
from app.v4_excel_rules_validator import validate_excel_render_rules
from app.v4_order_normalizer import normalize_flat_order_to_v4_order_object
from app.v4_examples import list_examples, load_example, save_example
from app.v4_pipeline_state import (
    get_pipeline_state,
    load_order_object_into_pipeline,
    merge_mapping_safety,
    reset_pipeline_state,
    set_block_operations,
    set_current_profile,
    set_current_template,
    set_excel_result,
    set_mapping_counts,
    set_mapping_safety,
    set_pipeline_result,
    set_render_preview,
    set_render_targets,
    set_structured_operations,
    set_table_operations,
    set_template_analysis,
    set_template_learning,
    set_unified_operations,
    set_validator_result,
)
from app.v4_renderer import render_example_to_description_fields
from app.v4_schema import get_product_form, get_product_forms, load_product_schema, save_product_schema
from app.v4_schema_version import check_schema_compatibility, get_current_schema_version
from app.v4_template_profiles import (
    create_template_profile,
    get_current_template_profile,
    list_template_profiles,
    load_template_profile,
    save_template_profile,
    validate_template_profile,
)
from app.v4_template_layout import build_layout_sections_from_template_analysis
from app.v4_template_cache import (
    delete_cached_template,
    get_cached_template_detail,
    list_cached_templates,
    save_fingerprint,
    update_template_info,
)
from app.v4_template_fingerprint import SUPPORTED_SUFFIXES, build_template_fingerprint
from app.v4_template_matcher import match_or_parse_template, match_template
from app.v4_template_rule_executor import (
    execute_ai_template_to_excel,
    execute_rules_to_template_excel,
    execute_rules_to_template_excel_with_preview,
    执行模板规则并生成Excel,
)
from app.v4_validator import validate_example_order


router = APIRouter()
logger = get_logger(__name__)


@router.get("/v4-template-settings")
def v4_template_settings_page():
    return FileResponse(get_base_dir() / "static" / "v4_template_settings.html")


def _save_v4_uploaded_template(file: UploadFile):
    original_name = Path(file.filename or "").name
    if not original_name:
        raise ValueError("上传文件为空")
    suffix = Path(original_name).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("仅支持 .xlsx、.xlsm、.xltx、.xltm 格式的 Excel 模板")

    upload_dir = get_base_dir() / "data" / "v4_template_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    stem = _safe_output_filename_part(Path(original_name).stem or "template")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{stem}_{timestamp}_{uuid.uuid4().hex[:8]}{suffix}"
    output_path = upload_dir / filename

    with output_path.open("wb") as buffer:
        file.file.seek(0)
        shutil.copyfileobj(file.file, buffer)

    if output_path.stat().st_size <= 0:
        output_path.unlink(missing_ok=True)
        raise ValueError("上传文件为空")

    return output_path


def _remove_v4_uploaded_template(path):
    if not path:
        return
    try:
        Path(path).unlink(missing_ok=True)
    except Exception:
        logger.warning("V4 temporary template cleanup failed: path=%s", path, exc_info=True)


def _system_template_relative_path(path):
    resolved = Path(path).resolve()
    try:
        return str(resolved.relative_to(get_base_dir().resolve())).replace("\\", "/")
    except ValueError:
        return str(resolved)


def _save_v4_system_template_file(profile_id, file: UploadFile):
    original_name = Path(file.filename or "").name
    if not original_name:
        raise ValueError("模板文件名不能为空")
    suffix = Path(original_name).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError("仅支持 .xlsx、.xlsm、.xltx、.xltm 格式的 Excel 模板")

    system_templates_dir = get_base_dir() / "v4" / "system_templates"
    system_templates_dir.mkdir(parents=True, exist_ok=True)

    safe_profile_id = _safe_output_filename_part(profile_id or "system_template")
    safe_stem = _safe_output_filename_part(Path(original_name).stem or "template")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_profile_id}_{safe_stem}_{timestamp}_{uuid.uuid4().hex[:8]}{suffix}"
    output_path = system_templates_dir / filename

    with output_path.open("wb") as buffer:
        file.file.seek(0)
        shutil.copyfileobj(file.file, buffer)

    if output_path.stat().st_size <= 0:
        output_path.unlink(missing_ok=True)
        raise ValueError("模板文件为空")

    return output_path, original_name


def _resolve_template_file_path_for_delete(path_value):
    raw_path = str(path_value or "").strip()
    if not raw_path:
        return None

    path = Path(raw_path)
    if ".." in path.parts:
        raise ValueError("template_file_path 不允许包含 ..")

    base_dir = get_base_dir().resolve()
    resolved = path if path.is_absolute() else base_dir / path
    resolved = resolved.resolve()
    try:
        resolved.relative_to(base_dir)
    except ValueError as exc:
        raise ValueError("模板文件不允许位于项目目录外") from exc
    return resolved


def _safe_template_profile_id(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^\w.-]+", "_", text, flags=re.UNICODE)
    text = text.strip("._")
    return text or "default_profile"


def _resolve_template_profile_json_path_for_delete(profile_id, profile):
    safe_id = _safe_template_profile_id(profile_id)
    profile_safe_id = _safe_template_profile_id(profile.get("profile_id") if isinstance(profile, dict) else "")
    if profile_safe_id != safe_id:
        raise ValueError("映射文件校验失败")

    profiles_dir = (get_base_dir() / "v4" / "template_profiles").resolve()
    target_path = (profiles_dir / f"{safe_id}.json").resolve()
    try:
        target_path.relative_to(profiles_dir)
    except ValueError as exc:
        raise ValueError("映射文件不允许位于项目目录外") from exc
    return target_path


def _template_configuration_from_profile(profile):
    render_config = profile.get("render_config") if isinstance(profile, dict) else {}
    if not isinstance(render_config, dict):
        return {}
    configuration = render_config.get("template_configuration")
    return deepcopy(configuration) if isinstance(configuration, dict) else {}


DEFAULT_EXCEL_FEATURE_FLAGS = {
    "image_fields": True,
    "dynamic_tables": False,
    "advanced_write_modes": False,
    "option_write_enhancement": False,
    "format_protection": True,
    "formula_protection": True,
    "export_readback_check": True,
}


def _get_excel_feature_flags(profile):
    configured = profile.get("excel_feature_flags") if isinstance(profile, dict) else {}
    configured = configured if isinstance(configured, dict) else {}
    return {
        key: bool(configured[key]) if key in configured else default
        for key, default in DEFAULT_EXCEL_FEATURE_FLAGS.items()
    }


def _section_configuration_from_profile(profile):
    render_config = profile.get("render_config") if isinstance(profile, dict) else {}
    if not isinstance(render_config, dict):
        return {}
    configuration = render_config.get("section_configuration")
    return deepcopy(configuration) if isinstance(configuration, dict) else {}


def _normalize_template_configuration_items(items):
    if not isinstance(items, list):
        raise ValueError("配置项必须是 list")

    configuration = {}
    for index, item in enumerate(items, 1):
        if not isinstance(item, dict):
            continue
        cell = str(item.get("cell") or "").strip().upper()
        if not cell:
            continue
        field_type = str(item.get("field_type") or "text").strip()
        write_mode = str(item.get("write_mode") or "").strip()
        if field_type in {"table", "dynamic_table"}:
            write_mode = "write_table_cell"
        configuration[cell] = {
            "label": str(item.get("label") or "").strip(),
            "show_in_workspace": bool(item.get("show_in_workspace", True)),
            "display_order": int(item.get("display_order") or index),
            "candidate_field_key": str(item.get("candidate_field_key") or item.get("field_key") or "").strip(),
            "candidate_field_label": str(item.get("candidate_field_label") or item.get("field_label") or "").strip(),
            "candidate_confidence": float(item.get("candidate_confidence") or 0),
            "candidate_source": str(item.get("candidate_source") or "").strip(),
            "candidate_reason": str(item.get("candidate_reason") or "").strip(),
            "confidence_breakdown": item.get("confidence_breakdown") if isinstance(item.get("confidence_breakdown"), dict) else {},
            "intent_type": str(item.get("intent_type") or "").strip(),
            "write_mode": write_mode,
            "label_cell": str(item.get("label_cell") or "").strip().upper(),
            "target_cell": str(item.get("target_cell") or "").strip().upper(),
            "sheet_name": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
            "option_value": str(item.get("option_value") or "").strip(),
            "intent_confidence": float(item.get("intent_confidence") or 0),
            "intent_reason": str(item.get("intent_reason") or "").strip(),
            "ai_extract_hint": str(item.get("ai_extract_hint") or "").strip(),
            "field_type": field_type,
            "image_fit": str(item.get("image_fit") or "contain").strip(),
            "image_anchor_cell": str(item.get("image_anchor_cell") or item.get("target_cell") or item.get("cell") or "").strip().upper(),
            "col_offset": int(item.get("col_offset") or item.get("table_col_offset") or item.get("column_offset") or 0),
            "table_col_offset": int(item.get("table_col_offset") or item.get("col_offset") or item.get("column_offset") or 0),
            "manual_override": bool(item.get("manual_override", False)),
            "user_edited": bool(item.get("user_edited", False)),
        }
    return configuration


_MAPPING_CANDIDATE_RULES = [
    {
        "field_key": "customer_name",
        "field_label": "客户名称",
        "keywords": ["客户名称", "客户公司", "客户", "公司名称", "customer name", "customer company", "customer"],
        "ai_extract_hint": "客户名称 / 客户公司",
    },
    {
        "field_key": "order_date",
        "field_label": "订单日期",
        "keywords": ["订单日期", "下单日期", "日期", "date", "order date"],
        "ai_extract_hint": "订单日期 / 下单日期 / 日期",
    },
    {
        "field_key": "product_name",
        "field_label": "产品名称",
        "keywords": ["产品名称", "产品名", "产品", "品名", "product name", "product"],
        "ai_extract_hint": "产品名称 / 品名",
    },
    {
        "field_key": "quantity",
        "field_label": "数量",
        "keywords": ["数量", "订单数量", "qty", "quantity", "count"],
        "ai_extract_hint": "数量 / 订单数量",
    },
    {
        "field_key": "specification",
        "field_label": "规格",
        "keywords": ["规格", "规格型号", "型号", "spec", "specification"],
        "ai_extract_hint": "规格 / 规格型号",
    },
    {
        "field_key": "packaging.container_type",
        "field_label": "容器类型",
        "keywords": ["容器要求", "容器类型", "容器", "瓶装", "袋装", "管装", "罐装", "包装方式", "container type", "bottle", "bag", "tube", "jar"],
        "ai_extract_hint": "容器类型 / 容器要求 / 瓶装 / 袋装 / 管装",
        "priority": 50,
    },
    {
        "field_key": "packaging.quantity_per_unit",
        "field_label": "装量",
        "keywords": ["装量", "包装数量", "包装规格", "每瓶", "每袋", "每管", "粒/瓶", "片/管", "粒每瓶", "片每管", "quantity per unit", "per bottle", "per bag", "per tube"],
        "ai_extract_hint": "装量 / 每瓶数量 / 每袋数量 / 每管数量",
        "priority": 58,
    },
    {
        "field_key": "packaging.capacity",
        "field_label": "容量",
        "keywords": ["容量", "毫升", "克重", "ml", "milliliter", "g", "gram", "grams", "capacity"],
        "ai_extract_hint": "容量 / ml / 毫升 / g / 克",
        "priority": 56,
    },
    {
        "field_key": "packaging.container_color",
        "field_label": "容器颜色",
        "keywords": ["瓶身颜色", "容器颜色", "瓶子颜色", "罐子颜色", "袋子颜色", "container color", "bottle color", "jar color"],
        "ai_extract_hint": "容器颜色 / 瓶身颜色 / 瓶子颜色",
        "priority": 66,
    },
    {
        "field_key": "packaging.cap_color",
        "field_label": "盖子颜色",
        "keywords": ["盖子颜色", "瓶盖颜色", "盖颜色", "cap color", "lid color"],
        "ai_extract_hint": "盖子颜色 / 瓶盖颜色",
        "priority": 64,
    },
    {
        "field_key": "packaging.seal_method",
        "field_label": "密封方式",
        "keywords": ["密封方式", "瓶口密封", "盖子密封", "袋口密封", "密封", "封口", "铝箔", "热封", "塑封", "seal method", "sealing", "foil seal", "heat seal"],
        "ai_extract_hint": "密封方式 / 瓶口密封 / 盖子密封 / 铝箔 / 热封 / 塑封",
        "priority": 62,
    },
    {
        "field_key": "packaging",
        "field_label": "包装",
        "keywords": ["包装", "包装规格", "包装要求", "package", "packaging"],
        "ai_extract_hint": "包装 / 包装规格 / 包装要求",
        "priority": 10,
    },
    {
        "field_key": "amount",
        "field_label": "金额",
        "keywords": ["金额", "总金额", "货值", "总价", "价格", "amount", "price", "total"],
        "ai_extract_hint": "金额 / 总金额 / 价格",
    },
]


_SECTION_FIELD_HINTS = [
    (
        ["header", "order", "customer", "client", "basic", "info", "表头", "订单", "客户", "基本信息"],
        {"customer_name": 0.22, "order_date": 0.18},
    ),
    (
        ["product", "item", "spec", "packing", "package", "detail", "table", "产品", "品名", "规格", "包装", "明细", "表格"],
        {
            "product_name": 0.22,
            "quantity": 0.16,
            "packaging": 0.18,
            "packaging.container_type": 0.18,
            "packaging.quantity_per_unit": 0.18,
            "packaging.capacity": 0.18,
            "packaging.container_color": 0.18,
            "packaging.cap_color": 0.18,
            "packaging.seal_method": 0.18,
            "specification": 0.18,
        },
    ),
    (
        ["image", "photo", "picture", "attachment", "图片", "图", "附件"],
        {"attachment": 0.24},
    ),
    (
        ["summary", "total", "amount", "price", "合计", "汇总", "金额", "总计", "价格"],
        {"amount": 0.22, "quantity": 0.08},
    ),
]

_GENERIC_LABEL_HINTS = {
    "名称": {"customer_name": 0.34, "product_name": 0.34},
    "name": {"customer_name": 0.34, "product_name": 0.34},
    "日期": {"order_date": 0.42},
    "date": {"order_date": 0.42},
    "数量": {"quantity": 0.48},
    "qty": {"quantity": 0.48},
    "规格": {"specification": 0.48},
    "型号": {"specification": 0.42},
    "容器": {"packaging.container_type": 0.58},
    "瓶装": {"packaging.container_type": 0.58},
    "袋装": {"packaging.container_type": 0.58},
    "管装": {"packaging.container_type": 0.58},
    "装量": {"packaging.quantity_per_unit": 0.58},
    "每瓶": {"packaging.quantity_per_unit": 0.58},
    "每袋": {"packaging.quantity_per_unit": 0.58},
    "每管": {"packaging.quantity_per_unit": 0.58},
    "容量": {"packaging.capacity": 0.58},
    "毫升": {"packaging.capacity": 0.58},
    "瓶身颜色": {"packaging.container_color": 0.58},
    "容器颜色": {"packaging.container_color": 0.58},
    "瓶子颜色": {"packaging.container_color": 0.58},
    "盖子颜色": {"packaging.cap_color": 0.58},
    "瓶盖颜色": {"packaging.cap_color": 0.58},
    "密封": {"packaging.seal_method": 0.58},
    "封口": {"packaging.seal_method": 0.58},
    "铝箔": {"packaging.seal_method": 0.58},
    "热封": {"packaging.seal_method": 0.58},
    "塑封": {"packaging.seal_method": 0.58},
    "包装": {"packaging": 0.48},
    "金额": {"amount": 0.48},
    "价格": {"amount": 0.42},
}


def load_field_catalog():
    path = get_base_dir() / "v4" / "schemas" / "field_catalog.json"
    if not path.is_file():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        logger.warning("V4 field catalog load failed: path=%s", path, exc_info=True)
        return {}
    if not isinstance(data, dict) or not isinstance(data.get("domains"), list):
        logger.warning("V4 field catalog invalid format: path=%s", path)
        return {}
    return data


def flatten_field_catalog(catalog=None):
    catalog = catalog if isinstance(catalog, dict) else load_field_catalog()
    domains = catalog.get("domains") if isinstance(catalog.get("domains"), list) else []
    fields = []
    for domain in domains:
        if not isinstance(domain, dict):
            continue
        domain_key = str(domain.get("key") or "").strip()
        domain_label = str(domain.get("label") or domain_key).strip()
        domain_fields = domain.get("fields") if isinstance(domain.get("fields"), list) else []
        for field in domain_fields:
            if not isinstance(field, dict):
                continue
            field_key = str(field.get("field_key") or "").strip()
            if not field_key:
                continue
            aliases = field.get("aliases") if isinstance(field.get("aliases"), list) else []
            keywords = field.get("keywords") if isinstance(field.get("keywords"), list) else []
            fields.append(
                {
                    "key": str(field.get("key") or field_key).strip(),
                    "field_key": field_key,
                    "label": str(field.get("label") or field_key).strip() or field_key,
                    "domain": domain_key,
                    "domain_label": domain_label,
                    "aliases": [str(item).strip() for item in aliases if str(item).strip()],
                    "keywords": [str(item).strip() for item in keywords if str(item).strip()],
                    "ai_extract_hint": str(field.get("ai_extract_hint") or field.get("label") or field_key).strip(),
                    "type": str(field.get("type") or "text").strip() or "text",
                    "enabled": field.get("enabled") is not False,
                    "priority": int(field.get("priority") or 0),
                }
            )
    return fields


def get_field_catalog_labels():
    labels = {}
    for field in flatten_field_catalog():
        if not field.get("enabled"):
            continue
        field_key = str(field.get("field_key") or "").strip()
        label = str(field.get("label") or field_key).strip()
        if field_key and label:
            labels[field_key] = label
    return labels


def _field_catalog_rule_from_field(field):
    keywords = []
    for value in [field.get("label"), *(field.get("aliases") or []), *(field.get("keywords") or [])]:
        text = str(value or "").strip()
        if text and text not in keywords:
            keywords.append(text)
    return {
        "field_key": field["field_key"],
        "field_label": field.get("label") or field["field_key"],
        "keywords": keywords,
        "ai_extract_hint": field.get("ai_extract_hint") or field.get("label") or field["field_key"],
        "priority": int(field.get("priority") or 0),
        "field_type": field.get("type") or "text",
        "source": "field_catalog",
    }


def get_field_catalog_candidate_rules():
    rules = []
    seen = set()
    for field in flatten_field_catalog():
        field_key = str(field.get("field_key") or "").strip()
        if not field.get("enabled") or not field_key or field_key in seen:
            continue
        rules.append(_field_catalog_rule_from_field(field))
        seen.add(field_key)
    for rule in _MAPPING_CANDIDATE_RULES:
        field_key = str(rule.get("field_key") or "").strip()
        if field_key and field_key not in seen:
            rules.append(rule)
            seen.add(field_key)
    return rules or list(_MAPPING_CANDIDATE_RULES)


def _field_key_from_catalog_keywords(text):
    normalized_text = _normalize_candidate_text(text)
    if not normalized_text:
        return ""
    matches = []
    for rule in get_field_catalog_candidate_rules():
        if rule.get("source") != "field_catalog":
            continue
        best_score = _keyword_score_for_rule(normalized_text, rule)
        if best_score <= 0:
            continue
        matches.append((best_score, int(rule.get("priority") or 0), str(rule.get("field_key") or "")))
    if not matches:
        return ""
    matches.sort(key=lambda item: (-item[0], -item[1], item[2]))
    return matches[0][2]


def _normalize_candidate_text(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s:：/\\|,，.。;；()（）\[\]【】_-]+", "", text)
    return text


def _cell_point(cell):
    match = re.match(r"^([A-Za-z]+)(\d+)$", str(cell or "").strip())
    if not match:
        return None
    col = 0
    for char in match.group(1).upper():
        col = col * 26 + ord(char) - 64
    return int(match.group(2)), col


def _cell_ref(row, col):
    if not row or not col or row < 1 or col < 1:
        return ""
    letters = ""
    value = int(col)
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return f"{letters}{int(row)}"


def _section_for_cell(layout_sections, cell):
    point = _cell_point(cell)
    if not point:
        return {}
    row, col = point
    for section in layout_sections if isinstance(layout_sections, list) else []:
        if not isinstance(section, dict):
            continue
        bounds = section.get("bounds") if isinstance(section.get("bounds"), dict) else {}
        try:
            start_row = int(bounds.get("start_row") or 0)
            end_row = int(bounds.get("end_row") or 0)
            start_col = int(bounds.get("start_col") or 0)
            end_col = int(bounds.get("end_col") or 0)
        except (TypeError, ValueError):
            continue
        if start_row <= row <= end_row and start_col <= col <= end_col:
            return section
    return {}


def _safe_cell_text(value):
    return "" if value is None else str(value).strip()


def _load_template_intent_context(template_path):
    if not template_path:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    try:
        workbook = load_workbook(template_path, data_only=False)
        worksheet = workbook.active
    except Exception:
        return {}

    merged_lookup = {}
    try:
        for merged_range in worksheet.merged_cells.ranges:
            start_cell = merged_range.start_cell.coordinate
            range_text = str(merged_range)
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_lookup[_cell_ref(row, col)] = {
                        "start_cell": start_cell,
                        "range": range_text,
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col,
                    }
    except Exception:
        merged_lookup = {}

    row_texts = {}
    cells = {}
    max_row = min(int(worksheet.max_row or 1), 200)
    max_col = min(int(worksheet.max_column or 1), 80)
    for row in range(1, max_row + 1):
        row_values = []
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            coord = cell.coordinate
            text = _safe_cell_text(cell.value)
            if text:
                row_values.append(text)
            merged = merged_lookup.get(coord, {})
            cells[coord] = {
                "row": row,
                "col": col,
                "value": text,
                "font_size": float(cell.font.sz or 0) if getattr(cell, "font", None) else 0,
                "bold": bool(cell.font.bold) if getattr(cell, "font", None) else False,
                "align": str(cell.alignment.horizontal or "").lower() if getattr(cell, "alignment", None) else "",
                "merged_start": merged.get("start_cell", ""),
                "merged_range": merged.get("range", ""),
                "is_merged": bool(merged),
            }
        row_texts[row] = row_values

    return {
        "cells": cells,
        "row_texts": row_texts,
        "max_row": max_row,
        "max_col": max_col,
    }


def _excel_color_text(color):
    if color is None:
        return ""
    rgb = getattr(color, "rgb", None)
    rgb_text = str(rgb or "").strip()
    if re.fullmatch(r"[A-Fa-f0-9]{6}|[A-Fa-f0-9]{8}", rgb_text) and rgb_text != "00000000":
        return rgb_text
    indexed = getattr(color, "indexed", None)
    if type(indexed) is int:
        return f"indexed:{indexed}"
    theme = getattr(color, "theme", None)
    if type(theme) is int:
        return f"theme:{theme}"
    return ""


def _excel_horizontal_align(alignment):
    value = str(getattr(alignment, "horizontal", "") or "").strip().lower()
    if value == "centercontinuous":
        return "center"
    if value in {"left", "center", "right", "general"}:
        return value
    if value in {"fill", "justify", "distributed"}:
        return "left"
    return "general"


def _excel_vertical_align(alignment):
    value = str(getattr(alignment, "vertical", "") or "").strip().lower()
    if value in {"top", "center", "bottom"}:
        return value
    if value in {"justify", "distributed"}:
        return "top"
    return "bottom"


def _merge_candidate_with_saved_configuration(candidate, saved_item):
    candidate = candidate if isinstance(candidate, dict) else {}
    saved_item = saved_item if isinstance(saved_item, dict) else {}
    def saved_or_candidate(saved_key, candidate_key, default=""):
        if saved_key in saved_item:
            return saved_item.get(saved_key)
        return candidate.get(candidate_key, default)

    return {
        "candidate_field_key": saved_or_candidate("candidate_field_key", "field_key"),
        "candidate_field_label": saved_or_candidate("candidate_field_label", "field_label"),
        "candidate_confidence": saved_or_candidate("candidate_confidence", "confidence", 0),
        "candidate_source": saved_or_candidate("candidate_source", "source"),
        "candidate_reason": saved_or_candidate("candidate_reason", "candidate_reason"),
        "confidence_breakdown": saved_item.get("confidence_breakdown") if isinstance(saved_item.get("confidence_breakdown"), dict) else (candidate.get("confidence_breakdown") if isinstance(candidate.get("confidence_breakdown"), dict) else {}),
        "intent_type": saved_or_candidate("intent_type", "intent_type"),
        "write_mode": saved_or_candidate("write_mode", "write_mode"),
        "label_cell": saved_or_candidate("label_cell", "label_cell"),
        "target_cell": saved_or_candidate("target_cell", "target_cell"),
        "option_value": saved_or_candidate("option_value", "option_value"),
        "semantic_promoted": saved_or_candidate("semantic_promoted", "semantic_promoted", False),
        "show_in_workspace": saved_or_candidate("show_in_workspace", "show_in_workspace", True),
        "intent_confidence": saved_or_candidate("intent_confidence", "intent_confidence", 0),
        "intent_reason": saved_or_candidate("intent_reason", "intent_reason"),
        "ai_extract_hint": saved_or_candidate("ai_extract_hint", "ai_extract_hint"),
        "field_type": saved_or_candidate("field_type", "field_type", "text"),
        "image_fit": saved_or_candidate("image_fit", "image_fit", "contain"),
        "image_anchor_cell": saved_or_candidate(
            "image_anchor_cell",
            "image_anchor_cell",
            saved_or_candidate("target_cell", "target_cell") or saved_item.get("cell") or candidate.get("cell") or ""
        ),
        "col_offset": saved_or_candidate("col_offset", "col_offset", 0),
        "table_col_offset": saved_or_candidate("table_col_offset", "table_col_offset", 0),
    }


def _build_visual_grid(template_path, mapping_candidates, template_configuration, semantic_by_cell=None):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(template_path, data_only=False)
    worksheet = workbook.active
    rows = min(int(worksheet.max_row or 1), 80)
    cols = min(int(worksheet.max_column or 1), 30)
    default_row_height = float(worksheet.sheet_format.defaultRowHeight or 15)
    default_col_width = float(worksheet.sheet_format.defaultColWidth or 8.43)
    row_heights = {}
    col_widths = {}

    for row in range(1, rows + 1):
        row_dimension = worksheet.row_dimensions.get(row)
        height = row_dimension.height if row_dimension is not None else None
        row_heights[str(row)] = float(height if height is not None else default_row_height)

    for col in range(1, cols + 1):
        column_letter = get_column_letter(col)
        column_dimension = worksheet.column_dimensions.get(column_letter)
        width = column_dimension.width if column_dimension is not None else None
        col_widths[column_letter] = float(width if width is not None else default_col_width)

    merges = []
    merge_lookup = {}
    for merged_range in worksheet.merged_cells.ranges:
        range_text = str(merged_range)
        start_cell = merged_range.start_cell.coordinate
        if merged_range.min_row <= rows and merged_range.min_col <= cols:
            merges.append(
                {
                    "range": range_text,
                    "start_cell": start_cell,
                    "start_row": merged_range.min_row,
                    "start_col": merged_range.min_col,
                    "end_row": min(merged_range.max_row, rows),
                    "end_col": min(merged_range.max_col, cols),
                }
            )
        for row in range(merged_range.min_row, min(merged_range.max_row, rows) + 1):
            for col in range(merged_range.min_col, min(merged_range.max_col, cols) + 1):
                merge_lookup[_cell_ref(row, col)] = {
                    "range": range_text,
                    "start_cell": start_cell,
                    "merge_start": _cell_ref(row, col) == start_cell,
                }

    candidates_by_cell = {
        str(item.get("cell") or "").strip().upper(): item
        for item in mapping_candidates if isinstance(item, dict) and item.get("cell")
    }
    saved = template_configuration if isinstance(template_configuration, dict) else {}
    semantic_lookup = semantic_by_cell if isinstance(semantic_by_cell, dict) else {}
    cells = []

    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell_obj = worksheet.cell(row=row, column=col)
            cell_ref = cell_obj.coordinate
            raw_value = _safe_cell_text(cell_obj.value)
            merge_info = merge_lookup.get(cell_ref, {})
            is_merged = bool(merge_info)
            merge_start = bool(merge_info.get("merge_start", False))
            display_value = raw_value if (not is_merged or merge_start) else ""
            font = cell_obj.font
            alignment = cell_obj.alignment
            horizontal_align = _excel_horizontal_align(alignment)
            vertical_align = _excel_vertical_align(alignment)
            saved_item = saved.get(cell_ref, {})
            v4_fields = _merge_candidate_with_saved_configuration(candidates_by_cell.get(cell_ref, {}), saved_item)
            semantic = _primary_semantic_for_cell(semantic_lookup, cell_ref)
            semantic_fields = {
                "semantic_type": semantic.get("type", ""),
                "semantic_region_id": semantic.get("region_id", ""),
                "semantic_confidence": semantic.get("confidence", 0),
                "semantic_reason": semantic.get("reason", ""),
            } if semantic else {}
            cells.append(
                {
                    "cell": cell_ref,
                    "row": row,
                    "col": col,
                    "value": raw_value,
                    "display_value": display_value,
                    "font_bold": bool(font.bold),
                    "font_size": float(font.sz or 0),
                    "font_name": str(font.name or ""),
                    "font_color": _excel_color_text(font.color),
                    "fill_color": _excel_color_text(cell_obj.fill.fgColor),
                    "align": horizontal_align,
                    "horizontal_align": horizontal_align,
                    "vertical_align": vertical_align,
                    "wrap_text": bool(alignment.wrap_text),
                    "shrink_to_fit": bool(alignment.shrink_to_fit),
                    "merged_range": merge_info.get("range", ""),
                    "is_merged": is_merged,
                    "merge_start": merge_start,
                    **v4_fields,
                    **semantic_fields,
                }
            )

    return {
        "rows": rows,
        "cols": cols,
        "row_heights": row_heights,
        "col_widths": col_widths,
        "cells": cells,
        "merges": merges,
    }


def _intent_cell_info(context, cell):
    cells = context.get("cells") if isinstance(context, dict) else {}
    return cells.get(str(cell or "").strip().upper(), {}) if isinstance(cells, dict) else {}


def _intent_cell_text(context, row, col):
    if row < 1 or col < 1:
        return ""
    return str(_intent_cell_info(context, _cell_ref(row, col)).get("value") or "").strip()


def _intent_blank_cell(context, row, col):
    if row < 1 or col < 1:
        return False
    ref = _cell_ref(row, col)
    info = _intent_cell_info(context, ref)
    if not info:
        return True
    start_cell = str(info.get("merged_start") or "").strip().upper()
    if start_cell and start_cell != ref:
        start_info = _intent_cell_info(context, start_cell)
        if str(start_info.get("value") or "").strip():
            return False
    return not str(info.get("value") or "").strip()


def _intent_target_cell(context, row, col):
    ref = _cell_ref(row, col)
    info = _intent_cell_info(context, ref)
    start_cell = str(info.get("merged_start") or "").strip().upper()
    return start_cell or ref


def _infer_right_target_cell(context, cell):
    point = _cell_point(cell)
    if not point:
        return ""
    row, col = point
    for offset in range(1, 5):
        target_col = col + offset
        if target_col > int(context.get("max_col") or target_col):
            break
        if _intent_blank_cell(context, row, target_col):
            target_cell = _intent_target_cell(context, row, target_col)
            if target_cell and target_cell != cell:
                return target_cell
    return ""


def _infer_below_target_cell(context, cell):
    point = _cell_point(cell)
    if not point:
        return ""
    row, col = point
    for offset in range(1, 4):
        target_row = row + offset
        if target_row > int(context.get("max_row") or target_row):
            break
        if _intent_blank_cell(context, target_row, col):
            return _intent_target_cell(context, target_row, col)
    return ""


def _row_short_text_count(context, row):
    row_texts = context.get("row_texts") if isinstance(context, dict) else {}
    values = row_texts.get(row, []) if isinstance(row_texts, dict) else []
    return sum(1 for value in values if 0 < len(str(value or "").strip()) <= 12)


def _intent_for_label(label_text, cell, section, label_index, context):
    text = str(label_text or "").strip()
    normalized = _normalize_candidate_text(text)
    point = _cell_point(cell)
    row, col = point if point else (0, 0)
    info = _intent_cell_info(context, cell)
    section_type = str(section.get("section_type") or section.get("semantic_type") or section.get("title") or "").lower() if isinstance(section, dict) else ""
    row_short_count = _row_short_text_count(context, row)
    text_len = len(text)
    has_colon = "：" in text or ":" in text
    ends_colon = text.endswith(("：", ":"))
    right_target = _infer_right_target_cell(context, cell)
    below_target = _infer_below_target_cell(context, cell)

    image_words = ["图片", "照片", "附图", "image", "photo"]
    attachment_words = ["附件", "上传", "附档", "attachment", "upload"]
    note_words = ["备注", "注意", "说明", "规则", "原则", "要求", "note", "instruction"]
    example_words = ["示例", "例如", "默认", "样例", "example", "sample", "default"]
    option_marks = ["□", "☐", "☑", "√", "✔", "[ ]", "（ ）", "( )"]

    if any(word in text.lower() for word in attachment_words):
        return {
            "intent_type": "attachment_hint",
            "write_mode": "skip",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.82,
            "intent_reason": "attachment_keyword",
        }
    if any(word in text.lower() for word in image_words):
        return {
            "intent_type": "image_area",
            "write_mode": "skip",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.82,
            "intent_reason": "image_keyword",
        }

    if any(mark in text for mark in option_marks):
        option_value = re.sub(r"[□☐☑√✔\[\]\(\)（）\s]+", "", text).strip("：:")
        return {
            "intent_type": "option_checkbox",
            "write_mode": "check_option",
            "label_cell": cell,
            "target_cell": cell,
            "option_value": option_value,
            "intent_confidence": 0.86,
            "intent_reason": "checkbox_mark",
        }

    if row_short_count >= 3 and text_len <= 12 and not has_colon:
        return {
            "intent_type": "option_text_choice",
            "write_mode": "select_option_text",
            "label_cell": cell,
            "target_cell": cell,
            "option_value": text,
            "intent_confidence": 0.68,
            "intent_reason": "multiple_short_options_in_row",
        }

    if row <= 3 and (
        info.get("is_merged")
        or info.get("align") == "center"
        or float(info.get("font_size") or 0) >= 14
        or info.get("bold")
    ):
        return {
            "intent_type": "title",
            "write_mode": "none",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.86,
            "intent_reason": "top_merged_center_or_large_text",
        }

    if (
        "table" in section_type
        and text_len <= 16
        and row <= int((section.get("bounds") or {}).get("start_row") or row) + 1
    ):
        return {
            "intent_type": "table_column_header",
            "write_mode": "write_table_column",
            "label_cell": cell,
            "target_cell": below_target,
            "option_value": "",
            "intent_confidence": 0.78 if below_target else 0.62,
            "intent_reason": "table_section_header_row",
        }

    if "table" in section_type and col <= int((section.get("bounds") or {}).get("start_col") or col):
        return {
            "intent_type": "table_row_field",
            "write_mode": "write_row_field",
            "label_cell": cell,
            "target_cell": right_target,
            "option_value": "",
            "intent_confidence": 0.72 if right_target else 0.58,
            "intent_reason": "table_left_row_label",
        }

    if any(word in text.lower() for word in example_words):
        return {
            "intent_type": "readonly_example",
            "write_mode": "skip",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.72,
            "intent_reason": "example_or_default_keyword",
        }

    if any(word in text.lower() for word in note_words) or text_len >= 28:
        return {
            "intent_type": "note_instruction",
            "write_mode": "skip",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.76,
            "intent_reason": "long_instruction_or_note_keyword",
        }

    if ends_colon and right_target:
        return {
            "intent_type": "label_fill_right",
            "write_mode": "write_right_cell",
            "label_cell": cell,
            "target_cell": right_target,
            "option_value": "",
            "intent_confidence": 0.84,
            "intent_reason": "colon_label_with_blank_right_cell",
        }

    if ends_colon and below_target:
        return {
            "intent_type": "label_fill_below",
            "write_mode": "write_below_cell",
            "label_cell": cell,
            "target_cell": below_target,
            "option_value": "",
            "intent_confidence": 0.74,
            "intent_reason": "colon_label_with_blank_below_cell",
        }

    if has_colon and not ends_colon:
        return {
            "intent_type": "inline_fill_after_colon",
            "write_mode": "append_after_colon",
            "label_cell": cell,
            "target_cell": cell,
            "option_value": "",
            "intent_confidence": 0.7,
            "intent_reason": "inline_colon_text",
        }

    if (
        text_len <= 18
        and (info.get("bold") or info.get("is_merged") or "section" in section_type or "block" in section_type)
        and not right_target
    ):
        return {
            "intent_type": "section_header",
            "write_mode": "none",
            "label_cell": cell,
            "target_cell": "",
            "option_value": "",
            "intent_confidence": 0.64,
            "intent_reason": "short_bold_or_merged_section_text",
        }

    if right_target:
        return {
            "intent_type": "label_fill_right",
            "write_mode": "write_right_cell",
            "label_cell": cell,
            "target_cell": right_target,
            "option_value": "",
            "intent_confidence": 0.56,
            "intent_reason": "short_label_with_blank_right_cell",
        }

    return {
        "intent_type": "unknown",
        "write_mode": "skip",
        "label_cell": cell,
        "target_cell": "",
        "option_value": "",
        "intent_confidence": 0.2,
        "intent_reason": "no_reliable_intent_rule_matched",
    }


def _mapping_rule_by_key():
    return {rule["field_key"]: rule for rule in get_field_catalog_candidate_rules()}


def _section_candidate_text(section):
    if not isinstance(section, dict):
        return ""
    values = [
        section.get("title"),
        section.get("source_region_name"),
        section.get("semantic_type"),
        section.get("section_type"),
        section.get("section_key"),
    ]
    matched_keywords = section.get("matched_keywords")
    if isinstance(matched_keywords, list):
        values.extend(matched_keywords)
    return _normalize_candidate_text(" ".join(str(value or "") for value in values))


def _keyword_score_for_rule(normalized_text, rule):
    score = 0
    for keyword in rule["keywords"]:
        normalized_keyword = _normalize_candidate_text(keyword)
        if not normalized_keyword:
            continue
        if normalized_keyword == "g" and not re.search(r"\d+g\b", normalized_text):
            continue
        if normalized_text == normalized_keyword:
            score = max(score, 0.68)
        elif normalized_keyword in normalized_text:
            score = max(score, 0.58)

    for generic_label, hints in _GENERIC_LABEL_HINTS.items():
        normalized_generic = _normalize_candidate_text(generic_label)
        if normalized_generic and normalized_generic in normalized_text:
            score = max(score, float(hints.get(rule["field_key"], 0) or 0))
    return score


def _section_score_for_rule(section_text, rule):
    if not section_text:
        return 0
    score = 0
    field_key = str(rule.get("field_key") or "")
    domain_key = field_key.split(".", 1)[0] if "." in field_key else field_key
    for keywords, field_scores in _SECTION_FIELD_HINTS:
        if any(_normalize_candidate_text(keyword) in section_text for keyword in keywords):
            score = max(score, float(field_scores.get(field_key, field_scores.get(domain_key, 0)) or 0))
    return score


def _reading_order_score_for_rule(rule, label_index, total_labels, cell):
    point = _cell_point(cell)
    row = point[0] if point else 0
    total = max(int(total_labels or 1), 1)
    ratio = max(min((int(label_index or 1) - 1) / total, 1), 0)
    field_key = rule["field_key"]

    if field_key in {"customer_name", "order_date"} and (ratio <= 0.3 or (row and row <= 8)):
        return 0.08
    if (field_key in {"product_name", "quantity", "specification", "packaging"} or field_key.startswith("packaging.")) and ratio >= 0.25:
        return 0.06
    if field_key == "amount" and ratio >= 0.65:
        return 0.07
    return 0


def _neighbor_score_for_rule(neighbor_text, rule):
    normalized_neighbor = _normalize_candidate_text(neighbor_text)
    if not normalized_neighbor:
        return 0
    score = 0
    for keyword in rule["keywords"]:
        normalized_keyword = _normalize_candidate_text(keyword)
        if normalized_keyword and normalized_keyword in normalized_neighbor:
            score = max(score, 0.08)
    return score


def _reason_from_breakdown(breakdown):
    reason_keys = [
        ("keyword", "keyword"),
        ("section", "section"),
        ("neighbor", "neighbor_label"),
        ("reading_order", "reading_order"),
    ]
    return "+".join(label for key, label in reason_keys if float(breakdown.get(key) or 0) > 0) or ""


def _candidate_for_label(label_text, section, label_index=1, total_labels=1, cell="", neighbor_text=""):
    normalized_text = _normalize_candidate_text(label_text)
    section_text = _section_candidate_text(section)
    scored_candidates = []
    for rule in get_field_catalog_candidate_rules():
        keyword_score = _keyword_score_for_rule(normalized_text, rule)
        section_score = _section_score_for_rule(section_text, rule)
        neighbor_score = _neighbor_score_for_rule(neighbor_text, rule)
        reading_order_score = _reading_order_score_for_rule(rule, label_index, total_labels, cell)

        if keyword_score <= 0 and neighbor_score <= 0:
            continue

        breakdown = {
            "keyword": round(keyword_score, 2),
            "section": round(section_score, 2),
            "neighbor": round(neighbor_score, 2),
            "reading_order": round(reading_order_score, 2),
        }
        confidence = min(0.98, sum(breakdown.values()))
        if confidence < 0.38:
            continue
        scored_candidates.append(
            {
                "field_key": rule["field_key"],
                "field_label": rule["field_label"],
                "confidence": round(confidence, 2),
                "ai_extract_hint": rule["ai_extract_hint"],
                "candidate_reason": _reason_from_breakdown(breakdown),
                "confidence_breakdown": {key: value for key, value in breakdown.items() if value > 0},
                "priority": int(rule.get("priority") or 0),
                "source": rule.get("source") or "hardcoded_rules",
            }
        )

    if scored_candidates:
        scored_candidates.sort(
            key=lambda item: (
                -int(item.get("source") == "field_catalog"),
                -item["confidence"],
                -int(item.get("priority") or 0),
                item["field_key"],
            )
        )
        return scored_candidates[0]

    return {
        "field_key": "",
        "field_label": "",
        "confidence": 0,
        "ai_extract_hint": "",
        "candidate_reason": "",
        "confidence_breakdown": {},
    }


def _slugify_semantic_field_key(label, semantic_type, cell, section_text=""):
    text = f"{label or ''} {section_text or ''}".strip().lower()
    semantic_type = str(semantic_type or "").strip()
    catalog_field_key = _field_key_from_catalog_keywords(text)
    if catalog_field_key:
        return catalog_field_key
    keyword_map = [
        (("瓶身颜色", "容器颜色", "瓶子颜色", "罐子颜色", "袋子颜色", "container color", "bottle color", "jar color"), "packaging.container_color"),
        (("盖子颜色", "瓶盖颜色", "盖颜色", "cap color", "lid color"), "packaging.cap_color"),
        (("瓶口密封", "盖子密封", "袋口密封", "密封方式", "密封", "封口", "铝箔", "热封", "塑封", "seal method", "sealing", "foil seal", "heat seal"), "packaging.seal_method"),
        (("装量", "包装数量", "每瓶", "每袋", "每管", "粒/瓶", "片/管", "粒每瓶", "片每管", "quantity per unit", "per bottle", "per bag", "per tube"), "packaging.quantity_per_unit"),
        (("容量", "毫升", "克重", "ml", "milliliter", "g", "gram", "grams", "capacity"), "packaging.capacity"),
        (("容器要求", "容器类型", "容器", "瓶装", "袋装", "管装", "罐装", "container type", "bottle", "bag", "tube", "jar"), "packaging.container_type"),
        (("客户名称", "客户", "customer"), "customer_name"),
        (("数量", "quantity", "qty", "amount"), "quantity"),
        (("日期", "date"), "date"),
        (("产品形式", "产品类型", "剂型", "形式", "product type", "product_type"), "product_type"),
        (("产品名称", "品名", "产品", "product name", "product"), "product_name"),
        (("包装要求", "包材", "包装", "package", "packaging"), "packaging"),
        (("包装", "包材", "package", "packaging"), "packaging"),
        (("标签", "label", "labeling"), "labeling"),
        (("批号", "批次", "batch"), "batch_code"),
        (("备注", "其他要求", "说明", "note", "other requirements"), "other_requirements"),
    ]
    for keywords, field_key in keyword_map:
        if any(keyword in text for keyword in keywords):
            return field_key
    if semantic_type in {"option_group", "option_item"}:
        point = _cell_point(cell)
        if point:
            return f"semantic_option_row_{point[0]}"
    suffix = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    return suffix or f"semantic_{_cell_key(cell).lower()}"


def _semantic_option_group_for_region(semantic, semantic_index):
    if not isinstance(semantic, dict) or not isinstance(semantic_index, dict):
        return {}
    semantic_id = str(semantic.get("region_id") or "")
    source_cell = _cell_key(semantic.get("source_cell"))
    target_cell = _cell_key(semantic.get("target_cell"))
    cells = {_cell_key(cell) for cell in semantic.get("cells", []) if cell} if isinstance(semantic.get("cells"), list) else set()
    cells.update(cell for cell in (source_cell, target_cell) if cell)
    for group in semantic_index.get("option_groups", []):
        if not isinstance(group, dict):
            continue
        if semantic_id and semantic_id == str(group.get("region_id") or ""):
            return group
        group_cells = {_cell_key(cell) for cell in group.get("cells", []) if cell} if isinstance(group.get("cells"), list) else set()
        group_cells.update(cell for cell in (_cell_key(group.get("source_cell")), _cell_key(group.get("target_cell"))) if cell)
        if cells & group_cells:
            return group
    return {}


def _semantic_candidate_rank(candidate):
    intent_type = str(candidate.get("intent_type") or "")
    write_mode = str(candidate.get("write_mode") or "")
    semantic_type = str(candidate.get("semantic_type") or "")
    if write_mode in {"skip", "none"} or intent_type in {"title", "section_header"}:
        if semantic_type in {"title", "section_header"}:
            return 20
        if semantic_type in {"note_instruction", "image_attachment_area"}:
            return 30
    if intent_type in {"option_checkbox", "option_text_choice"} or semantic_type in {"option_group", "option_item"}:
        return 80
    if intent_type in {"table_column_header", "table_row_field"} or semantic_type in {"table_header", "table_row_field", "table_region"}:
        return 70
    if write_mode not in {"skip", "none"} and intent_type not in {"unknown", ""}:
        return 90
    if semantic_type == "unknown" or intent_type == "unknown":
        return 10
    return 50


def _promote_candidate_with_semantic(candidate, intent, semantic, semantic_index, label_text, section_name, cell):
    if not semantic:
        return candidate, intent, {}
    semantic_type = str(semantic.get("type") or "").strip()
    semantic_label = str(semantic.get("label") or label_text or "").strip()
    semantic_confidence = float(semantic.get("confidence") or 0)
    source_cell = _cell_key(semantic.get("source_cell") or cell)
    if semantic_type == "option_group":
        source_cell = _cell_key(cell)
    target_cell = _cell_key(semantic.get("target_cell"))
    group = _semantic_option_group_for_region(semantic, semantic_index)
    group_label = str(group.get("label") or "").strip()
    field_key_label = group_label if semantic_type in {"option_group", "option_item"} and group_label else semantic_label
    field_key = _slugify_semantic_field_key(field_key_label, semantic_type, source_cell or cell, section_name)
    promoted = {
        "semantic_type": semantic_type,
        "semantic_region_id": str(semantic.get("region_id") or ""),
        "semantic_confidence": semantic_confidence,
        "semantic_reason": str(semantic.get("reason") or ""),
        "semantic_promoted": True,
    }

    if semantic_type == "field_label":
        candidate = {
            **candidate,
            "field_key": candidate.get("field_key") or field_key,
            "field_label": semantic_label or candidate.get("field_label", ""),
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence, 0.78),
            "ai_extract_hint": candidate.get("ai_extract_hint") or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
        }
        intent = {
            **intent,
            "intent_type": semantic.get("intent_type") or "label_fill_right",
            "write_mode": semantic.get("write_mode") or "write_right_cell",
            "label_cell": source_cell or cell,
            "target_cell": target_cell or intent.get("target_cell", ""),
            "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence, 0.78),
            "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
        }
    elif semantic_type == "inline_field":
        candidate = {
            **candidate,
            "field_key": candidate.get("field_key") or field_key,
            "field_label": semantic_label or candidate.get("field_label", ""),
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence, 0.80),
            "ai_extract_hint": candidate.get("ai_extract_hint") or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
        }
        intent = {
            **intent,
            "intent_type": "inline_fill_after_colon",
            "write_mode": "append_after_colon",
            "label_cell": source_cell or cell,
            "target_cell": source_cell or cell,
            "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence, 0.80),
            "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
        }
    elif semantic_type in {"option_group", "option_item"}:
        option_value = (semantic_label if semantic_type == "option_item" else str(label_text or "").strip()) or semantic_label
        group_field_key = _slugify_semantic_field_key(field_key_label or option_value, semantic_type, source_cell or cell, section_name)
        candidate = {
            **candidate,
            "field_key": group_field_key,
            "field_label": group_label or candidate.get("field_label") or section_name or "选择字段",
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence, 0.76),
            "ai_extract_hint": candidate.get("ai_extract_hint") or group_label or section_name or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
        }
        intent = {
            **intent,
            "intent_type": semantic.get("intent_type") or "option_text_choice",
            "write_mode": semantic.get("write_mode") or "select_option_text",
            "label_cell": source_cell or cell,
            "target_cell": target_cell or source_cell or cell,
            "option_value": option_value,
            "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence, 0.76),
            "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
        }
    elif semantic_type in {"note_instruction", "image_attachment_area"}:
        intent_type = "attachment_hint" if semantic_type == "image_attachment_area" else "note_instruction"
        candidate = {
            **candidate,
            "field_label": semantic_label or candidate.get("field_label", ""),
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence),
            "ai_extract_hint": candidate.get("ai_extract_hint") or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
            "show_in_workspace": False,
        }
        intent = {
            **intent,
            "intent_type": intent_type,
            "write_mode": "skip",
            "label_cell": source_cell or cell,
            "target_cell": "",
            "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence),
            "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
        }
    elif semantic_type in {"table_header", "table_row_field", "table_region"}:
        candidate = {
            **candidate,
            "field_key": candidate.get("field_key") or field_key,
            "field_label": semantic_label or candidate.get("field_label", ""),
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence, 0.70),
            "ai_extract_hint": candidate.get("ai_extract_hint") or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
        }
        intent = {
            **intent,
            "intent_type": semantic.get("intent_type") or ("table_row_field" if semantic_type == "table_row_field" else "table_column_header"),
            "write_mode": semantic.get("write_mode") or ("write_row_field" if semantic_type == "table_row_field" else "write_table_column"),
            "label_cell": source_cell or cell,
            "target_cell": target_cell or intent.get("target_cell", ""),
            "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence, 0.70),
            "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
        }
    elif semantic_type in {"title", "section_header"}:
        has_writable_catalog_intent = (
            semantic_type == "section_header"
            and bool(candidate.get("field_key"))
            and str(intent.get("write_mode") or "").strip() in {
                "write_right_cell",
                "write_below_cell",
                "append_after_colon",
                "write_table_column",
                "write_row_field",
                "write_table_cell",
            }
            and bool(_cell_key(intent.get("target_cell")))
        )
        candidate = {
            **candidate,
            "field_label": semantic_label or candidate.get("field_label", ""),
            "confidence": max(float(candidate.get("confidence") or 0), semantic_confidence),
            "ai_extract_hint": candidate.get("ai_extract_hint") or semantic_label,
            "candidate_reason": str(semantic.get("reason") or candidate.get("candidate_reason") or ""),
        }
        if not has_writable_catalog_intent:
            candidate["show_in_workspace"] = False
            intent = {
                **intent,
                "intent_type": semantic_type,
                "write_mode": "skip",
                "label_cell": source_cell or cell,
                "target_cell": "",
                "intent_confidence": max(float(intent.get("intent_confidence") or 0), semantic_confidence),
                "intent_reason": str(semantic.get("reason") or intent.get("intent_reason") or ""),
            }
    else:
        promoted["semantic_promoted"] = False

    breakdown = candidate.get("confidence_breakdown") if isinstance(candidate.get("confidence_breakdown"), dict) else {}
    if semantic_confidence:
        breakdown = {**breakdown, "semantic": round(semantic_confidence, 2)}
    candidate = {
        **candidate,
        "confidence": round(float(candidate.get("confidence") or 0), 2),
        "confidence_breakdown": breakdown,
        "source": "semantic_regions+template_analysis",
    }
    return candidate, intent, promoted


def _neighbor_label_text(labels, index, cell):
    point = _cell_point(cell)
    texts = []
    for offset in (-2, -1, 1, 2):
        neighbor_index = index + offset
        if 0 <= neighbor_index < len(labels):
            neighbor = labels[neighbor_index] if isinstance(labels[neighbor_index], dict) else {}
            neighbor_text = str(neighbor.get("value") or neighbor.get("name") or "").strip()
            if neighbor_text:
                texts.append(neighbor_text)

    if point:
        row, col = point
        for neighbor in labels:
            if not isinstance(neighbor, dict):
                continue
            neighbor_cell = str(neighbor.get("cell") or "").strip().upper()
            neighbor_point = _cell_point(neighbor_cell)
            if not neighbor_point or neighbor_cell == cell:
                continue
            neighbor_row, neighbor_col = neighbor_point
            if abs(neighbor_row - row) <= 1 and abs(neighbor_col - col) <= 2:
                neighbor_text = str(neighbor.get("value") or neighbor.get("name") or "").strip()
                if neighbor_text:
                    texts.append(neighbor_text)
    return " ".join(texts)


SEMANTIC_REGION_PRIORITY = {
    "field_label": 90,
    "inline_field": 88,
    "option_group": 84,
    "option_item": 82,
    "table_row_field": 78,
    "table_header": 76,
    "table_region": 74,
    "note_instruction": 62,
    "image_attachment_area": 60,
    "title": 52,
    "section_header": 50,
    "unknown": 0,
}


def _semantic_region_priority(region):
    if not isinstance(region, dict):
        return (0, 0)
    return (
        SEMANTIC_REGION_PRIORITY.get(str(region.get("type") or ""), 0),
        float(region.get("confidence") or 0),
    )


def _build_semantic_region_index(template_analysis):
    analysis = template_analysis if isinstance(template_analysis, dict) else {}
    regions = analysis.get("semantic_regions") if isinstance(analysis.get("semantic_regions"), list) else []
    by_source_cell = {}
    by_target_cell = {}
    by_cell = {}
    by_region_id = {}
    option_groups = []

    def add(bucket, cell, region):
        cell_key = str(cell or "").strip().upper()
        if not cell_key:
            return
        bucket.setdefault(cell_key, []).append(region)

    for region in regions:
        if not isinstance(region, dict):
            continue
        region_id = str(region.get("region_id") or "").strip()
        if region_id:
            by_region_id[region_id] = region
        if str(region.get("type") or "") == "option_group":
            option_groups.append(region)

        source_cell = str(region.get("source_cell") or "").strip().upper()
        target_cell = str(region.get("target_cell") or "").strip().upper()
        add(by_source_cell, source_cell, region)
        add(by_target_cell, target_cell, region)
        for cell in (source_cell, target_cell):
            add(by_cell, cell, region)
        if isinstance(region.get("cells"), list):
            for cell in region.get("cells", []):
                add(by_cell, cell, region)

    for bucket in (by_source_cell, by_target_cell, by_cell):
        for cell, items in bucket.items():
            bucket[cell] = sorted(
                [item for item in items if isinstance(item, dict)],
                key=_semantic_region_priority,
                reverse=True,
            )
    option_groups.sort(key=_semantic_region_priority, reverse=True)
    return {
        "by_source_cell": by_source_cell,
        "by_target_cell": by_target_cell,
        "by_cell": by_cell,
        "by_region_id": by_region_id,
        "option_groups": option_groups,
    }


def _semantic_by_cell_from_analysis(template_analysis):
    return _build_semantic_region_index(template_analysis).get("by_cell", {})
    return by_cell


def _primary_semantic_for_cell(semantic_by_cell, cell):
    regions = semantic_by_cell.get(str(cell or "").strip().upper(), []) if isinstance(semantic_by_cell, dict) else []
    if not regions:
        return {}
    return sorted(
        [region for region in regions if isinstance(region, dict)],
        key=_semantic_region_priority,
        reverse=True,
    )[0] if regions else {}


def _catalog_rule_for_field_key(field_key):
    field_key = str(field_key or "").strip()
    if not field_key:
        return {}
    for rule in get_field_catalog_candidate_rules():
        if str(rule.get("field_key") or "").strip() == field_key:
            return rule
    return {}


def _field_catalog_override_key_for_text(text):
    normalized = _normalize_candidate_text(text)
    if not normalized:
        return ""
    overrides = [
        (
            [
                "\u5305\u88c5\u6570\u91cf",
                "\u5305\u88c5\u89c4\u683c",
                "\u6bcf\u74f6",
                "\u6bcf\u888b",
                "\u7c92\u74f6",
                "\u7c92\u6bcf\u74f6",
            ],
            "packaging.quantity_per_unit",
        ),
        (
            ["\u74f6\u53e3\u5bc6\u5c01"],
            "packaging.bottle_seal_method",
        ),
        (
            ["\u6279\u53f7\u65e5\u671f", "\u6279\u53f7"],
            "batch_marking.requirement",
        ),
        (
            ["\u4ea7\u54c1\u63cf\u8ff0"],
            "product_name",
        ),
    ]
    for keywords, field_key in overrides:
        if any(_normalize_candidate_text(keyword) in normalized for keyword in keywords):
            return field_key
    return ""


def _field_catalog_candidate_for_text(text, section, label_index, total_labels, cell, neighbor_text=""):
    candidate = _candidate_for_label(text, section, label_index, total_labels, cell, neighbor_text)
    override_key = _field_catalog_override_key_for_text(text)
    if override_key:
        rule = _catalog_rule_for_field_key(override_key)
        if rule:
            confidence = max(float(candidate.get("confidence") or 0), 0.72)
            candidate = {
                **candidate,
                "field_key": rule.get("field_key") or override_key,
                "field_label": rule.get("field_label") or candidate.get("field_label", ""),
                "confidence": round(confidence, 2),
                "ai_extract_hint": rule.get("ai_extract_hint") or candidate.get("ai_extract_hint", ""),
                "candidate_reason": "field_catalog_override",
                "confidence_breakdown": {
                    **(candidate.get("confidence_breakdown") if isinstance(candidate.get("confidence_breakdown"), dict) else {}),
                    "catalog_override": 0.72,
                },
                "priority": int(rule.get("priority") or candidate.get("priority") or 0),
                "source": "field_catalog",
            }
    return candidate


def _split_field_catalog_candidate_lines(text):
    parts = []
    for part in re.split(r"[\r\n]+", str(text or "")):
        cleaned = part.strip()
        if cleaned:
            parts.append(cleaned)
    return parts


def _field_catalog_expansion_target(context, cell, text):
    if "\uff1a" in str(text or "") or ":" in str(text or ""):
        return cell, "append_after_colon", "inline_fill_after_colon"
    point = _cell_point(cell)
    if point:
        row, _ = point
        right = _infer_right_target_cell(context, cell)
        if right:
            return right, "write_right_cell", "label_fill_right"
        below = _infer_below_target_cell(context, cell)
        if below and row <= 8:
            return below, "write_below_cell", "label_fill_below"
    return cell, "append_after_colon", "inline_fill_after_colon"


def _generate_field_catalog_expansion_candidates(template_analysis, layout_sections, template_path, existing_candidates):
    analysis = template_analysis if isinstance(template_analysis, dict) else {}
    labels = analysis.get("labels") if isinstance(analysis.get("labels"), list) else []
    if not labels:
        return []
    intent_context = _load_template_intent_context(template_path)
    used_field_keys = {
        str(item.get("field_key") or "").strip()
        for item in existing_candidates if isinstance(item, dict)
        and str(item.get("field_key") or "").strip()
        and str(item.get("write_mode") or "").strip() not in {"skip", "none", ""}
    }
    used_cells = {
        str(item.get("cell") or "").strip().upper()
        for item in existing_candidates if isinstance(item, dict) and item.get("cell")
    }
    expansion = []
    synthetic_index = 0
    total_labels = len(labels)

    for label_index, label in enumerate(labels, 1):
        if not isinstance(label, dict):
            continue
        cell = str(label.get("cell") or "").strip().upper()
        label_text = str(label.get("value") or label.get("name") or "").strip()
        if not cell or not label_text:
            continue
        lines = _split_field_catalog_candidate_lines(label_text)
        if len(lines) <= 1 and "\uff1a" not in label_text and ":" not in label_text:
            preview = _field_catalog_candidate_for_text(label_text, _section_for_cell(layout_sections, cell), label_index, total_labels, cell)
            preview_key = str(preview.get("field_key") or "").strip()
            if _field_catalog_override_key_for_text(label_text) != "product_name" and preview_key not in {"customer_name", "order_date", "quantity"}:
                continue
            lines = [label_text]
        section = _section_for_cell(layout_sections, cell)
        section_name = (
            section.get("title")
            or section.get("source_region_name")
            or section.get("section_key")
            or ""
        ) if isinstance(section, dict) else ""
        for line_index, line in enumerate(lines, 1):
            candidate = _field_catalog_candidate_for_text(line, section, label_index, total_labels, cell)
            field_key = str(candidate.get("field_key") or "").strip()
            if not field_key or field_key in {"packaging"}:
                continue
            if field_key in used_field_keys:
                continue
            confidence = max(float(candidate.get("confidence") or 0), 0.72)
            if confidence < 0.70:
                continue
            target_cell, write_mode, intent_type = _field_catalog_expansion_target(intent_context, cell, line)
            synthetic_index += 1
            synthetic_cell = f"{cell}__FC{synthetic_index:02d}"
            while synthetic_cell in used_cells:
                synthetic_index += 1
                synthetic_cell = f"{cell}__FC{synthetic_index:02d}"
            used_cells.add(synthetic_cell)
            used_field_keys.add(field_key)
            expansion.append(
                {
                    "cell": synthetic_cell,
                    "field_key": field_key,
                    "field_label": candidate.get("field_label") or line,
                    "section": section_name,
                    "section_key": section.get("section_key", "") if isinstance(section, dict) else "",
                    "confidence": round(confidence, 2),
                    "source": "field_catalog_v2_expansion",
                    "ai_extract_hint": candidate.get("ai_extract_hint") or line,
                    "candidate_reason": candidate.get("candidate_reason") or "field_catalog_line",
                    "confidence_breakdown": candidate.get("confidence_breakdown") if isinstance(candidate.get("confidence_breakdown"), dict) else {},
                    "intent_type": intent_type,
                    "write_mode": write_mode,
                    "label_cell": cell,
                    "target_cell": target_cell,
                    "option_value": "",
                    "intent_confidence": 0.80,
                    "intent_reason": "field_catalog_v2_expansion",
                    "semantic_type": "field_catalog_line",
                    "semantic_region_id": "",
                    "semantic_confidence": round(confidence, 2),
                    "semantic_reason": "field_catalog_v2_expansion",
                    "semantic_promoted": True,
                    "show_in_workspace": True,
                    "label_text": line,
                    "display_order": (label_index * 100) + line_index,
                }
            )

    return expansion


def _generate_mapping_candidates(template_analysis, layout_sections, template_path=None):
    analysis = template_analysis if isinstance(template_analysis, dict) else {}
    labels = analysis.get("labels") if isinstance(analysis.get("labels"), list) else []
    semantic_index = _build_semantic_region_index(analysis)
    semantic_by_cell = semantic_index.get("by_cell", {})
    intent_context = _load_template_intent_context(template_path)
    candidates = []
    seen_cells = set()
    for index, label in enumerate(labels, 1):
        if not isinstance(label, dict):
            continue
        cell = str(label.get("cell") or "").strip().upper()
        label_text = str(label.get("value") or label.get("name") or "").strip()
        if not cell or not label_text or cell in seen_cells:
            continue
        seen_cells.add(cell)
        section = _section_for_cell(layout_sections, cell)
        neighbor_text = _neighbor_label_text(labels, index - 1, cell)
        candidate = _field_catalog_candidate_for_text(label_text, section, index, len(labels), cell, neighbor_text)
        intent = _intent_for_label(label_text, cell, section, index, intent_context)
        semantic = _primary_semantic_for_cell(semantic_by_cell, cell)
        section_name = (
            section.get("title")
            or section.get("source_region_name")
            or section.get("section_key")
            or ""
        ) if isinstance(section, dict) else ""
        candidate, intent, semantic_fields = _promote_candidate_with_semantic(
            candidate,
            intent,
            semantic,
            semantic_index,
            label_text,
            section_name,
            cell,
        )
        candidates.append(
            {
                "cell": cell,
                "field_key": candidate["field_key"],
                "field_label": candidate["field_label"],
                "section": section_name,
                "section_key": section.get("section_key", "") if isinstance(section, dict) else "",
                "confidence": candidate["confidence"],
                "source": candidate.get("source") or "template_analysis+layout_sections",
                "ai_extract_hint": candidate["ai_extract_hint"],
                "candidate_reason": candidate.get("candidate_reason", ""),
                "confidence_breakdown": candidate.get("confidence_breakdown", {}),
                "intent_type": intent.get("intent_type", "unknown"),
                "write_mode": intent.get("write_mode", "skip"),
                "label_cell": intent.get("label_cell", cell),
                "target_cell": intent.get("target_cell", ""),
                "option_value": intent.get("option_value", ""),
                "intent_confidence": intent.get("intent_confidence", 0),
                "intent_reason": intent.get("intent_reason", ""),
                "semantic_type": semantic_fields.get("semantic_type", semantic.get("type", "") if semantic else ""),
                "semantic_region_id": semantic_fields.get("semantic_region_id", semantic.get("region_id", "") if semantic else ""),
                "semantic_confidence": semantic_fields.get("semantic_confidence", semantic.get("confidence", 0) if semantic else 0),
                "semantic_reason": semantic_fields.get("semantic_reason", semantic.get("reason", "") if semantic else ""),
                "semantic_promoted": bool(semantic_fields.get("semantic_promoted", False)),
                "show_in_workspace": candidate.get("show_in_workspace", True),
                "label_text": label_text,
                "display_order": index,
            }
        )
    candidates.extend(_generate_field_catalog_expansion_candidates(analysis, layout_sections, template_path, candidates))
    candidates.sort(
        key=lambda item: (
            -_semantic_candidate_rank(item),
            -int(bool(item.get("semantic_promoted"))),
            -float(item.get("confidence") or 0),
            int(item.get("display_order") or 0),
        )
    )
    return candidates


def _contract_field_type_from_workspace_field(field):
    explicit_field_type = str(field.get("field_type") or field.get("type") or "").strip()
    if explicit_field_type == "image":
        return "image"
    intent_type = str(field.get("intent_type") or "").strip()
    write_mode = str(field.get("write_mode") or "").strip()
    label_key = f"{field.get('label') or ''} {field.get('field_key') or ''}".lower()
    if intent_type in {"option_checkbox", "option_text_choice"}:
        return "single_choice"
    if "日期" in label_key or "date" in label_key:
        return "date"
    if any(keyword in label_key for keyword in ("数量", "quantity", "qty", "number", "amount")):
        return "number"
    if any(keyword in label_key for keyword in ("备注", "要求", "描述", "说明", "other", "note", "description")):
        return "textarea"
    if write_mode in {"skip", "none"}:
        return "readonly"
    return "text"


def _contract_field_required(field):
    label_key = f"{field.get('label') or ''} {field.get('field_key') or ''}".lower()
    return any(keyword in label_key for keyword in ("客户名称", "数量", "日期", "产品", "product", "customer", "quantity"))


def _build_ai_extraction_contract_from_workspace_fields(workspace_fields):
    fields = []
    option_groups_by_key = {}
    option_field_by_key = {}
    seen_field_keys = set()
    hidden_intents = {"title", "section_header", "note_instruction", "image_area", "attachment_hint", "readonly_example"}
    hidden_write_modes = {"skip", "none"}

    source_fields = workspace_fields if isinstance(workspace_fields, list) else []
    sorted_fields = sorted(
        [field for field in source_fields if isinstance(field, dict)],
        key=lambda field: int(field.get("display_order") or 0) if str(field.get("display_order") or "").isdigit() else 0,
    )

    for field in sorted_fields:
        field_key = str(field.get("field_key") or "").strip()
        if not field_key:
            continue

        intent_type = str(field.get("intent_type") or "").strip()
        write_mode = str(field.get("write_mode") or "").strip()
        explicit_field_type = str(field.get("field_type") or field.get("type") or "").strip()
        is_image_field = explicit_field_type == "image"
        is_table_field = explicit_field_type in {"table", "dynamic_table"} or write_mode == "write_table_cell"
        if not is_image_field and not is_table_field and (intent_type in hidden_intents or write_mode in hidden_write_modes):
            continue

        label = str(field.get("label") or field_key).strip() or field_key
        field_type = _contract_field_type_from_workspace_field(field)
        option_value = str(field.get("option_value") or label).strip()

        if field_type == "single_choice":
            group = option_groups_by_key.setdefault(
                field_key,
                {
                    "field_key": field_key,
                    "label": str(field.get("ai_extract_hint") or field.get("section") or label or field_key).strip() or field_key,
                    "options": [],
                },
            )
            if option_value and option_value not in group["options"]:
                group["options"].append(option_value)
            option_field_by_key.setdefault(field_key, field)
            continue

        if field_key in seen_field_keys:
            continue

        fields.append(
            {
                "field_key": field_key,
                "label": label,
                "type": field_type,
                "section": str(field.get("section") or "").strip(),
                "required": _contract_field_required(field),
                "ai_extract_hint": str(field.get("ai_extract_hint") or label).strip(),
                "intent_type": intent_type,
                "write_mode": write_mode,
                "target_cell": _cell_key(field.get("cell")),
                "source_cell": _cell_key(field.get("source_cell")),
                "sheet_name": _sheet_key(field.get("sheet_name") or field.get("target_sheet") or field.get("worksheet") or field.get("sheet")),
                "col_offset": int(field.get("col_offset") or field.get("table_col_offset") or field.get("column_offset") or 0),
                "table_col_offset": int(field.get("table_col_offset") or field.get("col_offset") or field.get("column_offset") or 0),
                "options": [],
            }
        )
        seen_field_keys.add(field_key)

    for field_key, group in option_groups_by_key.items():
        source_field = option_field_by_key.get(field_key, {})
        label = str(group.get("label") or source_field.get("label") or field_key).strip() or field_key
        fields.append(
            {
                "field_key": field_key,
                "label": label,
                "type": "single_choice",
                "section": str(source_field.get("section") or "").strip(),
                "required": _contract_field_required(source_field),
                "ai_extract_hint": str(source_field.get("ai_extract_hint") or label).strip(),
                "intent_type": str(source_field.get("intent_type") or "").strip(),
                "write_mode": str(source_field.get("write_mode") or "").strip(),
                "target_cell": _cell_key(source_field.get("cell")),
                "source_cell": _cell_key(source_field.get("source_cell")),
                "sheet_name": _sheet_key(source_field.get("sheet_name") or source_field.get("target_sheet") or source_field.get("worksheet") or source_field.get("sheet")),
                "col_offset": int(source_field.get("col_offset") or source_field.get("table_col_offset") or source_field.get("column_offset") or 0),
                "table_col_offset": int(source_field.get("table_col_offset") or source_field.get("col_offset") or source_field.get("column_offset") or 0),
                "options": group["options"],
            }
        )

    fields.sort(key=lambda field: (str(field.get("section") or ""), str(field.get("field_key") or "")))
    option_groups = list(option_groups_by_key.values())
    return {
        "fields": fields,
        "option_groups": option_groups,
    }


def _build_ai_extraction_contract_from_profile(profile):
    return _build_ai_extraction_contract_from_workspace_fields(_build_workspace_fields_from_profile(profile))


def _ai_extraction_contract_summary(contract):
    contract = contract if isinstance(contract, dict) else {}
    fields = contract.get("fields") if isinstance(contract.get("fields"), list) else []
    option_groups = contract.get("option_groups") if isinstance(contract.get("option_groups"), list) else []
    return {
        "fields_count": len(fields),
        "option_groups_count": len(option_groups),
        "field_keys": [str(field.get("field_key") or "").strip() for field in fields if isinstance(field, dict) and str(field.get("field_key") or "").strip()],
    }


SEMANTIC_WORKSPACE_SECTIONS = [
    {"key": "document_info", "title": "文档编号信息", "order": 10},
    {"key": "basic_info", "title": "基础订单信息", "order": 20},
    {"key": "product_info", "title": "产品信息", "order": 30},
    {"key": "specification", "title": "产品规格", "order": 40},
    {"key": "packaging", "title": "包装要求", "order": 50},
    {"key": "labeling", "title": "标签要求", "order": 60},
    {"key": "attachments", "title": "附件资料", "order": 70},
    {"key": "other", "title": "其他要求", "order": 80},
    {"key": "debug_other", "title": "其他字段", "order": 90},
]
SEMANTIC_WORKSPACE_SECTION_BY_KEY = {item["key"]: item for item in SEMANTIC_WORKSPACE_SECTIONS}


def _semantic_workspace_analysis_for_profile(profile):
    profile = profile if isinstance(profile, dict) else {}
    profile_id = str(profile.get("profile_id") or "").strip()
    state = get_pipeline_state()
    state_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    state_analysis = state.get("template_analysis") if isinstance(state.get("template_analysis"), dict) else {}
    if state_profile.get("profile_id") == profile_id and isinstance(state_analysis.get("semantic_regions"), list):
        return state_analysis

    template_file_path = str(profile.get("template_file_path") or "").strip()
    if not template_file_path:
        return {}
    try:
        from app.v4_template_analysis import analyze_template

        return analyze_template(_resolve_bound_template_file_path(template_file_path))
    except Exception:
        logger.info("V4 semantic workspace analysis unavailable: profile_id=%s", profile_id, exc_info=True)
        return {}


def _semantic_workspace_cell_index(workspace_fields):
    by_source = {}
    by_target = {}
    by_field_key = {}
    for field in workspace_fields if isinstance(workspace_fields, list) else []:
        if not isinstance(field, dict):
            continue
        source_cell = _cell_key(field.get("source_cell"))
        target_cell = _cell_key(field.get("cell"))
        field_key = str(field.get("field_key") or "").strip()
        if source_cell:
            by_source.setdefault(source_cell, field)
        if target_cell:
            by_target.setdefault(target_cell, field)
        if field_key:
            by_field_key.setdefault(field_key, field)
    return {"by_source": by_source, "by_target": by_target, "by_field_key": by_field_key}


def _semantic_workspace_match_field(region, index):
    source_cell = _cell_key(region.get("source_cell"))
    target_cell = _cell_key(region.get("target_cell"))
    for cell in (source_cell, target_cell):
        if cell and cell in index["by_source"]:
            return index["by_source"][cell]
        if cell and cell in index["by_target"]:
            return index["by_target"][cell]
    return {}


def _semantic_workspace_field_key(label, source_cell):
    text = str(label or "").strip().lower()
    if "客户" in text or "customer" in text:
        return "customer_name"
    if "数量" in text or "quantity" in text or "qty" in text:
        return "quantity"
    if "日期" in text or "date" in text:
        return "date"
    if "产品" in text or "品名" in text or "product" in text:
        return "product_name"
    suffix = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    return suffix or f"semantic_{_cell_key(source_cell).lower()}"


def _semantic_workspace_type(region, field=None):
    field = field if isinstance(field, dict) else {}
    semantic_type = str(region.get("type") or "").strip()
    intent_type = str(region.get("intent_type") or field.get("intent_type") or "").strip()
    label_key = f"{region.get('label') or ''} {field.get('label') or ''} {field.get('field_key') or ''}".lower()
    if any(keyword in label_key for keyword in ("日期", "date")):
        return "date"
    if any(keyword in label_key for keyword in ("数量", "quantity", "qty", "amount", "number")):
        return "number"
    if semantic_type in {"option_group", "option_item"} or intent_type in {"option_checkbox", "option_text_choice"}:
        return "single_choice"
    if semantic_type in {"note_instruction"} or intent_type == "readonly_example":
        return "readonly_note"
    if semantic_type == "image_attachment_area" or intent_type in {"attachment_hint", "image_area"}:
        return "image_upload"
    if semantic_type == "inline_field" or intent_type == "inline_fill_after_colon":
        if any(keyword in label_key for keyword in ("要求", "描述", "备注", "说明", "other", "note", "description")):
            return "textarea"
        return "text"
    if semantic_type in {"table_region", "table_header", "table_row_field"} or intent_type in {"table_row_field", "table_column_header"}:
        return "text"
    if semantic_type == "unknown":
        return "text"
    return "text"


def _semantic_workspace_section_from_text(text):
    value = str(text or "").lower()
    if not value:
        return None
    if any(keyword in value for keyword in ("文档", "编号", "document")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["document_info"]
    if any(keyword in value for keyword in ("客户", "基础", "负责人", "customer", "basic")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["basic_info"]
    if any(keyword in value for keyword in ("产品", "品名", "product")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["product_info"]
    if any(keyword in value for keyword in ("规格", "数量", "含量", "配方", "spec", "quantity")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["specification"]
    if any(keyword in value for keyword in ("包装", "瓶", "袋", "package")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["packaging"]
    if any(keyword in value for keyword in ("标签", "批号", "label")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["labeling"]
    if any(keyword in value for keyword in ("附件", "图片", "照片", "attachment", "image", "photo")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["attachments"]
    if any(keyword in value for keyword in ("备注", "要求", "说明", "note", "other")):
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["other"]
    return None


def _semantic_workspace_section(region, field, section_headers):
    field_section = _semantic_workspace_section_from_text(field.get("section") if isinstance(field, dict) else "")
    if field_section:
        return field_section
    row = int(region.get("row") or 0)
    prior_headers = [header for header in section_headers if int(header.get("row") or 0) <= row]
    if prior_headers:
        header_match = _semantic_workspace_section_from_text(prior_headers[-1].get("label"))
        if header_match:
            return header_match
    label_match = _semantic_workspace_section_from_text(region.get("label"))
    if label_match:
        return label_match
    if str(region.get("type") or "") == "image_attachment_area":
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["attachments"]
    if str(region.get("type") or "") == "unknown":
        return SEMANTIC_WORKSPACE_SECTION_BY_KEY["debug_other"]
    return SEMANTIC_WORKSPACE_SECTION_BY_KEY["other"]


def _semantic_workspace_required(field, label):
    if isinstance(field, dict) and field.get("required") is True:
        return True
    text = f"{label or ''} {field.get('field_key') if isinstance(field, dict) else ''}".lower()
    return any(keyword in text for keyword in ("客户名称", "数量", "日期", "产品", "product", "customer", "quantity"))


def _semantic_workspace_option_group_key(region, field, section):
    if region.get("type") == "option_group":
        return str(region.get("region_id") or "")
    field_key = str(field.get("field_key") or "").strip() if isinstance(field, dict) else ""
    if field_key:
        return f"field:{field_key}"
    return f"row:{section['key']}:{region.get('row') or 0}"


def _build_semantic_workspace_schema(profile):
    profile = profile if isinstance(profile, dict) else {}
    analysis = _semantic_workspace_analysis_for_profile(profile)
    semantic_regions = analysis.get("semantic_regions") if isinstance(analysis.get("semantic_regions"), list) else []
    workspace_fields = _build_workspace_fields_from_profile(profile)
    field_index = _semantic_workspace_cell_index(workspace_fields)
    section_headers = [region for region in semantic_regions if isinstance(region, dict) and region.get("type") == "section_header"]
    section_headers.sort(key=lambda item: (int(item.get("row") or 0), int(item.get("col") or 0)))
    fields = []
    option_groups = {}

    for region in semantic_regions:
        if not isinstance(region, dict):
            continue
        semantic_type = str(region.get("type") or "")
        if semantic_type in {"title", "section_header"}:
            continue
        field = _semantic_workspace_match_field(region, field_index)
        section = _semantic_workspace_section(region, field, section_headers)
        field_type = _semantic_workspace_type(region, field)
        source_cell = _cell_key(region.get("source_cell") or field.get("source_cell"))
        target_cell = _cell_key(region.get("target_cell") or field.get("cell") or source_cell)
        label = str(field.get("label") or region.get("label") or field.get("field_key") or source_cell).strip()
        field_key = str(field.get("field_key") or _semantic_workspace_field_key(label, source_cell)).strip()
        write_mode = str(region.get("write_mode") or field.get("write_mode") or "skip").strip()
        intent_type = str(region.get("intent_type") or field.get("intent_type") or "unknown").strip()
        base = {
            "id": f"{field_key}_{target_cell or source_cell}",
            "type": field_type,
            "label": label,
            "field_key": field_key,
            "cell": target_cell,
            "source_cell": source_cell,
            "intent_type": intent_type,
            "write_mode": write_mode,
            "required": _semantic_workspace_required(field, label),
            "options": [],
            "section_key": section["key"],
            "section_title": section["title"],
            "section_order": section["order"],
            "semantic_type": semantic_type,
            "semantic_region_id": str(region.get("region_id") or ""),
            "semantic_confidence": float(region.get("confidence") or 0),
            "semantic_reason": str(region.get("reason") or ""),
            "ai_extract_hint": str(field.get("ai_extract_hint") or label).strip(),
            "display_order": int(region.get("row") or 0) * 100 + int(region.get("col") or 0),
        }

        if field_type == "single_choice":
            group_key = _semantic_workspace_option_group_key(region, field, section)
            group = option_groups.setdefault(
                group_key,
                {
                    **base,
                    "id": group_key,
                    "label": str(field.get("ai_extract_hint") or section["title"] if semantic_type == "option_item" else label).strip() or label,
                    "options": [],
                },
            )
            option_value = str(field.get("option_value") or region.get("label") or label).strip()
            option = {
                "label": option_value,
                "value": option_value,
                "cell": target_cell or source_cell,
                "display_cell": target_cell or source_cell,
                "source_cell": source_cell,
                "field_key": field_key,
                "write_mode": write_mode,
                "intent_type": intent_type,
                "option_value": option_value,
                "semantic_region_id": str(region.get("region_id") or ""),
            }
            if option_value and option_value not in {item.get("value") for item in group["options"]}:
                group["options"].append(option)
            continue

        fields.append(base)

    fields.extend(option_groups.values())
    sections_by_key = {}
    for field in fields:
        key = field.get("section_key") or "other"
        section = sections_by_key.setdefault(
            key,
            {
                "key": key,
                "title": field.get("section_title") or SEMANTIC_WORKSPACE_SECTION_BY_KEY.get(key, {}).get("title", "其他要求"),
                "order": int(field.get("section_order") or 999),
                "fields": [],
            },
        )
        section["fields"].append(field)

    for section in sections_by_key.values():
        section["fields"].sort(key=lambda item: (int(item.get("display_order") or 0), str(item.get("label") or "")))
    sections = sorted(sections_by_key.values(), key=lambda item: (int(item.get("order") or 999), str(item.get("title") or "")))
    fallback_used = not bool(sections)
    return {
        "source": "semantic_regions" if not fallback_used else "workspace_fields",
        "sections": sections,
        "summary": {
            "sections_count": len(sections),
            "fields_count": sum(len(section.get("fields", [])) for section in sections),
            "source": "semantic_regions" if not fallback_used else "workspace_fields",
        },
        "fallback_used": fallback_used,
    }


def _count_runtime_saved_configuration_fields(profile):
    configuration = _template_configuration_from_profile(profile)
    count = 0
    for item in configuration.values():
        if not isinstance(item, dict):
            continue
        show_in_workspace = item.get("show_in_workspace") is not False
        write_mode = str(item.get("write_mode") or "").strip()
        if show_in_workspace and write_mode not in {"skip", "none"}:
            count += 1
    return count


def _count_semantic_workspace_schema_fields(schema):
    schema = schema if isinstance(schema, dict) else {}
    summary = schema.get("summary") if isinstance(schema.get("summary"), dict) else {}
    fields_count = summary.get("fields_count")
    if isinstance(fields_count, int):
        return fields_count
    sections = schema.get("sections") if isinstance(schema.get("sections"), list) else []
    return sum(len(section.get("fields", [])) for section in sections if isinstance(section, dict) and isinstance(section.get("fields"), list))


def _get_runtime_mapping_source(profile):
    saved_fields_count = _count_runtime_saved_configuration_fields(profile)
    semantic_workspace_schema = _build_semantic_workspace_schema(profile)
    semantic_fields_count = _count_semantic_workspace_schema_fields(semantic_workspace_schema)
    if saved_fields_count > 0:
        source = "saved_configuration"
    elif semantic_fields_count > 0:
        source = "semantic_fallback"
    else:
        source = "empty"
    return {
        "source": source,
        "saved_fields_count": saved_fields_count,
        "semantic_fields_count": semantic_fields_count,
        "using_saved_configuration": source == "saved_configuration",
    }



def _build_runtime_mapping_trace_report(profile):
    profile = profile if isinstance(profile, dict) else {}
    profile_id = str(profile.get("profile_id") or "").strip()
    configuration = _template_configuration_from_profile(profile)
    workspace_fields = _build_workspace_fields_from_profile(profile)
    ai_contract = _build_ai_extraction_contract_from_workspace_fields(workspace_fields)
    ai_fields = ai_contract.get("fields") if isinstance(ai_contract.get("fields"), list) else []
    runtime_mapping_source = _get_runtime_mapping_source(profile)

    workspace_keys = {
        str(item.get("field_key") or "").strip()
        for item in workspace_fields
        if isinstance(item, dict) and str(item.get("field_key") or "").strip()
    }
    workspace_cells = {
        _cell_key(item.get("cell") or item.get("target_cell"))
        for item in workspace_fields
        if isinstance(item, dict) and _cell_key(item.get("cell") or item.get("target_cell"))
    }
    ai_keys = {
        str(item.get("field_key") or "").strip()
        for item in ai_fields
        if isinstance(item, dict) and str(item.get("field_key") or "").strip()
    }

    export_write_modes = {
        "write_right_cell",
        "write_below_cell",
        "append_after_colon",
        "check_option",
        "select_option_text",
        "write_table_column",
        "write_row_field",
        "write_table_cell",
    }
    skipped_write_modes = {"skip", "none"}
    skipped_intents = {"title", "section_header", "note_instruction", "image_area", "attachment_hint", "readonly_example"}

    trace_items = []
    warnings = []
    errors = []

    for source_cell, item in sorted(configuration.items(), key=lambda pair: _cell_key(pair[0])):
        if not isinstance(item, dict):
            continue

        cell = _cell_key(source_cell)
        target_cell = _cell_key(item.get("target_cell"))
        field_key = str(item.get("candidate_field_key") or item.get("field_key") or "").strip()
        label = (
            str(item.get("label") or "").strip()
            or str(item.get("candidate_field_label") or "").strip()
            or str(item.get("field_label") or "").strip()
            or field_key
        )
        write_mode = str(item.get("write_mode") or "").strip()
        intent_type = str(item.get("intent_type") or "").strip()
        show_in_workspace = item.get("show_in_workspace") is not False
        is_skipped = write_mode in skipped_write_modes or intent_type in skipped_intents

        in_workspace = bool(
            not is_skipped
            and show_in_workspace
            and (
                (field_key and field_key in workspace_keys)
                or (target_cell and target_cell in workspace_cells)
                or (cell and cell in workspace_cells)
            )
        )
        in_ai_contract = bool(not is_skipped and field_key and field_key in ai_keys)
        export_ready = bool(not is_skipped and write_mode in export_write_modes and target_cell)

        problems = []
        if show_in_workspace and not field_key and not is_skipped:
            problems.append("显示到工作页但 field_key 为空。")
        if write_mode in export_write_modes and not target_cell:
            problems.append(f"write_mode={write_mode} 需要 target_cell。")
        if field_key and not in_workspace and show_in_workspace and not is_skipped:
            problems.append("已保存配置未进入 workspace_fields。")
        if field_key and not in_ai_contract and not is_skipped:
            problems.append("已保存配置未进入 AI extraction contract。")
        if not export_ready and write_mode in export_write_modes:
            problems.append("当前配置尚不可导出。")

        if problems:
            warnings.append({
                "cell": cell,
                "target_cell": target_cell,
                "field_key": field_key,
                "label": label,
                "problems": problems,
            })

        trace_items.append({
            "cell": cell,
            "target_cell": target_cell,
            "field_key": field_key,
            "label": label,
            "show_in_workspace": show_in_workspace,
            "write_mode": write_mode,
            "intent_type": intent_type,
            "in_workspace": in_workspace,
            "in_ai_contract": in_ai_contract,
            "export_ready": export_ready,
            "source": "saved_configuration",
            "problems": problems,
        })

    summary = {
        "saved_configuration_count": len(configuration),
        "workspace_fields_count": len(workspace_fields),
        "ai_contract_fields_count": len(ai_fields),
        "export_ready_count": sum(1 for item in trace_items if item.get("export_ready")),
        "warnings_count": len(warnings),
        "errors_count": len(errors),
    }

    return {
        "success": True,
        "profile_id": profile_id,
        "runtime_mapping_source": runtime_mapping_source,
        "summary": summary,
        "saved_configuration_count": summary["saved_configuration_count"],
        "workspace_fields_count": summary["workspace_fields_count"],
        "ai_contract_fields_count": summary["ai_contract_fields_count"],
        "export_ready_count": summary["export_ready_count"],
        "trace_items": trace_items,
        "warnings": warnings,
        "errors": errors,
    }



def _normalize_audit_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _values_match_for_audit(expected, actual, write_mode=""):
    expected_text = _normalize_audit_value(expected)
    actual_text = _normalize_audit_value(actual)
    mode = str(write_mode or "").strip()

    if not expected_text:
        return not actual_text

    if expected_text == actual_text:
        return True

    if mode == "append_after_colon":
        return actual_text.endswith(expected_text) or expected_text in actual_text

    if mode in {"check_option", "select_option_text"}:
        return expected_text in actual_text or actual_text in expected_text

    return False


def _export_audit_root_cause(item, expected="", actual="", write_mode="", has_mapping_config=True):
    status = str(item.get("status") or "").strip()
    field_key = str(item.get("field_key") or "").strip()
    cell = _cell_key(item.get("cell"))
    expected_text = _normalize_audit_value(expected)
    actual_text = _normalize_audit_value(actual)
    mode = str(write_mode or "").strip()

    if status == "matched":
        return "ok", "已正确导出"

    if status == "skipped":
        return "skipped", "已跳过检查"

    if not has_mapping_config:
        return "mapping_missing", "缺少已保存映射配置"

    if not field_key:
        return "mapping_missing", "缺少 field_key，无法形成稳定链路"

    if not cell:
        return "mapping_missing", "缺少目标单元格"

    if expected_text and not actual_text:
        return "missing_export", "期望值存在，但导出单元格为空"

    if mode in {"append_after_colon", "check_option", "select_option_text", "write_table_column", "write_row_field", "write_table_cell"}:
        return "write_mode_issue", "可能与特殊写入模式有关"

    if expected_text != actual_text:
        return "value_mismatch", "导出值与期望值不一致"

    return "unknown", "未知原因"


def _readback_cell_with_offsets(cell_ref, row_offset=0, col_offset=0):
    match = re.match(r"^([A-Za-z]+)([1-9][0-9]*)$", str(cell_ref or "").strip())
    if not match:
        return _cell_key(cell_ref)

    from openpyxl.utils import column_index_from_string, get_column_letter

    col_letters = match.group(1).upper()
    row = int(match.group(2))
    col = column_index_from_string(col_letters)

    try:
        row_delta = int(row_offset or 0)
    except (TypeError, ValueError):
        row_delta = 0

    try:
        col_delta = int(col_offset or 0)
    except (TypeError, ValueError):
        col_delta = 0

    target_row = max(1, row + row_delta)
    target_col = max(1, col + col_delta)
    return f"{get_column_letter(target_col)}{target_row}"


def _readback_target_cell_for_confirmed_item(cell, merged):
    write_mode = str(merged.get("write_mode") or "").strip()
    field_type = str(merged.get("field_type") or merged.get("type") or "").strip().lower()
    if write_mode != "write_table_cell" and field_type not in {"table", "dynamic_table"}:
        return _cell_key(cell)

    return _readback_cell_with_offsets(
        cell,
        row_offset=merged.get("row_offset"),
        col_offset=merged.get("col_offset") or merged.get("table_col_offset") or merged.get("column_offset"),
    )



def _build_export_readback_audit(exported_file_path, confirmed_cells, profile=None):
    from openpyxl import load_workbook

    profile = profile if isinstance(profile, dict) else {}
    confirmed_cells = confirmed_cells if isinstance(confirmed_cells, list) else []

    summary = {
        "total": 0,
        "checked": 0,
        "matched": 0,
        "mismatched": 0,
        "skipped": 0,
    }
    items = []
    warnings = []
    errors = []

    try:
        workbook = load_workbook(exported_file_path, data_only=False)
        sheet = workbook.active
    except Exception as exc:
        return {
            "success": False,
            "error": f"导出文件回读失败：{exc}",
            "summary": summary,
            "items": [],
            "warnings": [],
            "errors": [{"message": f"导出文件回读失败：{exc}"}],
        }

    lookup = _confirmed_config_lookup_from_profile(profile)

    for raw_item in confirmed_cells:
        if not isinstance(raw_item, dict):
            continue

        summary["total"] += 1
        config = _lookup_confirmed_mapping_config(raw_item, lookup)
        merged = _confirmed_item_with_mapping_config(raw_item, config)

        base_cell = _cell_key(merged.get("cell") or merged.get("target_cell") or merged.get("display_cell"))
        cell = _readback_target_cell_for_confirmed_item(base_cell, merged)
        value = merged.get("value")
        write_mode = str(merged.get("write_mode") or "").strip()
        field_key = str(merged.get("field_key") or "").strip()
        label = str(merged.get("label") or field_key or cell or "").strip()

        if not cell:
            summary["skipped"] += 1
            warnings.append({
                "cell": "",
                "field_key": field_key,
                "label": label,
                "message": "confirmed cell 缺少目标 cell，跳过回读。",
            })
            root_cause, root_cause_label = _export_audit_root_cause(
                {"status": "skipped", "field_key": field_key, "cell": ""},
                expected=value,
                actual="",
                write_mode=write_mode,
                has_mapping_config=bool(config),
            )
            items.append({
                "cell": "",
                "field_key": field_key,
                "label": label,
                "expected": _normalize_audit_value(value),
                "actual": "",
                "write_mode": write_mode,
                "matched": False,
                "status": "skipped",
                "message": "缺少目标 cell",
                "root_cause": root_cause,
                "root_cause_label": root_cause_label,
            })
            continue

        if write_mode in {"skip", "none"}:
            summary["skipped"] += 1
            root_cause, root_cause_label = _export_audit_root_cause(
                {"status": "skipped", "field_key": field_key, "cell": cell},
                expected=value,
                actual="",
                write_mode=write_mode,
                has_mapping_config=bool(config),
            )
            items.append({
                "cell": cell,
                "field_key": field_key,
                "label": label,
                "expected": _normalize_audit_value(value),
                "actual": "",
                "write_mode": write_mode,
                "matched": False,
                "status": "skipped",
                "message": "write_mode 为 skip/none",
                "root_cause": root_cause,
                "root_cause_label": root_cause_label,
            })
            continue

        try:
            actual = sheet[cell].value
        except Exception as exc:
            summary["skipped"] += 1
            warnings.append({
                "cell": cell,
                "field_key": field_key,
                "label": label,
                "message": f"读取单元格失败：{exc}",
            })
            root_cause, root_cause_label = _export_audit_root_cause(
                {"status": "skipped", "field_key": field_key, "cell": cell},
                expected=value,
                actual="",
                write_mode=write_mode,
                has_mapping_config=bool(config),
            )
            items.append({
                "cell": cell,
                "field_key": field_key,
                "label": label,
                "expected": _normalize_audit_value(value),
                "actual": "",
                "write_mode": write_mode,
                "matched": False,
                "status": "skipped",
                "message": f"读取单元格失败：{exc}",
                "root_cause": root_cause,
                "root_cause_label": root_cause_label,
            })
            continue

        expected_text = _normalize_audit_value(value)
        actual_text = _normalize_audit_value(actual)
        matched = _values_match_for_audit(expected_text, actual_text, write_mode)

        summary["checked"] += 1
        if matched:
            summary["matched"] += 1
            status = "matched"
            message = ""
        else:
            summary["mismatched"] += 1
            status = "mismatched"
            message = "导出回读值与期望值不一致"
            warnings.append({
                "cell": cell,
                "field_key": field_key,
                "label": label,
                "message": message,
                "expected": expected_text,
                "actual": actual_text,
            })

        root_cause, root_cause_label = _export_audit_root_cause(
            {"status": status, "field_key": field_key, "cell": cell},
            expected=expected_text,
            actual=actual_text,
            write_mode=write_mode,
            has_mapping_config=bool(config),
        )
        items.append({
            "cell": cell,
            "field_key": field_key,
            "label": label,
            "expected": expected_text,
            "actual": actual_text,
            "write_mode": write_mode,
            "matched": matched,
            "status": status,
            "message": message,
            "root_cause": root_cause,
            "root_cause_label": root_cause_label,
        })

    root_cause_summary = {}
    for item in items:
        cause = str(item.get("root_cause") or "unknown")
        root_cause_summary[cause] = root_cause_summary.get(cause, 0) + 1

    return {
        "success": True,
        "summary": {
            **summary,
            "root_cause_summary": root_cause_summary,
        },
        "items": items,
        "warnings": warnings,
        "errors": errors,
        "root_cause_summary": root_cause_summary,
    }



def _mapping_health_issue(level, cell, label, field_key, message):
    return {
        "level": level,
        "cell": _cell_key(cell),
        "label": str(label or "").strip(),
        "field_key": str(field_key or "").strip(),
        "message": message,
    }


def _build_mapping_health_report(profile):
    profile = profile if isinstance(profile, dict) else {}
    profile_id = str(profile.get("profile_id") or "").strip()
    configuration = _template_configuration_from_profile(profile)
    workspace_fields = _build_workspace_fields_from_profile(profile)
    ai_contract = _build_ai_extraction_contract_from_workspace_fields(workspace_fields)
    ai_summary = _ai_extraction_contract_summary(ai_contract)
    runtime_mapping_source = _get_runtime_mapping_source(profile)
    required_target_write_modes = {
        "write_right_cell",
        "write_below_cell",
        "append_after_colon",
        "check_option",
        "select_option_text",
        "write_table_column",
        "write_row_field",
        "write_table_cell",
    }
    executable_write_modes = required_target_write_modes
    skipped_intents = {"title", "section_header", "note_instruction", "image_area", "attachment_hint", "readonly_example"}
    skipped_write_modes = {"skip", "none"}
    excel_cell_re = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
    checks = []
    errors = []
    warnings = []
    field_usage = {}
    target_usage = {}

    def add_problem(problems, issue_list, level, cell, label, field_key, message):
        problems.append(message)
        issue_list.append(_mapping_health_issue(level, cell, label, field_key, message))

    for source_cell, item in sorted(configuration.items(), key=lambda pair: _cell_key(pair[0])):
        if not isinstance(item, dict):
            continue
        cell = _cell_key(source_cell)
        field_key = str(item.get("candidate_field_key") or item.get("field_key") or "").strip()
        label = (
            str(item.get("label") or "").strip()
            or str(item.get("candidate_field_label") or "").strip()
            or str(item.get("field_label") or "").strip()
            or field_key
        )
        intent_type = str(item.get("intent_type") or "").strip()
        write_mode = str(item.get("write_mode") or "").strip()
        target_cell = _cell_key(item.get("target_cell"))
        show_in_workspace = item.get("show_in_workspace") is not False
        source_label_cell = _cell_key(item.get("label_cell") or cell)
        section = str(item.get("section") or item.get("section_key") or "").strip()
        semantic_type = str(item.get("semantic_type") or "").strip()
        problems = []
        status = "ok"
        is_skipped = write_mode in skipped_write_modes or intent_type in skipped_intents

        if is_skipped:
            status = "skipped"
        else:
            if show_in_workspace and not field_key:
                add_problem(problems, errors, "error", cell, label, field_key, "show_in_workspace=true 但 field_key 为空，AI 和工作页无法稳定使用。")
            if write_mode in required_target_write_modes and not target_cell:
                add_problem(problems, errors, "error", cell, label, field_key, f"write_mode={write_mode} 需要 target_cell。")
            if target_cell and not excel_cell_re.match(target_cell):
                add_problem(problems, errors, "error", cell, label, field_key, f"target_cell 格式非法：{target_cell}")
            if intent_type in {"option_checkbox", "option_text_choice"} and not str(item.get("option_value") or "").strip():
                add_problem(problems, warnings, "warning", cell, label, field_key, "选项字段缺少 option_value。")

        ai_contract_ready = bool(field_key and write_mode not in skipped_write_modes and intent_type not in skipped_intents)
        workspace_ready = bool(show_in_workspace and field_key and not is_skipped)
        export_ready = bool(target_cell and write_mode in executable_write_modes and not is_skipped and (not target_cell or excel_cell_re.match(target_cell)))

        if field_key and not is_skipped and show_in_workspace:
            field_usage.setdefault(field_key, []).append({
                "cell": cell,
                "label": label,
                "section": section,
                "semantic_type": semantic_type,
            })
        if target_cell and not is_skipped:
            target_usage.setdefault(target_cell, []).append(cell)

        if problems and status != "skipped":
            status = "error" if any(issue["cell"] == cell and issue["level"] == "error" for issue in errors) else "warning"

        checks.append(
            {
                "cell": cell,
                "source_cell": source_label_cell,
                "field_key": field_key,
                "label": label,
                "intent_type": intent_type,
                "write_mode": write_mode,
                "target_cell": target_cell,
                "show_in_workspace": show_in_workspace,
                "ai_contract_ready": ai_contract_ready,
                "workspace_ready": workspace_ready,
                "export_ready": export_ready,
                "status": status,
                "problems": problems,
            }
        )

    check_by_cell = {item["cell"]: item for item in checks}
    
    def _is_same_section_or_semantic(fields):
        if len(fields) < 2:
            return True
        sections = {f["section"] for f in fields if f["section"]}
        semantic_types = {f["semantic_type"] for f in fields if f["semantic_type"]}
        
        if len(sections) == 1:
            return True
        if len(semantic_types) == 1:
            return True
        
        domain_parts = set()
        for f in fields:
            field_key = check_by_cell.get(f["cell"], {}).get("field_key", "")
            if "." in field_key:
                domain = field_key.split(".")[0]
                domain_parts.add(domain)
        if len(domain_parts) == 1:
            return True
        
        return False
    
    for field_key, field_items in field_usage.items():
        if len(field_items) <= 1:
            continue
        
        is_shared_field = _is_same_section_or_semantic(field_items)
        cells = [f["cell"] for f in field_items]
        labels = [f["label"] for f in field_items]
        
        if is_shared_field:
            message = f"共享字段：{field_key} 被 {len(cells)} 个相关单元格使用（{', '.join(cells)}）。"
            level = "info"
        else:
            message = f"field_key 重复：{field_key} 被 {len(cells)} 个配置项使用（{', '.join(cells)}）。"
            level = "warning"
        
        for item in field_items:
            cell = item["cell"]
            label = item["label"]
            check = check_by_cell.get(cell)
            if check:
                check["problems"].append(message)
                if level == "warning" and check["status"] == "ok":
                    check["status"] = "warning"
            if level == "warning":
                warnings.append(_mapping_health_issue(level, cell, label, field_key, message))

    append_write_modes = {
        "append_after_colon",
        "append_value",
        "append_text",
        "append_line",
        "write_composite"
    }
    regular_write_modes = {
        "write_value",
        "write_cell",
        "write_right_cell",
        "write_below_cell",
        "write_table_cell"
    }
    
    for target_cell, cells in target_usage.items():
        if len(cells) <= 1:
            continue
        
        cell_write_modes = []
        for cell in cells:
            check = check_by_cell.get(cell)
            if check:
                cell_write_modes.append(check.get("write_mode", ""))
        
        append_count = sum(1 for wm in cell_write_modes if wm in append_write_modes)
        regular_count = sum(1 for wm in cell_write_modes if wm in regular_write_modes)
        
        if append_count >= 2 and regular_count == 0:
            shared_target_cell = True
            message = f"共享写入单元格：{target_cell} 被 {len(cells)} 个组合字段共同写入（{', '.join(cells)}）。"
            level = "info"
        else:
            shared_target_cell = False
            message = f"target_cell 重复：{target_cell} 被 {len(cells)} 个配置项写入（{', '.join(cells)}）。"
            level = "warning"
        
        for cell in cells:
            check = check_by_cell.get(cell)
            if check:
                check["problems"].append(message)
                if not shared_target_cell and check["status"] == "ok":
                    check["status"] = "warning"
            if level == "warning":
                warnings.append(_mapping_health_issue(level, cell, check.get("label") if check else "", check.get("field_key") if check else "", message))

    export_ready_fields = sum(1 for item in checks if item.get("export_ready"))
    return {
        "success": True,
        "profile_id": profile_id,
        "summary": {
            "total_config_items": len(configuration),
            "workspace_fields": len(workspace_fields),
            "ai_contract_fields": ai_summary.get("fields_count", 0),
            "export_ready_fields": export_ready_fields,
            "errors_count": len(errors),
            "warnings_count": len(warnings),
            "runtime_mapping_source": runtime_mapping_source,
        },
        "runtime_mapping_source": runtime_mapping_source,
        "checks": checks,
        "errors": errors,
        "warnings": warnings,
    }


_WORKSPACE_DOMAIN_ORDER = {
    "basic": 10,
    "product": 20,
    "packaging": 30,
    "labeling": 40,
    "batch_marking": 50,
    "formula": 60,
    "attachment": 70,
    "other": 90,
}

_WORKSPACE_FIELD_KEY_ORDER = {
    "customer_name": 10,
    "order_date": 20,
    "product_name": 30,
    "quantity": 40,
    "product.product_form": 50,
    "product.soft_capsule.shell_size": 60,
    "packaging.container_type": 100,
    "packaging.quantity_per_unit": 110,
    "packaging.bottle_seal_method": 120,
    "packaging.cap_seal_method": 130,
    "packaging.desiccant": 140,
    "packaging.shrink_wrap_full": 150,
    "packaging.protective_bag": 160,
    "labeling.label_requirement": 200,
    "labeling.design_source": 210,
    "labeling.no_label": 220,
    "batch_marking.requirement": 300,
    "formula.bilingual_formula": 400,
}


def _workspace_domain_for_field_key(field_key):
    key = str(field_key or "").strip()
    if key in {"customer_name", "order_date", "quantity", "amount", "specification"}:
        return "basic"
    if key == "product_name" or key.startswith("product."):
        return "product"
    if "." in key:
        domain = key.split(".", 1)[0]
        return domain if domain in _WORKSPACE_DOMAIN_ORDER else "other"
    return "other"


def _workspace_field_order(field_key, display_order):
    try:
        fallback_order = int(display_order or 0)
    except (TypeError, ValueError):
        fallback_order = 0
    return _WORKSPACE_FIELD_KEY_ORDER.get(str(field_key or "").strip(), 1000 + fallback_order)


def _workspace_field_sort_key(item):
    domain = str(item.get("workspace_domain") or "").strip()
    return (
        _WORKSPACE_DOMAIN_ORDER.get(domain, _WORKSPACE_DOMAIN_ORDER["other"]),
        int(item.get("workspace_order") or 0),
        int(item.get("display_order") or 0),
        str(item.get("field_key") or ""),
        str(item.get("cell") or ""),
    )


def _build_workspace_fields_from_profile(profile):
    configuration = _template_configuration_from_profile(profile)
    catalog_labels = get_field_catalog_labels()
    workspace_fields = []
    hidden_intents = {"title", "section_header", "note_instruction", "readonly_example"}
    hidden_write_modes = {"skip", "none"}

    def sort_key(item):
        try:
            return int(item.get("display_order") or 0)
        except (TypeError, ValueError):
            return 0

    for source_cell, item in sorted(configuration.items(), key=lambda pair: sort_key(pair[1])):
        if not isinstance(item, dict):
            continue
        if item.get("show_in_workspace") is False:
            continue

        field_key = str(item.get("candidate_field_key") or item.get("field_key") or "").strip()
        if not field_key:
            continue

        intent_type = str(item.get("intent_type") or "").strip()
        raw_field_type = str(item.get("field_type") or "text").strip()
        is_image_intent = intent_type in {"image_area", "attachment_hint"}
        field_type = "image" if raw_field_type == "image" or is_image_intent else raw_field_type
        is_image_field = field_type == "image"
        write_mode = str(item.get("write_mode") or "").strip()
        is_table_field = field_type in {"table", "dynamic_table"} or write_mode == "write_table_cell"
        if not is_image_field and not is_table_field and intent_type in hidden_intents:
            continue

        if not is_image_field and write_mode in hidden_write_modes:
            continue

        normalized_source_cell = str(source_cell or "").strip().upper()
        target_cell = str(item.get("target_cell") or "").strip().upper()
        catalog_label = str(catalog_labels.get(field_key) or "").strip()
        label = (
            catalog_label
            or str(item.get("label") or "").strip()
            or str(item.get("candidate_field_label") or "").strip()
            or str(item.get("field_label") or "").strip()
            or field_key
        )
        workspace_domain = _workspace_domain_for_field_key(field_key)
        workspace_order = _workspace_field_order(field_key, sort_key(item))
        workspace_fields.append(
            {
                "cell": target_cell or normalized_source_cell,
                "source_cell": str(item.get("label_cell") or normalized_source_cell).strip().upper(),
                "field_key": field_key,
                "label": label,
                "workspace_domain": workspace_domain,
                "workspace_order": workspace_order,
                "intent_type": intent_type,
                "write_mode": write_mode,
                "option_value": str(item.get("option_value") or "").strip(),
                "ai_extract_hint": str(item.get("ai_extract_hint") or "").strip(),
                "section": str(item.get("section") or item.get("section_key") or "").strip(),
                "display_order": sort_key(item),
                "field_type": field_type,
                "image_fit": str(item.get("image_fit") or "contain").strip(),
                "image_anchor_cell": str(item.get("image_anchor_cell") or item.get("target_cell") or normalized_source_cell or "").strip().upper(),
                "col_offset": int(item.get("col_offset") or item.get("table_col_offset") or item.get("column_offset") or 0),
                "table_col_offset": int(item.get("table_col_offset") or item.get("col_offset") or item.get("column_offset") or 0),
            }
        )

    workspace_fields.sort(key=_workspace_field_sort_key)
    return workspace_fields


def _is_blank_extracted_value(value):
    if value is None:
        return True
    if isinstance(value, str):
        text = value.strip()
        return not text or text.lower() in {"null", "none", "n/a", "na"}
    if isinstance(value, (list, dict)):
        return len(value) == 0
    return False


def _stringify_extracted_value(value):
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _table_row_values(raw_value):
    if not isinstance(raw_value, list):
        return []
    rows = []
    for item in raw_value:
        if isinstance(item, dict):
            rows.append(item)
        else:
            rows.append({"value": item})
    return rows


def _table_cell_value(row, field_key):
    if not isinstance(row, dict):
        return ""
    if field_key in row:
        return row.get(field_key)
    if "value" in row:
        return row.get("value")
    return ""


def _table_rows_for_field(parsed_fields, field_key, raw_value):
    direct_rows = _table_row_values(raw_value)
    if direct_rows:
        return direct_rows

    if not isinstance(parsed_fields, dict):
        return []

    preferred_keys = [
        "products",
        "items",
        "rows",
        "table",
        "table_rows",
        "product_items",
        "order_items",
        "details",
    ]

    for key in preferred_keys:
        value = parsed_fields.get(key)
        rows = _table_row_values(value)
        if any(isinstance(row, dict) and field_key in row for row in rows):
            return rows

    for value in parsed_fields.values():
        rows = _table_row_values(value)
        if any(isinstance(row, dict) and field_key in row for row in rows):
            return rows

    return []


def _table_col_offset_for_field(item):
    if not isinstance(item, dict):
        return 0
    for key in ("col_offset", "table_col_offset", "column_offset"):
        raw_value = item.get(key)
        if raw_value in (None, ""):
            continue
        try:
            return int(raw_value)
        except (TypeError, ValueError):
            continue
    return 0


def _bind_parsed_fields_to_template_cells(parsed, profile):
    if not isinstance(parsed, dict):
        parsed_fields = {}
    elif isinstance(parsed.get("fields"), dict):
        parsed_fields = parsed.get("fields", {})
    else:
        parsed_fields = parsed
    contract = _build_ai_extraction_contract_from_profile(profile)
    contract_fields = contract.get("fields") if isinstance(contract, dict) and isinstance(contract.get("fields"), list) else []
    confirmed_cells = []
    operations = []

    for item in contract_fields:
        field_key = str(item.get("field_key") or "").strip()
        cell = _cell_key(item.get("target_cell") or item.get("cell"))
        if not field_key or not cell:
            continue

        raw_value = parsed_fields.get(field_key) if isinstance(parsed_fields, dict) else None
        table_rows = _table_rows_for_field(parsed_fields, field_key, raw_value)
        if table_rows:
            col_offset = _table_col_offset_for_field(item)
            label = str(item.get("label") or item.get("field_label") or field_key).strip() or field_key
            source = "AI Extraction Contract"
            for row_index, row in enumerate(table_rows):
                row_value = _table_cell_value(row, field_key)
                if _is_blank_extracted_value(row_value):
                    continue
                value = _stringify_extracted_value(row_value)
                confirmed_cells.append(
                    {
                        "cell": cell,
                        "display_cell": cell,
                        "source_cell": item.get("source_cell") or "",
                        "source": source,
                        "label": label,
                        "value": value,
                        "field_key": field_key,
                        "write_mode": item.get("write_mode") or "",
                        "intent_type": item.get("intent_type") or "",
                        "option_value": "",
                        "row_offset": row_index,
                        "col_offset": col_offset,
                        "sheet_name": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
                    }
                )
                operations.append(
                    {
                        "op_type": "write_table_cell",
                        "target_cell": cell,
                        "row_offset": row_index,
                        "col_offset": col_offset,
                        "value": value,
                        "source": source,
                        "field_key": field_key,
                        "field_label": label,
                        "mapping_confirmed": True,
                        "ai_extraction_contract": True,
                        "write_mode": item.get("write_mode") or "",
                        "intent_type": item.get("intent_type") or "",
                        "source_cell": item.get("source_cell") or "",
                        "target_sheet": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
                    }
                )
            continue

        if field_key not in parsed_fields:
            continue

        if _is_blank_extracted_value(raw_value):
            continue

        value = _stringify_extracted_value(raw_value)
        label = str(item.get("label") or item.get("field_label") or field_key).strip() or field_key
        source = "AI Extraction Contract"
        confirmed_cells.append(
            {
                "cell": cell,
                "display_cell": cell,
                "source_cell": item.get("source_cell") or "",
                "source": source,
                "label": label,
                "value": value,
                "field_key": field_key,
                "write_mode": item.get("write_mode") or "",
                "intent_type": item.get("intent_type") or "",
                "option_value": "",
                "sheet_name": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
            }
        )
        operations.append(
            {
                "op_type": "write_text",
                "target_cell": cell,
                "value": value,
                "source": source,
                "field_key": field_key,
                "field_label": label,
                "mapping_confirmed": True,
                "ai_extraction_contract": True,
                "write_mode": item.get("write_mode") or "",
                "intent_type": item.get("intent_type") or "",
                "source_cell": item.get("source_cell") or "",
                "target_sheet": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
            }
        )

    return {
        "confirmed_cells": confirmed_cells,
        "operations": operations,
    }


def _operation_merge_key(operation):
    if not isinstance(operation, dict):
        return ""
    cell = _cell_key(operation.get("target_cell") or operation.get("cell") or operation.get("display_cell"))
    if not cell:
        return ""
    row_offset = str(operation.get("row_offset") or 0).strip()
    col_offset = str(operation.get("col_offset") or 0).strip()
    sheet = _sheet_key(operation.get("target_sheet") or operation.get("sheet_name") or operation.get("worksheet") or operation.get("sheet"))
    return f"{sheet}|{cell}|{row_offset}|{col_offset}"


def _merge_field_bound_operations(processed_operations, field_bound_operations):
    operations = deepcopy(processed_operations) if isinstance(processed_operations, list) else []
    bound_operations = field_bound_operations if isinstance(field_bound_operations, list) else []
    operation_index_by_key = {}
    for index, operation in enumerate(operations):
        if not isinstance(operation, dict):
            continue
        merge_key = _operation_merge_key(operation)
        if merge_key:
            operation_index_by_key[merge_key] = index

    added_count = 0
    override_count = 0
    for bound_operation in bound_operations:
        if not isinstance(bound_operation, dict):
            continue
        merge_key = _operation_merge_key(bound_operation)
        if not merge_key:
            continue
        existing_index = operation_index_by_key.get(merge_key)
        if existing_index is None:
            operations.append(deepcopy(bound_operation))
            operation_index_by_key[merge_key] = len(operations) - 1
            added_count += 1
            continue

        operations[existing_index] = {
            **operations[existing_index],
            **deepcopy(bound_operation),
            "ai_extraction_override": True,
        }
        override_count += 1

    return {
        "processed_operations": operations,
        "added_count": added_count,
        "override_count": override_count,
    }


def _normalize_section_configuration_items(items):
    if items is None:
        return {}
    if not isinstance(items, list):
        raise ValueError("分区配置项必须是 list")

    configuration = {}
    for index, item in enumerate(items, 1):
        if not isinstance(item, dict):
            continue
        section_key = str(item.get("section_key") or "").strip()
        if not section_key:
            continue
        configuration[section_key] = {
            "section_label": str(item.get("section_label") or "").strip(),
            "section_order": int(item.get("section_order") or index),
        }
    return configuration


def _current_template_profile_for_export():
    state = get_pipeline_state()
    profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not profile:
        profile = get_current_template_profile()
    return profile if isinstance(profile, dict) else {}


def _clear_mapping_runtime_state():
    set_structured_operations([])
    set_table_operations([])
    set_block_operations([])
    set_unified_operations([])
    set_pipeline_result([], [])
    set_mapping_safety(
        {
            "has_conflicts": False,
            "conflicts": [],
            "warnings": [],
            "skipped_operations": [],
            "overwrite_warnings": [],
        }
    )
    set_mapping_counts(
        {
            "enabled_structured_mappings": 0,
            "enabled_table_mappings": 0,
            "enabled_block_rules": 0,
        }
    )
    set_validator_result({})
    set_render_preview(
        {
            "cells": [],
            "cell_preview": [],
            "table_preview": [],
            "block_preview": [],
            "skipped_preview": [],
            "mapping_safety": {},
            "generated_time": None,
        }
    )
    set_render_targets({"html_preview": "", "excel_preview": []})
    set_excel_result(None)
    return get_pipeline_state()


def _resolve_bound_template_file_path(path_value):
    raw_path = str(path_value or "").strip()
    if not raw_path:
        return None

    path = Path(raw_path)
    if ".." in path.parts:
        raise ValueError("template_file_path 不允许包含 ..")

    resolved = path if path.is_absolute() else get_base_dir() / path
    resolved = resolved.resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"系统模板绑定文件不存在: {raw_path}")
    return resolved


def _resolve_export_template_source():
    profile = _current_template_profile_for_export()
    profile_id = profile.get("profile_id") or ""
    if not profile_id:
        raise ValueError("未选择系统模板")

    template_file_path = str(profile.get("template_file_path") or "").strip()
    if not template_file_path:
        raise ValueError("当前系统模板未绑定模板文件")

    try:
        bound_path = _resolve_bound_template_file_path(template_file_path)
    except FileNotFoundError as exc:
        logger.info(
            "V4 system template file missing: profile_id=%s template_file_path=%s",
            profile_id,
            template_file_path,
        )
        raise ValueError("系统模板文件不存在") from exc

    set_current_template(str(bound_path))
    logger.info(
        "V4 export using system template file: profile_id=%s path=%s",
        profile_id,
        bound_path,
    )
    return bound_path, False, "profile"


def _cell_key(value):
    return str(value or "").strip().upper()


def _sheet_key(value):
    return str(value or "").strip()


def _confirmed_export_empty_summary():
    return {
        "total": 0,
        "written": 0,
        "skipped": 0,
        "by_mode": {},
    }


def _increment_write_mode_summary(summary, write_mode, status):
    mode = str(write_mode or "").strip() or "legacy"
    by_mode = summary.setdefault("by_mode", {})
    if mode not in by_mode:
        by_mode[mode] = {"total": 0, "written": 0, "skipped": 0}
    by_mode[mode]["total"] += 1
    by_mode[mode][status] += 1


def _confirmed_config_lookup_from_profile(profile):
    configuration = _template_configuration_from_profile(profile)
    by_source_cell = {}
    by_target_cell = {}
    by_field_key = {}

    def field_config_rank(config):
        if not isinstance(config, dict):
            return 0
        rank = 0
        if config.get("show_in_workspace") is not False:
            rank += 10
        write_mode = str(config.get("write_mode") or "").strip()
        if write_mode not in {"", "skip", "none"}:
            rank += 5
        if _cell_key(config.get("target_cell")):
            rank += 3
        return rank

    for source_cell, config in configuration.items():
        if not isinstance(config, dict):
            continue
        normalized_source_cell = _cell_key(source_cell)
        item = deepcopy(config)
        item["_source_cell"] = normalized_source_cell
        if normalized_source_cell:
            by_source_cell[normalized_source_cell] = item

        target_cell = _cell_key(item.get("target_cell"))
        if target_cell and target_cell not in by_target_cell:
            by_target_cell[target_cell] = item

        field_key = str(item.get("candidate_field_key") or item.get("field_key") or "").strip()
        if field_key and (
            field_key not in by_field_key
            or field_config_rank(item) > field_config_rank(by_field_key[field_key])
        ):
            by_field_key[field_key] = item

    return {
        "by_source_cell": by_source_cell,
        "by_target_cell": by_target_cell,
        "by_field_key": by_field_key,
    }


def _lookup_confirmed_mapping_config(item, lookup):
    if not isinstance(item, dict) or not isinstance(lookup, dict):
        return {}

    field_key = str(item.get("field_key") or "").strip()
    if field_key and field_key in lookup.get("by_field_key", {}):
        return lookup["by_field_key"][field_key]

    for key_name in ("source_cell", "display_cell", "cell"):
        cell = _cell_key(item.get(key_name))
        if cell and cell in lookup.get("by_source_cell", {}):
            return lookup["by_source_cell"][cell]

    for key_name in ("cell", "display_cell", "source_cell"):
        cell = _cell_key(item.get(key_name))
        if cell and cell in lookup.get("by_target_cell", {}):
            return lookup["by_target_cell"][cell]

    return {}


def _confirmed_item_with_mapping_config(item, config):
    config = config if isinstance(config, dict) else {}
    merged = deepcopy(item) if isinstance(item, dict) else {}
    merged["write_mode"] = str(config.get("write_mode") or merged.get("write_mode") or "").strip()
    merged["intent_type"] = str(config.get("intent_type") or merged.get("intent_type") or "").strip()
    merged["option_value"] = str(config.get("option_value") or merged.get("option_value") or "").strip()
    merged["field_key"] = str(
        config.get("candidate_field_key")
        or config.get("field_key")
        or merged.get("field_key")
        or ""
    ).strip()
    merged["label"] = str(
        config.get("label")
        or config.get("candidate_field_label")
        or merged.get("label")
        or merged.get("field_key")
        or ""
    ).strip()
    merged["source_cell"] = _cell_key(merged.get("source_cell") or config.get("label_cell") or config.get("_source_cell"))
    merged["cell"] = _cell_key(config.get("target_cell") or merged.get("target_cell") or merged.get("image_anchor_cell") or merged.get("cell") or merged.get("display_cell") or merged.get("source_cell"))
    merged["target_cell"] = _cell_key(config.get("target_cell") or merged.get("target_cell") or merged.get("image_anchor_cell") or merged.get("cell"))
    merged["sheet_name"] = _sheet_key(
        config.get("sheet_name")
        or config.get("target_sheet")
        or config.get("worksheet")
        or config.get("sheet")
        or merged.get("sheet_name")
        or merged.get("target_sheet")
        or merged.get("worksheet")
        or merged.get("sheet")
    )
    merged["row_offset"] = int(merged.get("row_offset") or config.get("row_offset") or 0)
    merged["col_offset"] = int(
        merged.get("col_offset")
        or merged.get("table_col_offset")
        or config.get("col_offset")
        or config.get("table_col_offset")
        or config.get("column_offset")
        or 0
    )
    merged["table_col_offset"] = merged["col_offset"]
    return merged


def _load_template_worksheet_for_confirmed_export(template_path):
    if not template_path:
        return None
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(template_path, data_only=False)
        return workbook.active
    except Exception:
        logger.warning("V4 confirmed export template read failed: path=%s", template_path, exc_info=True)
        return None


def _template_cell_text(worksheet, cell_ref):
    cell = _cell_key(cell_ref)
    if not worksheet or not cell or not re.match(r"^[A-Z]{1,3}[1-9][0-9]*$", cell):
        return ""
    try:
        value = worksheet[cell].value
    except Exception:
        return ""
    return "" if value is None else str(value).strip()


def _format_append_after_colon_value(original_text, label, value):
    value_text = str(value or "").strip()
    text = str(original_text or "").strip()
    for separator in ("：", ":"):
        if separator in text:
            prefix = text.split(separator, 1)[0].strip()
            return f"{prefix}{separator}{value_text}" if prefix else value_text
    prefix = str(label or text or "").strip()
    return f"{prefix}：{value_text}" if prefix else value_text


def _is_confirmed_option_selected(value):
    text = str(value or "").strip().lower()
    if not text:
        return False
    return text not in {"false", "0", "no", "n", "none", "null", "unchecked", "否", "不选", "未选"}


def _format_check_option_value(original_text, option_value, value):
    option_text = str(original_text or option_value or value or "").strip()
    option_text = re.sub(r"^[☑☒☐□■\s]+", "", option_text).strip()
    return f"☑ {option_text}" if option_text else "☑"


def _confirmed_write_value(item, worksheet):
    write_mode = str(item.get("write_mode") or "").strip()
    value = item.get("value", "")
    target_cell = item.get("target_cell") or item.get("cell")
    original_text = _template_cell_text(worksheet, target_cell) or _template_cell_text(worksheet, item.get("source_cell"))

    if write_mode == "append_after_colon":
        return _format_append_after_colon_value(original_text, item.get("label"), value)
    if write_mode == "check_option":
        if not _is_confirmed_option_selected(value):
            return ""
        return _format_check_option_value(original_text, item.get("option_value"), value)
    if write_mode == "select_option_text":
        return str(value or "").strip()
    return value


def _materialize_image_data_url_to_temp_file(item):
    if not isinstance(item, dict):
        return ""

    image = item.get("image")
    image = image if isinstance(image, dict) else {}

    data = (
        str(image.get("data_url") or "").strip()
        or str(item.get("image_data") or "").strip()
        or str(item.get("image_base64") or "").strip()
    )
    if not data:
        return ""

    mime_type = str(image.get("mime_type") or item.get("mime_type") or "").strip().lower()
    encoded = data
    if data.startswith("data:"):
        if "," not in data:
            return ""
        header, encoded = data.split(",", 1)
        header_mime = header[5:].split(";")[0].strip().lower()
        mime_type = mime_type or header_mime

    ext = _image_extension_from_mime_type(mime_type or "image/png")
    if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
        return ""

    try:
        raw = base64.b64decode(encoded)
    except Exception:
        logger.warning("V4 confirmed image data_url decode failed", exc_info=True)
        return ""

    if not raw:
        return ""

    tmp_dir = get_base_dir() / "output" / "tmp_images"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    image_path = tmp_dir / f"confirmed_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}{ext}"
    try:
        image_path.write_bytes(raw)
    except Exception:
        logger.warning("V4 confirmed image temp file write failed: path=%s", image_path, exc_info=True)
        return ""
    return str(image_path)


def _confirmed_operation_from_item(item, worksheet):
    target_cell = _cell_key(item.get("target_cell") or item.get("cell"))
    field_type = str(item.get("field_type") or item.get("type") or "").strip().lower()
    image = item.get("image") if isinstance(item.get("image"), dict) else {}
    is_image_field = (
        field_type == "image"
        or bool(item.get("image_path"))
        or bool(item.get("image_data"))
        or bool(item.get("image_base64"))
        or bool(image.get("data_url"))
    )
    if is_image_field:
        image_path = str(item.get("image_path") or "").strip()
        if not image_path:
            image_path = _materialize_image_data_url_to_temp_file(item)
        return {
            "op_type": "write_image",
            "target_cell": target_cell,
            "source": item.get("source") or "Confirmed Workspace",
            "field_key": item.get("field_key") or "",
            "field_label": item.get("label") or "",
            "mapping_confirmed": True,
            "confirmed_override": True,
            "confirmed_label": item.get("label") or "",
            "confirmed_source": item.get("source") or "Confirmed Workspace",
            "write_mode": item.get("write_mode") or "",
            "intent_type": item.get("intent_type") or "",
            "image_anchor_cell": item.get("image_anchor_cell") or item.get("target_cell") or item.get("cell"),
            "image_path": image_path,
            "image_data": item.get("image_data") or "",
            "image_base64": item.get("image_base64") or "",
            "image_fit": item.get("image_fit") or image.get("image_fit") or "contain",
            "confirmed": True,
            "target_sheet": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
        }
    value = _confirmed_write_value(item, worksheet)
    write_mode = str(item.get("write_mode") or "").strip()
    is_table_field = field_type in {"table", "dynamic_table"} or write_mode == "write_table_cell"
    return {
        "op_type": "write_table_cell" if is_table_field else "write_text",
        "target_cell": target_cell,
        "value": value,
        "source": item.get("source") or "Confirmed Workspace",
        "field_key": item.get("field_key") or "",
        "field_label": item.get("label") or "",
        "mapping_confirmed": True,
        "confirmed_override": True,
        "confirmed_label": item.get("label") or "",
        "confirmed_source": item.get("source") or "Confirmed Workspace",
        "write_mode": write_mode,
        "intent_type": item.get("intent_type") or "",
        "source_cell": item.get("source_cell") or "",
        "option_value": item.get("option_value") or "",
        "target_sheet": _sheet_key(item.get("sheet_name") or item.get("target_sheet") or item.get("worksheet") or item.get("sheet")),
        "row_offset": item.get("row_offset") or 0,
        "col_offset": item.get("col_offset") or 0,
    }


def _override_operations_with_confirmed_cells(processed_operations, confirmed_cells, profile=None, template_path=None):
    operations = deepcopy(processed_operations) if isinstance(processed_operations, list) else []
    confirmed_items = _merge_append_after_colon_confirmed_cells_for_export(
        confirmed_cells if isinstance(confirmed_cells, list) else [],
        profile=profile if isinstance(profile, dict) else {},
    )
    confirmed_by_key = {}
    config_lookup = _confirmed_config_lookup_from_profile(profile if isinstance(profile, dict) else {})
    worksheet = _load_template_worksheet_for_confirmed_export(template_path)
    write_mode_summary = _confirmed_export_empty_summary()

    for item in confirmed_items:
        if not isinstance(item, dict):
            continue
        key = _operation_merge_key(item)
        if key:
            confirmed_by_key[key] = item

    override_count = 0
    for operation in operations:
        if not isinstance(operation, dict):
            continue
        operation_key = _operation_merge_key(operation)
        confirmed_item = confirmed_by_key.get(operation_key)
        if not confirmed_item:
            continue

        config = {} if confirmed_item.get("merged_append_after_colon") else _lookup_confirmed_mapping_config(confirmed_item, config_lookup)
        enriched_item = _confirmed_item_with_mapping_config(confirmed_item, config)
        write_mode = str(enriched_item.get("write_mode") or "").strip()
        is_image_item = _confirmed_cell_is_image_item(enriched_item)
        write_mode_summary["total"] += 1
        if write_mode in {"skip", "none"} and not is_image_item:
            operation["value"] = ""
            operation["confirmed_override"] = True
            operation["confirmed_skip_reason"] = f"write_mode={write_mode}"
            write_mode_summary["skipped"] += 1
            _increment_write_mode_summary(write_mode_summary, write_mode, "skipped")
            continue

        confirmed_value = _confirmed_write_value(enriched_item, worksheet)
        operation["confirmed_override"] = True
        operation["mapping_confirmed"] = True
        operation["confirmed_label"] = enriched_item.get("label", "")
        operation["confirmed_source"] = enriched_item.get("source", "")
        operation["write_mode"] = write_mode
        operation["intent_type"] = enriched_item.get("intent_type", "")
        operation["source_cell"] = enriched_item.get("source_cell", "")
        operation["option_value"] = enriched_item.get("option_value", "")
        field_type = str(enriched_item.get("field_type") or enriched_item.get("type") or "").strip().lower()
        if field_type in {"table", "dynamic_table"} or write_mode == "write_table_cell":
            operation["op_type"] = "write_table_cell"
            operation["row_offset"] = enriched_item.get("row_offset") or 0
            operation["col_offset"] = enriched_item.get("col_offset") or 0
        if enriched_item.get("field_key"):
            operation["field_key"] = enriched_item.get("field_key")
        if enriched_item.get("label"):
            operation["field_label"] = enriched_item.get("label")
        override_count += 1
        operation["value"] = confirmed_value
        if str(confirmed_value or "").strip() == "":
            write_mode_summary["skipped"] += 1
            _increment_write_mode_summary(write_mode_summary, write_mode, "skipped")
            continue
        write_mode_summary["written"] += 1
        _increment_write_mode_summary(write_mode_summary, write_mode, "written")

    operation_index_by_key = {}
    for index, operation in enumerate(operations):
        if not isinstance(operation, dict):
            continue
        merge_key = _operation_merge_key(operation)
        if merge_key:
            operation_index_by_key[merge_key] = index

    added_count = 0
    for item in confirmed_items:
        if not isinstance(item, dict):
            continue
        config = {} if item.get("merged_append_after_colon") else _lookup_confirmed_mapping_config(item, config_lookup)
        enriched_item = _confirmed_item_with_mapping_config(item, config)
        target_cell = _cell_key(enriched_item.get("target_cell") or enriched_item.get("cell"))
        merge_key = _operation_merge_key(enriched_item)
        if not target_cell or not merge_key or merge_key in operation_index_by_key:
            continue

        write_mode = str(enriched_item.get("write_mode") or "").strip()
        is_image_item = _confirmed_cell_is_image_item(enriched_item)
        write_mode_summary["total"] += 1
        if write_mode in {"skip", "none"} and not is_image_item:
            write_mode_summary["skipped"] += 1
            _increment_write_mode_summary(write_mode_summary, write_mode, "skipped")
            continue

        operation = _confirmed_operation_from_item(enriched_item, worksheet)
        is_image = operation.get("op_type") == "write_image"
        if not is_image and str(operation.get("value") or "").strip() == "":
            write_mode_summary["skipped"] += 1
            _increment_write_mode_summary(write_mode_summary, write_mode, "skipped")
            continue

        operations.append(operation)
        operation_index_by_key[merge_key] = len(operations) - 1
        added_count += 1
        write_mode_summary["written"] += 1
        _increment_write_mode_summary(write_mode_summary, write_mode, "written")

    return {
        "processed_operations": operations,
        "override_count": override_count,
        "added_count": added_count,
        "write_mode_summary": write_mode_summary,
    }


@router.get("/api/v4/health")
def api_v4_health():
    schema_loaded = False
    examples_count = 0
    excel_rules_loaded = False

    try:
        base_dir = get_base_dir()
        schema_path = base_dir / "v4" / "schemas" / "product_schema.json"
        examples_dir = base_dir / "v4" / "examples"
        rules_path = base_dir / "v4" / "schemas" / "excel_render_rules.json"

        if not schema_path.is_file():
            raise FileNotFoundError(f"Product Schema not found: {schema_path}")
        with schema_path.open("r", encoding="utf-8") as f:
            schema = json.load(f)
        schema_loaded = isinstance(schema, dict) and bool(schema)
        if not schema_loaded:
            raise ValueError("Product Schema is empty or invalid")

        if not examples_dir.is_dir():
            raise FileNotFoundError(f"Examples directory not found: {examples_dir}")
        examples_count = len(list(examples_dir.glob("*.json")))

        if not rules_path.is_file():
            raise FileNotFoundError(f"Excel render rules not found: {rules_path}")
        with rules_path.open("r", encoding="utf-8") as f:
            excel_rules = json.load(f)
        excel_rules_loaded = isinstance(excel_rules, dict) and bool(excel_rules)
        if not excel_rules_loaded:
            raise ValueError("Excel render rules are empty or invalid")

        return {
            "success": True,
            "version": "v4-dev",
            "module": "v4",
            "schema_loaded": schema_loaded,
            "examples_count": examples_count,
            "excel_rules_loaded": excel_rules_loaded,
            "message": "V4 experimental chain is available",
        }
    except Exception as exc:
        logger.exception("V4 health check failed")
        return {
            "success": False,
            "version": "v4-dev",
            "module": "v4",
            "schema_loaded": schema_loaded,
            "examples_count": examples_count,
            "excel_rules_loaded": excel_rules_loaded,
            "error": str(exc),
        }


def _safe_output_filename_part(value: str) -> str:
    text = str(value or "").strip()
    if text.endswith(".json"):
        text = text[:-5]
    text = re.sub(r"[^\w.-]+", "_", text, flags=re.UNICODE)
    return text.strip("._") or "output"


def _resolve_template_path(template_path: str):
    raw_path = str(template_path or "").strip()
    if not raw_path:
        return None, "template_path \u4e0d\u80fd\u4e3a\u7a7a"

    path = Path(raw_path)
    if ".." in path.parts:
        return None, "\u6a21\u677f\u8def\u5f84\u4e0d\u5141\u8bb8\u5305\u542b .."

    resolved = path.resolve()
    if not resolved.is_file():
        return None, "\u6a21\u677f\u6587\u4ef6\u4e0d\u5b58\u5728"

    return resolved, ""


def _resolve_v4_output_file(filename: str):
    requested_name = str(filename or "").strip()
    requested_path = Path(requested_name)
    if (
        not requested_name
        or "/" in requested_name
        or "\\" in requested_name
        or requested_name in {".", ".."}
        or requested_path.is_absolute()
        or requested_path.name != requested_name
    ):
        return None

    if requested_name.endswith(".json"):
        media_type = "application/json"
    elif requested_name.endswith(".xlsx"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        return None

    output_dir = (get_base_dir() / "v4" / "output").resolve()
    output_path = (output_dir / requested_name).resolve()

    try:
        output_path.relative_to(output_dir)
    except ValueError:
        return None

    if not output_path.is_file():
        return None

    return output_path, media_type


def _v4_rules_dir():
    rules_dir = get_base_dir() / "v4" / "rules"
    rules_dir.mkdir(parents=True, exist_ok=True)
    return rules_dir


def _read_v4_rule_file(filename: str, default_data: dict):
    path = _v4_rules_dir() / filename
    if not path.is_file():
        _write_v4_rule_file(filename, default_data)
        return json.loads(json.dumps(default_data, ensure_ascii=False))
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, json.JSONDecodeError):
        data = json.loads(json.dumps(default_data, ensure_ascii=False))
        _write_v4_rule_file(filename, data)
    return data if isinstance(data, dict) else json.loads(json.dumps(default_data, ensure_ascii=False))


def _write_v4_rule_file(filename: str, data: dict):
    path = _v4_rules_dir() / filename
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def _structured_rule_default():
    return {"version": "V4-Rebuild", "mappings": []}


def _table_rule_default():
    return {"version": "V4-Rebuild", "tables": []}


def _block_rule_default():
    return {"version": "V4-Rebuild", "blocks": []}


@router.get("/api/v4/product-schema")
def api_v4_product_schema():
    logger.info("V4 product schema requested")
    return {
        "success": True,
        "data": load_product_schema(),
    }


@router.get("/api/v4/field-catalog")
def api_v4_field_catalog():
    logger.info("V4 field catalog requested")
    fields = flatten_field_catalog()
    return {
        "success": True,
        "fields": fields,
        "labels": get_field_catalog_labels(),
    }


def _load_v4_order_object_for_schema_version():
    order_object_path = get_base_dir() / "v4" / "examples" / "order_object.json"
    if not order_object_path.is_file():
        return None

    try:
        with order_object_path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        logger.warning("V4 order object schema version read failed: path=%s", order_object_path, exc_info=True)
        return None


@router.get("/api/v4/schema-version")
def api_v4_schema_version():
    logger.info("V4 schema version requested")
    compatibility = check_schema_compatibility(_load_v4_order_object_for_schema_version())
    return {
        "success": True,
        "current_version": get_current_schema_version(),
        "order_object_version": compatibility.get("order_object_version"),
        "compatible": compatibility.get("compatible", True),
        "level": compatibility.get("level", "warning"),
        "message": compatibility.get("message", ""),
    }


@router.get("/api/v4/template-profiles")
def api_v4_template_profiles():
    logger.info("V4 template profiles requested")
    profiles = list_template_profiles()
    return {
        "success": True,
        "profiles": profiles,
    }


@router.get("/api/v4/current-template-profile")
def api_v4_current_template_profile():
    logger.info("V4 current template profile requested")
    state = get_pipeline_state()
    profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not profile:
        profile = get_current_template_profile()
    validation = validate_template_profile(profile)
    runtime_mapping_source = _get_runtime_mapping_source(profile)
    return {
        "success": True,
        "profile": profile,
        "workspace_fields": _build_workspace_fields_from_profile(profile),
        "runtime_mapping_source": runtime_mapping_source,
        "validation": validation,
    }


@router.get("/api/v4/pipeline-profile-debug")
def api_v4_pipeline_profile_debug():
    logger.info("V4 pipeline profile debug requested")

    state = get_pipeline_state()
    profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not profile:
        profile = get_current_template_profile()
    profile = profile if isinstance(profile, dict) else {}

    return {
        "success": True,
        "active_profile_source": "pipeline_state" if state.get("current_profile") else "default_profile",
        "profile_id": profile.get("profile_id"),
        "profile_name": profile.get("profile_name"),
        "template_file_path": profile.get("template_file_path"),
        "structured_mapping_file": profile.get("structured_mapping_file"),
        "table_mapping_file": profile.get("table_mapping_file"),
        "block_rules_file": profile.get("block_rules_file"),
    }


@router.post("/api/v4/set-current-template-profile")
def api_v4_set_current_template_profile(payload: Any = Body(None)):
    logger.info("V4 set current template profile requested")

    payload = payload if isinstance(payload, dict) else {}
    profile_id = str(payload.get("profile_id") or "").strip()
    if not profile_id:
        return {
            "success": False,
            "error": "profile_id 不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "Template Profile 不存在",
            "pipeline_state": get_pipeline_state(),
        }

    state = set_current_profile(profile)
    state = _clear_mapping_runtime_state()
    template_file_path = str(profile.get("template_file_path") or "").strip()
    if template_file_path:
        try:
            bound_template_path = _resolve_bound_template_file_path(template_file_path)
            state = set_current_template(str(bound_template_path))
            logger.info(
                "V4 current template profile bound template resolved: profile_id=%s path=%s",
                profile.get("profile_id"),
                bound_template_path,
            )
        except (OSError, ValueError) as exc:
            state = set_current_template(None)
            logger.info(
                "V4 current template profile bound template unavailable: profile_id=%s error=%s",
                profile.get("profile_id"),
                exc,
            )
    else:
        state = set_current_template(None)
    return {
        "success": True,
        "message": "Current Template Profile 已设置",
        "profile": profile,
        "workspace_fields": _build_workspace_fields_from_profile(profile),
        "runtime_mapping_source": _get_runtime_mapping_source(profile),
        "pipeline_state": state,
    }


@router.get("/api/v4/template-profiles/{profile_id}")
def api_v4_template_profile_detail(profile_id: str):
    logger.info("V4 template profile detail requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "Template Profile 不存在",
            "profile": {},
            "validation": {
                "valid": False,
                "warnings": [],
                "errors": ["Template Profile 不存在"],
                "file_status": {},
            },
        }

    validation = validate_template_profile(profile)
    return {
        "success": True,
        "profile": profile,
        "validation": validation,
    }


@router.get("/api/v4/template-profiles/{profile_id}/configuration")
def api_v4_template_profile_configuration(profile_id: str):
    from app.v4_template_analysis import analyze_template

    logger.info("V4 template profile configuration requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "映射不存在",
            "layout_sections": [],
            "template_analysis": {},
        }

    template_file_path = str(profile.get("template_file_path") or "").strip()
    if not template_file_path:
        return {
            "success": True,
            "profile": profile,
            "has_template_file": False,
            "layout_sections": [],
            "template_analysis": {},
            "template_labels": [],
            "template_analysis_summary": {},
            "semantic_summary": {},
            "template_configuration": _template_configuration_from_profile(profile),
            "section_configuration": _section_configuration_from_profile(profile),
            "runtime_mapping_source": _get_runtime_mapping_source(profile),
            "excel_feature_flags": _get_excel_feature_flags(profile),
            "mapping_candidates": [],
        }

    try:
        bound_template_path = _resolve_bound_template_file_path(template_file_path)
        analysis = analyze_template(bound_template_path)
        layout_result = build_layout_sections_from_template_analysis(analysis)
        layout_sections = layout_result.get("layout_sections", [])
        mapping_candidates = _generate_mapping_candidates(analysis, layout_sections, bound_template_path)
        return {
            "success": True,
            "profile": profile,
            "has_template_file": True,
            "layout_sections": layout_sections,
            "layout_summary": layout_result.get("summary", {}),
            "template_analysis": analysis if isinstance(analysis, dict) else {},
            "template_labels": analysis.get("labels", []) if isinstance(analysis, dict) else [],
            "template_analysis_summary": analysis.get("summary", {}) if isinstance(analysis, dict) else {},
            "semantic_summary": analysis.get("semantic_summary", {}) if isinstance(analysis, dict) else {},
            "template_configuration": _template_configuration_from_profile(profile),
            "section_configuration": _section_configuration_from_profile(profile),
            "runtime_mapping_source": _get_runtime_mapping_source(profile),
            "excel_feature_flags": _get_excel_feature_flags(profile),
            "mapping_candidates": mapping_candidates,
        }
    except (BadZipFile, InvalidFileException) as exc:
        logger.warning("V4 template profile configuration invalid Excel: profile_id=%s", profile_id, exc_info=True)
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "layout_sections": [],
            "template_analysis": {},
        }
    except (OSError, ValueError) as exc:
        return {
            "success": False,
            "error": str(exc),
            "layout_sections": [],
            "template_analysis": {},
        }
    except Exception as exc:
        logger.exception("V4 template profile configuration failed")
        return {
            "success": False,
            "error": str(exc) or "模板配置加载失败",
            "layout_sections": [],
            "template_analysis": {},
        }


@router.post("/api/v4/template-profiles/{profile_id}/regenerate-field-catalog-candidates")
def api_v4_template_profile_regenerate_field_catalog_candidates(profile_id: str):
    from app.v4_template_analysis import analyze_template

    logger.info("V4 field catalog candidate regeneration requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "Template Profile 不存在",
            "mapping_candidates": [],
        }

    template_file_path = str(profile.get("template_file_path") or "").strip()
    if not template_file_path:
        return {
            "success": False,
            "error": "当前映射尚未绑定模板文件。",
            "mapping_candidates": [],
        }

    try:
        bound_template_path = _resolve_bound_template_file_path(template_file_path)
        analysis = analyze_template(bound_template_path)
        layout_result = build_layout_sections_from_template_analysis(analysis)
        layout_sections = layout_result.get("layout_sections", [])
        mapping_candidates = _generate_mapping_candidates(analysis, layout_sections, bound_template_path)
        catalog = load_field_catalog()
        return {
            "success": True,
            "profile_id": profile.get("profile_id", ""),
            "field_catalog_version": catalog.get("version") or catalog.get("schema_version") or "",
            "layout_sections": layout_sections,
            "layout_summary": layout_result.get("summary", {}),
            "template_analysis": analysis if isinstance(analysis, dict) else {},
            "template_labels": analysis.get("labels", []) if isinstance(analysis, dict) else [],
            "template_analysis_summary": analysis.get("summary", {}) if isinstance(analysis, dict) else {},
            "semantic_summary": analysis.get("semantic_summary", {}) if isinstance(analysis, dict) else {},
            "mapping_candidates": mapping_candidates,
            "summary": {
                "mapping_candidates_count": len(mapping_candidates),
                "field_catalog_candidates_count": sum(
                    1 for item in mapping_candidates
                    if isinstance(item, dict) and str(item.get("source") or "").startswith("field_catalog")
                ),
                "applicable_candidates_count": sum(
                    1 for item in mapping_candidates
                    if isinstance(item, dict)
                    and item.get("semantic_promoted") is True
                    and float(item.get("confidence") or 0) >= 0.70
                ),
            },
        }
    except (BadZipFile, InvalidFileException) as exc:
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "mapping_candidates": [],
        }
    except (OSError, ValueError) as exc:
        return {
            "success": False,
            "error": str(exc),
            "mapping_candidates": [],
        }
    except Exception as exc:
        logger.exception("V4 field catalog candidate regeneration failed")
        return {
            "success": False,
            "error": str(exc) or "字段库重新识别失败",
            "mapping_candidates": [],
        }


@router.get("/api/v4/template-profiles/{profile_id}/visual-grid")
def api_v4_template_profile_visual_grid(profile_id: str):
    from app.v4_template_analysis import analyze_template

    logger.info("V4 template profile visual grid requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "映射不存在",
            "visual_grid": {"rows": 0, "cols": 0, "cells": [], "merges": []},
        }

    template_file_path = str(profile.get("template_file_path") or "").strip()
    if not template_file_path:
        return {
            "success": False,
            "error": "当前映射尚未绑定模板文件。",
            "visual_grid": {"rows": 0, "cols": 0, "cells": [], "merges": []},
        }

    try:
        bound_template_path = _resolve_bound_template_file_path(template_file_path)
        analysis = analyze_template(bound_template_path)
        layout_result = build_layout_sections_from_template_analysis(analysis)
        layout_sections = layout_result.get("layout_sections", [])
        mapping_candidates = _generate_mapping_candidates(analysis, layout_sections, bound_template_path)
        template_configuration = _template_configuration_from_profile(profile)
        semantic_by_cell = _semantic_by_cell_from_analysis(analysis)
        visual_grid = _build_visual_grid(bound_template_path, mapping_candidates, template_configuration, semantic_by_cell)
        return {
            "success": True,
            "profile_id": profile.get("profile_id", ""),
            "template_filename": profile.get("template_filename", ""),
            "visual_grid": visual_grid,
            "semantic_summary": analysis.get("semantic_summary", {}) if isinstance(analysis, dict) else {},
        }
    except (BadZipFile, InvalidFileException) as exc:
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "visual_grid": {"rows": 0, "cols": 0, "cells": [], "merges": []},
        }
    except (OSError, ValueError) as exc:
        return {
            "success": False,
            "error": str(exc),
            "visual_grid": {"rows": 0, "cols": 0, "cells": [], "merges": []},
        }
    except Exception as exc:
        logger.exception("V4 template profile visual grid failed")
        return {
            "success": False,
            "error": str(exc) or "可视化模板配置加载失败",
            "visual_grid": {"rows": 0, "cols": 0, "cells": [], "merges": []},
        }


@router.get("/api/v4/template-profiles/{profile_id}/mapping-health")
def api_v4_template_profile_mapping_health(profile_id: str):
    logger.info("V4 template profile mapping health requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "profile_id": profile_id,
            "error": "Template Profile 不存在",
            "summary": {
                "total_config_items": 0,
                "workspace_fields": 0,
                "ai_contract_fields": 0,
                "export_ready_fields": 0,
                "errors_count": 1,
                "warnings_count": 0,
                "runtime_mapping_source": {
                    "source": "empty",
                    "saved_fields_count": 0,
                    "semantic_fields_count": 0,
                    "using_saved_configuration": False,
                },
            },
            "runtime_mapping_source": {
                "source": "empty",
                "saved_fields_count": 0,
                "semantic_fields_count": 0,
                "using_saved_configuration": False,
            },
            "checks": [],
            "errors": [_mapping_health_issue("error", "", "", "", "Template Profile 不存在")],
            "warnings": [],
        }
    return _build_mapping_health_report(profile)


@router.get("/api/v4/template-profiles/{profile_id}/runtime-trace")
def api_v4_template_profile_runtime_trace(profile_id: str):
    logger.info("V4 template profile runtime trace requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        empty_runtime_source = {
            "source": "empty",
            "saved_fields_count": 0,
            "semantic_fields_count": 0,
            "using_saved_configuration": False,
        }
        return {
            "success": False,
            "profile_id": profile_id,
            "error": "Template Profile 不存在",
            "runtime_mapping_source": empty_runtime_source,
            "summary": {
                "saved_configuration_count": 0,
                "workspace_fields_count": 0,
                "ai_contract_fields_count": 0,
                "export_ready_count": 0,
                "warnings_count": 0,
                "errors_count": 1,
            },
            "saved_configuration_count": 0,
            "workspace_fields_count": 0,
            "ai_contract_fields_count": 0,
            "export_ready_count": 0,
            "trace_items": [],
            "warnings": [],
            "errors": [{
                "cell": "",
                "target_cell": "",
                "field_key": "",
                "label": "",
                "problems": ["Template Profile 不存在"],
            }],
        }
    return _build_runtime_mapping_trace_report(profile)



@router.post("/api/v4/template-profiles/{profile_id}/configuration")
def api_v4_template_profile_configuration_save(profile_id: str, payload: Any = Body(None)):
    logger.info("V4 template profile configuration save requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "映射不存在",
        }

    try:
        payload = payload if isinstance(payload, dict) else {}
        configuration = _normalize_template_configuration_items(payload.get("items"))
        section_configuration = _normalize_section_configuration_items(payload.get("sections"))
        excel_feature_flags = _get_excel_feature_flags({
            "excel_feature_flags": payload.get("excel_feature_flags", profile.get("excel_feature_flags", {}))
        })
        render_config = profile.get("render_config") if isinstance(profile.get("render_config"), dict) else {}
        profile["render_config"] = {
            **render_config,
            "template_configuration": configuration,
            "section_configuration": section_configuration,
        }
        profile["excel_feature_flags"] = excel_feature_flags
        saved_profile = save_template_profile(profile)
        state = get_pipeline_state()
        current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
        if current_profile.get("profile_id") == saved_profile.get("profile_id"):
            state = set_current_profile(saved_profile)
        return {
            "success": True,
            "message": "模板配置已保存",
            "profile": saved_profile,
            "template_configuration": _template_configuration_from_profile(saved_profile),
            "section_configuration": _section_configuration_from_profile(saved_profile),
            "runtime_mapping_source": _get_runtime_mapping_source(saved_profile),
            "excel_feature_flags": _get_excel_feature_flags(saved_profile),
            "pipeline_state": state,
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
        }
    except Exception as exc:
        logger.exception("V4 template profile configuration save failed")
        return {
            "success": False,
            "error": str(exc) or "模板配置保存失败",
        }


@router.post("/api/v4/template-profiles/create")
def api_v4_template_profile_create(payload: Any = Body(None)):
    logger.info("V4 template profile create requested")
    try:
        profile = create_template_profile(payload if isinstance(payload, dict) else {})
        validation = validate_template_profile(profile)
        return {
            "success": True,
            "message": "Template Profile 已创建",
            "profile": profile,
            "validation": validation,
        }
    except Exception as exc:
        logger.exception("V4 template profile create failed")
        return {
            "success": False,
            "error": str(exc) or "Template Profile 创建失败",
        }


@router.post("/api/v4/template-profiles/create-empty")
def api_v4_template_profile_create_empty(payload: Any = Body(None)):
    logger.info("V4 template profile create empty requested")
    try:
        payload = payload if isinstance(payload, dict) else {}
        profile_name = str(payload.get("profile_name") or "").strip()
        if not profile_name:
            return {
                "success": False,
                "error": "映射名称不能为空",
            }

        base_id = _safe_template_profile_id(profile_name)
        profile_id = base_id
        if load_template_profile(profile_id):
            profile_id = f"{base_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}"

        profile = create_template_profile(
            {
                "profile_id": profile_id,
                "profile_name": profile_name,
                "template_name": "",
                "template_filename": "",
                "template_file_path": "",
            }
        )
        validation = validate_template_profile(profile)
        return {
            "success": True,
            "message": "映射已创建",
            "profile": profile,
            "validation": validation,
        }
    except Exception as exc:
        logger.exception("V4 template profile create empty failed")
        return {
            "success": False,
            "error": str(exc) or "映射创建失败",
        }


@router.post("/api/v4/template-profiles/save")
def api_v4_template_profile_save(payload: Any = Body(None)):
    logger.info("V4 template profile save requested")
    try:
        if not isinstance(payload, dict):
            return {
                "success": False,
                "error": "payload 必须是 object",
            }

        profile = save_template_profile(payload)
        validation = validate_template_profile(profile)
        return {
            "success": True,
            "message": "Template Profile 已保存",
            "profile": profile,
            "validation": validation,
        }
    except Exception as exc:
        logger.exception("V4 template profile save failed")
        return {
            "success": False,
            "error": str(exc) or "Template Profile 保存失败",
        }


@router.post("/api/v4/template-profiles/{profile_id}/template-file")
def api_v4_template_profile_template_file_update(profile_id: str, payload: Any = Body(None)):
    logger.info("V4 template profile template file update requested: profile_id=%s", profile_id)
    payload = payload if isinstance(payload, dict) else {}
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "Template Profile 不存在",
        }

    profile["template_file_path"] = str(payload.get("template_file_path") or "").strip()
    try:
        saved_profile = save_template_profile(profile)
        state = get_pipeline_state()
        current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
        if current_profile.get("profile_id") == saved_profile.get("profile_id"):
            state = set_current_profile(saved_profile)
        validation = validate_template_profile(saved_profile)
        return {
            "success": True,
            "message": "Template Profile 模板文件已保存",
            "profile": saved_profile,
            "validation": validation,
            "pipeline_state": state,
        }
    except Exception as exc:
        logger.exception("V4 template profile template file update failed")
        return {
            "success": False,
            "error": str(exc) or "Template Profile 模板文件保存失败",
        }


@router.post("/api/v4/template-profiles/{profile_id}/replace-template-file")
def api_v4_template_profile_replace_template_file(profile_id: str, template_file: UploadFile = File(...)):
    from app.v4_template_analysis import analyze_template

    logger.info(
        "V4 template profile replace template file requested: profile_id=%s filename=%s",
        profile_id,
        template_file.filename,
    )
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "系统模板不存在",
        }

    template_path = None
    try:
        template_path, original_name = _save_v4_system_template_file(profile_id, template_file)
        analysis = analyze_template(template_path)
        state = set_current_template(template_path.name)
        state = set_template_analysis(analysis)
        layout_result = build_layout_sections_from_template_analysis(analysis)

        profile["template_name"] = Path(original_name or template_path.name).stem
        profile["template_filename"] = original_name or template_path.name
        profile["template_file_path"] = _system_template_relative_path(template_path)
        saved_profile = save_template_profile(profile)
        state = set_current_profile(saved_profile)
        validation = validate_template_profile(saved_profile)

        return {
            "success": True,
            "message": "模板文件已更新",
            "profile": saved_profile,
            "validation": validation,
            "template_analysis_summary": analysis.get("summary", {}) if isinstance(analysis, dict) else {},
            "semantic_summary": analysis.get("semantic_summary", {}) if isinstance(analysis, dict) else {},
            "template_labels": analysis.get("labels", []) if isinstance(analysis, dict) else [],
            "layout_summary": layout_result.get("summary", {}),
            "layout_sections": layout_result.get("layout_sections", []),
            "pipeline_state": state,
        }
    except (BadZipFile, InvalidFileException) as exc:
        if template_path:
            template_path.unlink(missing_ok=True)
        logger.warning("V4 template profile replace invalid Excel: path=%s", template_path, exc_info=True)
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "pipeline_state": get_pipeline_state(),
        }
    except ValueError as exc:
        if template_path:
            template_path.unlink(missing_ok=True)
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        if template_path:
            template_path.unlink(missing_ok=True)
        logger.exception("V4 template profile replace template file failed")
        return {
            "success": False,
            "error": str(exc) or "模板文件更新失败",
            "pipeline_state": get_pipeline_state(),
        }


@router.post("/api/v4/template-profiles/{profile_id}/delete-template-file")
def api_v4_template_profile_delete_template_file(profile_id: str):
    logger.info("V4 template profile delete template file requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "系统模板不存在",
        }

    try:
        template_file_path = str(profile.get("template_file_path") or "").strip()
        deleted = False
        if template_file_path:
            target_path = _resolve_template_file_path_for_delete(template_file_path)
            if target_path and target_path.exists():
                if not target_path.is_file():
                    raise ValueError("模板文件路径不是文件")
                target_path.unlink()
                deleted = True

        profile["template_file_path"] = ""
        profile["template_filename"] = ""
        profile["template_name"] = ""
        saved_profile = save_template_profile(profile)
        state = set_current_profile(saved_profile)
        state = set_current_template(None)
        validation = validate_template_profile(saved_profile)
        return {
            "success": True,
            "message": "模板文件已删除" if deleted else "模板文件绑定已清空",
            "profile": saved_profile,
            "validation": validation,
            "file_deleted": deleted,
            "pipeline_state": state,
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 template profile delete template file failed")
        return {
            "success": False,
            "error": str(exc) or "模板文件删除失败",
            "pipeline_state": get_pipeline_state(),
        }


@router.post("/api/v4/template-profiles/{profile_id}/delete")
def api_v4_template_profile_delete(profile_id: str):
    logger.info("V4 template profile delete requested: profile_id=%s", profile_id)
    profile = load_template_profile(profile_id)
    if not profile:
        return {
            "success": False,
            "error": "映射不存在",
        }

    try:
        template_file_path = str(profile.get("template_file_path") or "").strip()
        template_deleted = False
        if template_file_path:
            target_path = _resolve_template_file_path_for_delete(template_file_path)
            if target_path and target_path.exists():
                if not target_path.is_file():
                    raise ValueError("模板文件路径不是文件")
                target_path.unlink()
                template_deleted = True

        profile_path = _resolve_template_profile_json_path_for_delete(profile_id, profile)
        if not profile_path.is_file():
            return {
                "success": False,
                "error": "映射不存在",
            }
        profile_path.unlink()

        state = get_pipeline_state()
        current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
        if _safe_template_profile_id(current_profile.get("profile_id")) == _safe_template_profile_id(profile_id):
            set_current_profile({})
            set_current_template(None)

        return {
            "success": True,
            "message": "映射已删除",
            "template_file_deleted": template_deleted,
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 template profile delete failed")
        return {
            "success": False,
            "error": str(exc) or "映射删除失败",
            "pipeline_state": get_pipeline_state(),
        }


@router.post("/api/v4/template-profiles/analyze-template")
def api_v4_template_profile_analyze_template(
    profile_name: str = Form(...),
    template_file: UploadFile = File(...),
):
    from app.v4_template_analysis import analyze_template

    requested_profile_name = str(profile_name or "").strip()
    logger.info(
        "V4 template profile analyze template requested: profile_name=%s filename=%s",
        requested_profile_name,
        template_file.filename,
    )
    if not requested_profile_name:
        return {
            "success": False,
            "error": "系统模板名称不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    template_path = None
    try:
        original_name = Path(template_file.filename or "").name
        if not original_name:
            return {
                "success": False,
                "error": "模板文件名不能为空",
                "pipeline_state": get_pipeline_state(),
            }

        safe_name = _safe_output_filename_part(requested_profile_name)
        profile_id = f"system_template_{safe_name}"
        if load_template_profile(profile_id):
            profile_id = f"{profile_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}"

        profile = create_template_profile(
            {
                "profile_id": profile_id,
                "profile_name": requested_profile_name,
                "template_name": Path(original_name).stem,
                "template_filename": original_name,
            }
        )

        safe_profile_id = _safe_output_filename_part(profile_id)
        safe_original_name = _safe_output_filename_part(Path(original_name).stem) + Path(original_name).suffix.lower()
        system_templates_dir = get_base_dir() / "v4" / "system_templates"
        system_templates_dir.mkdir(parents=True, exist_ok=True)
        template_path = system_templates_dir / f"{safe_profile_id}_{safe_original_name}"
        with template_path.open("wb") as buffer:
            template_file.file.seek(0)
            shutil.copyfileobj(template_file.file, buffer)

        if template_path.stat().st_size <= 0:
            template_path.unlink(missing_ok=True)
            return {
                "success": False,
                "error": "模板文件为空",
                "pipeline_state": get_pipeline_state(),
            }

        analysis = analyze_template(template_path)
        state = set_current_template(template_path.name)
        state = set_template_analysis(analysis)
        layout_result = build_layout_sections_from_template_analysis(analysis)

        template_file_path = _system_template_relative_path(template_path)

        profile["template_name"] = Path(original_name or template_path.name).stem
        profile["template_filename"] = original_name or template_path.name
        profile["template_file_path"] = template_file_path
        saved_profile = save_template_profile(profile)
        state = set_current_profile(saved_profile)

        validation = validate_template_profile(saved_profile)
        return {
            "success": True,
            "message": "系统模板已生成",
            "profile": saved_profile,
            "validation": validation,
            "template_path": str(template_path),
            "template_file_path": template_file_path,
            "template_analysis_summary": analysis.get("summary", {}) if isinstance(analysis, dict) else {},
            "semantic_summary": analysis.get("semantic_summary", {}) if isinstance(analysis, dict) else {},
            "template_labels": analysis.get("labels", []) if isinstance(analysis, dict) else [],
            "layout_summary": layout_result.get("summary", {}),
            "layout_sections": layout_result.get("layout_sections", []),
            "pipeline_state": state,
        }
    except (BadZipFile, InvalidFileException) as exc:
        logger.warning("V4 template profile analyze invalid Excel: path=%s", template_path, exc_info=True)
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "pipeline_state": get_pipeline_state(),
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 template profile analyze template failed")
        return {
            "success": False,
            "error": str(exc) or "公司模板分析失败",
            "pipeline_state": get_pipeline_state(),
        }


@router.get("/api/v4/structured-mapping")
def api_v4_structured_mapping():
    logger.info("V4 structured mapping requested")
    return {
        "success": True,
        "data": _read_v4_rule_file("structured_excel_mapping.json", _structured_rule_default()),
    }


@router.post("/api/v4/structured-mapping/save")
def api_v4_structured_mapping_save(payload: Any = Body(None)):
    logger.info("V4 structured mapping save requested")
    payload = payload if isinstance(payload, dict) else {}
    mappings = payload.get("mappings", [])
    if not isinstance(mappings, list):
        mappings = []
    data = {
        "version": str(payload.get("version") or "V4-Rebuild"),
        "mappings": mappings,
    }
    _write_v4_rule_file("structured_excel_mapping.json", data)
    return {"success": True, "message": "结构化字段映射已保存", "data": data}


@router.get("/api/v4/table-mapping")
def api_v4_table_mapping():
    logger.info("V4 table mapping requested")
    return {
        "success": True,
        "data": _read_v4_rule_file("table_mapping.json", _table_rule_default()),
    }


@router.post("/api/v4/table-mapping/save")
def api_v4_table_mapping_save(payload: Any = Body(None)):
    logger.info("V4 table mapping save requested")
    payload = payload if isinstance(payload, dict) else {}
    tables = payload.get("tables", [])
    if not isinstance(tables, list):
        tables = []
    data = {
        "version": str(payload.get("version") or "V4-Rebuild"),
        "tables": tables,
    }
    _write_v4_rule_file("table_mapping.json", data)
    return {"success": True, "message": "动态表格映射已保存", "data": data}


@router.get("/api/v4/block-merge-rules")
def api_v4_block_merge_rules():
    logger.info("V4 block merge rules requested")
    return {
        "success": True,
        "data": _read_v4_rule_file("block_merge_rules.json", _block_rule_default()),
    }


@router.post("/api/v4/block-merge-rules/save")
def api_v4_block_merge_rules_save(payload: Any = Body(None)):
    logger.info("V4 block merge rules save requested")
    payload = payload if isinstance(payload, dict) else {}
    blocks = payload.get("blocks", [])
    if not isinstance(blocks, list):
        blocks = []
    data = {
        "version": str(payload.get("version") or "V4-Rebuild"),
        "blocks": blocks,
    }
    _write_v4_rule_file("block_merge_rules.json", data)
    return {"success": True, "message": "区块合并规则已保存", "data": data}


@router.get("/api/v4/pipeline-state")
def api_v4_pipeline_state():
    logger.info("V4 pipeline state requested")
    return {
        "success": True,
        "pipeline_state": get_pipeline_state(),
    }


@router.post("/api/v4/reset-pipeline-state")
def api_v4_reset_pipeline_state():
    logger.info("V4 pipeline state reset requested")
    return {
        "success": True,
        "message": "Pipeline State 已重置",
        "pipeline_state": reset_pipeline_state(),
    }


@router.post("/api/v4/maintenance/clear-runtime-state")
def api_v4_maintenance_clear_runtime_state():
    logger.info("V4 maintenance clear runtime state requested")
    cleared = []

    load_order_object_into_pipeline({})
    cleared.append("current_order_object")

    set_structured_operations([])
    set_table_operations([])
    set_block_operations([])
    set_unified_operations([])
    cleared.extend([
        "structured_operations",
        "table_operations",
        "block_operations",
        "unified_operations",
    ])

    set_pipeline_result([], [])
    cleared.append("pipeline_result")

    set_render_targets({"html_preview": "", "excel_preview": []})
    cleared.append("render_targets")

    set_excel_result(None)
    cleared.append("excel_result")

    set_validator_result({})
    cleared.append("validator_result")

    set_mapping_safety({})
    cleared.append("mapping_safety")

    return {
        "success": True,
        "cleared": cleared,
        "pipeline_state": get_pipeline_state(),
    }


def _maintenance_is_within(base_path, candidate_path):
    try:
        candidate_path.resolve().relative_to(base_path.resolve())
        return True
    except ValueError:
        return False


def _maintenance_cleanup_old_files(directory, patterns, cutoff_time):
    deleted_files = []
    skipped = []
    errors = []
    directory = Path(directory)
    if not directory.is_dir():
        return {"deleted_files": [], "skipped": [str(directory)], "errors": []}

    for pattern in patterns:
        for file_path in directory.glob(pattern):
            try:
                if not file_path.is_file() or not _maintenance_is_within(directory, file_path):
                    continue
                if datetime.fromtimestamp(file_path.stat().st_mtime) >= cutoff_time:
                    continue
                file_path.unlink()
                deleted_files.append(str(file_path))
            except Exception as exc:
                errors.append({"file": str(file_path), "error": str(exc)})

    return {"deleted_files": deleted_files, "skipped": skipped, "errors": errors}


@router.post("/api/v4/maintenance/cleanup-temp-files")
def api_v4_maintenance_cleanup_temp_files():
    logger.info("V4 maintenance cleanup temp files requested")
    base_dir = get_base_dir().resolve()
    cutoff_time = datetime.now() - timedelta(days=7)
    cleanup_targets = [
        {
            "directory": base_dir / "v4" / "output",
            "patterns": ["*.xlsx"],
        },
        {
            "directory": base_dir / "output",
            "patterns": ["*.xlsx"],
        },
        {
            "directory": base_dir / "output" / "tmp_images",
            "patterns": ["*.png", "*.jpg", "*.jpeg", "*.webp", "*.bmp", "*.gif"],
        },
        {
            "directory": base_dir / "output" / "layout_cache",
            "patterns": ["*.png", "*.jpg", "*.jpeg", "*.webp", "*.bmp", "*.gif"],
        },
        {
            "directory": base_dir / "output" / "readback",
            "patterns": ["*.xlsx", "*.json", "*.tmp"],
        },
        {
            "directory": base_dir / "v4" / "output" / "readback",
            "patterns": ["*.xlsx", "*.json", "*.tmp"],
        },
    ]
    deleted_files = []
    skipped = []
    errors = []

    for target in cleanup_targets:
        directory = target["directory"].resolve()
        if not _maintenance_is_within(base_dir, directory):
            skipped.append(str(directory))
            continue
        result = _maintenance_cleanup_old_files(directory, target["patterns"], cutoff_time)
        deleted_files.extend(result.get("deleted_files", []))
        skipped.extend(result.get("skipped", []))
        errors.extend(result.get("errors", []))

    relative_deleted = []
    for file_path in deleted_files[:20]:
        try:
            relative_deleted.append(str(Path(file_path).resolve().relative_to(base_dir)))
        except ValueError:
            relative_deleted.append(str(file_path))

    return {
        "success": not bool(errors),
        "deleted_count": len(deleted_files),
        "deleted_files": relative_deleted,
        "skipped": skipped,
        "errors": errors[:20],
        "retention_days": 7,
    }


@router.post("/api/v4/load-order-object")
def api_v4_load_order_object():
    logger.info("V4 load order object into pipeline requested")
    order_object_path = get_base_dir() / "v4" / "examples" / "order_object.json"
    if not order_object_path.is_file():
        return {
            "success": False,
            "error": "v4/examples/order_object.json 不存在",
            "pipeline_state": get_pipeline_state(),
        }

    try:
        with order_object_path.open("r", encoding="utf-8") as f:
            order_object = json.load(f)
    except Exception as exc:
        logger.warning("V4 order object load into pipeline failed: path=%s", order_object_path, exc_info=True)
        return {
            "success": False,
            "error": f"Order Object 读取失败：{exc}",
            "pipeline_state": get_pipeline_state(),
        }

    state = load_order_object_into_pipeline(order_object if isinstance(order_object, dict) else {})
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        profile = get_current_template_profile()
        if profile:
            state = set_current_profile(profile)

    return {
        "success": True,
        "message": "Order Object 已加载",
        "pipeline_state": state,
    }


@router.post("/api/v4/load-order-object-from-payload")
def api_v4_load_order_object_from_payload(payload: Any = Body(None)):
    logger.info("V4 load order object from payload requested")

    payload = payload if isinstance(payload, dict) else {}
    order_object = payload.get("order_object") if isinstance(payload.get("order_object"), dict) else payload

    if not isinstance(order_object, dict) or not order_object:
        return {
            "success": False,
            "error": "Order Object 不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    state = load_order_object_into_pipeline(order_object)
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        profile = get_current_template_profile()
        if profile:
            state = set_current_profile(profile)

    return {
        "success": True,
        "message": "Order Object 已从请求数据加载",
        "order_object_keys": list(order_object.keys()),
        "pipeline_state": state,
    }


@router.post("/api/v4/normalize-flat-order")
def api_v4_normalize_flat_order(payload: Any = Body(None)):
    logger.info("V4 normalize flat order requested")

    payload = payload if isinstance(payload, dict) else {}
    flat_data = payload.get("data") if isinstance(payload.get("data"), dict) else payload

    if not isinstance(flat_data, dict) or not flat_data:
        return {
            "success": False,
            "error": "flat order data 不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    normalized = normalize_flat_order_to_v4_order_object(flat_data)
    order_object = normalized.get("order_object") if isinstance(normalized, dict) else {}
    if not isinstance(order_object, dict) or not order_object:
        return {
            "success": False,
            "error": "Order Object 转换失败",
            "normalized": normalized,
            "pipeline_state": get_pipeline_state(),
        }

    state = load_order_object_into_pipeline(order_object)
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        profile = get_current_template_profile()
        if profile:
            state = set_current_profile(profile)

    return {
        "success": True,
        "message": "Flat order data 已转换为 V4 Order Object 并加载",
        "warnings": normalized.get("warnings", []),
        "source_keys": normalized.get("source_keys", []),
        "order_object": order_object,
        "pipeline_state": state,
    }


@router.post("/api/v4/parse-chat-to-order-object")
def api_v4_parse_chat_to_order_object(payload: Any = Body(None)):
    logger.info("V4 parse chat to order object requested")

    payload = payload if isinstance(payload, dict) else {}
    message = payload.get("chat_text")
    if message is None:
        message = payload.get("message", "")
    message = str(message or "")

    if not message.strip():
        return {
            "success": False,
            "error": "message不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    chat_preprocess = preprocess_chat_text(message)
    clean_message = str(chat_preprocess.get("clean_text") or "").strip()
    preprocess_payload = {
        "stats": chat_preprocess.get("stats") or {},
        "removed_lines": chat_preprocess.get("removed_lines") or [],
    }

    if not clean_message:
        return {
            "success": False,
            "error": "message 清洗后为空，无法解析",
            "chat_preprocess": preprocess_payload,
            "pipeline_state": get_pipeline_state(),
        }

    state = get_pipeline_state()
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        current_profile = get_current_template_profile() or {}
    workspace_fields = _build_workspace_fields_from_profile(current_profile)
    semantic_workspace_schema = _build_semantic_workspace_schema(current_profile)
    runtime_mapping_source = _get_runtime_mapping_source(current_profile)
    excel_feature_flags = _get_excel_feature_flags(current_profile)
    extraction_contract = _build_ai_extraction_contract_from_workspace_fields(workspace_fields)
    extraction_contract_summary = _ai_extraction_contract_summary(extraction_contract)

    parsed = parse_message(clean_message, extraction_contract=extraction_contract)
    if isinstance(parsed, dict) and parsed.get("error"):
        return {
            "success": False,
            "error": parsed.get("error"),
            "parsed": parsed,
            "last_extraction_contract": extraction_contract.get("fields", []),
            "extraction_contract": extraction_contract,
            "ai_extraction_contract_summary": extraction_contract_summary,
            "runtime_mapping_source": runtime_mapping_source,
            "excel_feature_flags": excel_feature_flags,
            "semantic_workspace_schema": semantic_workspace_schema,
            "workspace_fields": workspace_fields,
            "chat_preprocess": preprocess_payload,
            "pipeline_state": get_pipeline_state(),
        }

    if not isinstance(parsed, dict) or not parsed:
        return {
            "success": False,
            "error": "AI parse 未返回有效字段",
            "parsed": parsed,
            "last_extraction_contract": extraction_contract.get("fields", []),
            "extraction_contract": extraction_contract,
            "ai_extraction_contract_summary": extraction_contract_summary,
            "runtime_mapping_source": runtime_mapping_source,
            "excel_feature_flags": excel_feature_flags,
            "semantic_workspace_schema": semantic_workspace_schema,
            "workspace_fields": workspace_fields,
            "chat_preprocess": preprocess_payload,
            "pipeline_state": get_pipeline_state(),
        }

    field_binding_result = _bind_parsed_fields_to_template_cells(parsed, current_profile)

    normalized = normalize_flat_order_to_v4_order_object(parsed)
    order_object = normalized.get("order_object") if isinstance(normalized, dict) else {}
    if not isinstance(order_object, dict) or not order_object:
        return {
            "success": False,
            "error": "Order Object 转换失败",
            "parsed": parsed,
            "normalized": normalized,
            "last_extraction_contract": extraction_contract.get("fields", []),
            "extraction_contract": extraction_contract,
            "ai_extraction_contract_summary": extraction_contract_summary,
            "runtime_mapping_source": runtime_mapping_source,
            "excel_feature_flags": excel_feature_flags,
            "semantic_workspace_schema": semantic_workspace_schema,
            "confirmed_cells": field_binding_result.get("confirmed_cells", []),
            "field_bound_operations": field_binding_result.get("operations", []),
            "workspace_fields": workspace_fields,
            "chat_preprocess": preprocess_payload,
            "pipeline_state": get_pipeline_state(),
        }

    state = load_order_object_into_pipeline(order_object)
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        profile = get_current_template_profile()
        if profile:
            state = set_current_profile(profile)

    return {
        "success": True,
        "message": "Chat 已解析为 V4 Order Object 并加载",
        "parsed": parsed,
        "last_extraction_contract": extraction_contract.get("fields", []),
        "extraction_contract": extraction_contract,
        "ai_extraction_contract_summary": extraction_contract_summary,
        "runtime_mapping_source": runtime_mapping_source,
        "excel_feature_flags": excel_feature_flags,
        "semantic_workspace_schema": semantic_workspace_schema,
        "confirmed_cells": field_binding_result.get("confirmed_cells", []),
        "field_bound_operations": field_binding_result.get("operations", []),
        "workspace_fields": workspace_fields,
        "warnings": normalized.get("warnings", []),
        "source_keys": normalized.get("source_keys", []),
        "order_object": order_object,
        "chat_preprocess": preprocess_payload,
        "pipeline_state": state,
    }


@router.post("/api/v4/parse-chat-run-pipeline")
def api_v4_parse_chat_run_pipeline(payload: Any = Body(None)):
    logger.info("V4 parse chat and run pipeline requested")

    parse_result = api_v4_parse_chat_to_order_object(payload)
    if not parse_result.get("success"):
        return {
            "success": False,
            "stage": "parse_chat_to_order_object",
            "error": parse_result.get("error", "Chat 解析为 Order Object 失败"),
            "parse_result": parse_result,
            "pipeline_state": get_pipeline_state(),
        }

    pipeline_result = api_v4_core_pipeline_run()
    if not pipeline_result.get("success"):
        return {
            "success": False,
            "stage": "core_pipeline",
            "error": pipeline_result.get("error", "Core Pipeline 执行失败"),
            "parse_result": parse_result,
            "pipeline_result": pipeline_result,
            "pipeline_state": get_pipeline_state(),
        }

    field_bound_operations = parse_result.get("field_bound_operations", [])
    workspace_fields = parse_result.get("workspace_fields", [])
    semantic_workspace_schema = parse_result.get("semantic_workspace_schema", {})
    runtime_mapping_source = parse_result.get("runtime_mapping_source", {})
    excel_feature_flags = parse_result.get("excel_feature_flags", _get_excel_feature_flags({}))
    field_binding_merge = _merge_field_bound_operations(
        pipeline_result.get("processed_operations", []),
        field_bound_operations,
    )
    if field_bound_operations:
        from app.v4_render_preview import build_render_preview

        processed_operations = field_binding_merge.get("processed_operations", [])
        pipeline_result["processed_operations"] = processed_operations
        pipeline_result["field_bound_operation_count"] = len(field_bound_operations)
        pipeline_result["field_bound_added_count"] = field_binding_merge.get("added_count", 0)
        pipeline_result["field_bound_override_count"] = field_binding_merge.get("override_count", 0)
        state = get_pipeline_state()
        set_pipeline_result(processed_operations, pipeline_result.get("stages", []))
        render_preview = build_render_preview(
            processed_operations,
            state.get("mapping_safety", {}),
            state.get("current_template_path"),
        )
        set_render_preview(render_preview)
        pipeline_result["render_preview"] = render_preview

    return {
        "success": True,
        "message": "Chat 已解析并完成 V4 Core Pipeline",
        "parse_result": {
            "parsed": parse_result.get("parsed", {}),
            "warnings": parse_result.get("warnings", []),
            "source_keys": parse_result.get("source_keys", []),
            "order_object": parse_result.get("order_object", {}),
            "last_extraction_contract": parse_result.get("last_extraction_contract", []),
            "extraction_contract": parse_result.get("extraction_contract", {}),
            "ai_extraction_contract_summary": parse_result.get("ai_extraction_contract_summary", {}),
            "runtime_mapping_source": runtime_mapping_source,
            "excel_feature_flags": excel_feature_flags,
            "confirmed_cells": parse_result.get("confirmed_cells", []),
            "field_bound_operations": field_bound_operations,
            "workspace_fields": workspace_fields,
            "semantic_workspace_schema": semantic_workspace_schema,
            "chat_preprocess": parse_result.get("chat_preprocess", {}),
        },
        "pipeline_result": {
            "validation": pipeline_result.get("validation", {}),
            "warnings": pipeline_result.get("warnings", []),
            "mapping_counts": pipeline_result.get("mapping_counts", {}),
            "structured_operations": pipeline_result.get("structured_operations", []),
            "table_operations": pipeline_result.get("table_operations", []),
            "block_operations": pipeline_result.get("block_operations", []),
            "processed_operations": pipeline_result.get("processed_operations", []),
            "field_bound_operation_count": pipeline_result.get("field_bound_operation_count", len(field_bound_operations)),
            "field_bound_added_count": pipeline_result.get("field_bound_added_count", 0),
            "field_bound_override_count": pipeline_result.get("field_bound_override_count", 0),
            "workspace_fields": workspace_fields,
            "render_preview": pipeline_result.get("render_preview", {}),
            "render_ready": pipeline_result.get("render_ready", False),
        },
        "workspace_fields": workspace_fields,
        "semantic_workspace_schema": semantic_workspace_schema,
        "runtime_mapping_source": runtime_mapping_source,
        "excel_feature_flags": excel_feature_flags,
        "ai_extraction_contract_summary": parse_result.get("ai_extraction_contract_summary", {}),
        "pipeline_state": get_pipeline_state(),
    }


@router.post("/api/v4/core-pipeline/run")
def api_v4_core_pipeline_run():
    from app.v4_pipeline_executor import run_operation_pipeline

    logger.info("V4 core pipeline run requested")
    state = get_pipeline_state()
    order_object = state.get("current_order_object") if isinstance(state.get("current_order_object"), dict) else {}
    if not order_object:
        load_result = api_v4_load_order_object()
        if not load_result.get("success"):
            return {
                "success": False,
                "error": load_result.get("error", "请先加载 Order Object"),
                "pipeline_state": get_pipeline_state(),
            }
        state = get_pipeline_state()
        order_object = state.get("current_order_object") if isinstance(state.get("current_order_object"), dict) else {}
    
    current_profile = state.get("current_profile") if isinstance(state.get("current_profile"), dict) else {}
    if not current_profile:
        profile = get_current_template_profile()
        state = set_current_profile(profile)
        current_profile = state.get("current_profile", {})
    
    current_template_path = state.get("current_template_path") if isinstance(state.get("current_template_path"), str) else None

    result = run_operation_pipeline(order_object, profile=current_profile, template_path=current_template_path)
    validation = result.get("validation", {})
    set_validator_result(validation)
    if not result.get("success"):
        return {
            "success": False,
            "error": "Validator 校验失败",
            "validation": validation,
            "pipeline_state": get_pipeline_state(),
        }

    structured_ops = result.get("structured_operations", [])
    set_structured_operations(structured_ops)

    table_ops = result.get("table_operations", [])
    set_table_operations(table_ops)

    block_ops = result.get("block_operations", [])
    set_block_operations(block_ops)

    unified_ops = result.get("unified_operations", [])
    set_unified_operations(unified_ops)

    processed_ops = result.get("processed_operations", [])
    stages = result.get("stages", [])
    mapping_safety = result.get("mapping_safety", {})
    set_mapping_safety(mapping_safety)
    mapping_counts = result.get("mapping_counts", {})
    set_mapping_counts(mapping_counts)
    set_pipeline_result(processed_ops, stages)
    set_render_preview(result.get("render_preview", {}))

    return {
        "success": True,
        "validation": validation,
        "warnings": result.get("warnings", []),
        "structured_operations": structured_ops,
        "table_operations": table_ops,
        "block_operations": block_ops,
        "unified_operations": unified_ops,
        "processed_operations": processed_ops,
        "mapping_safety": mapping_safety,
        "mapping_counts": mapping_counts,
        "stages": stages,
        "render_ready": result.get("render_ready", False),
        "render_preview": result.get("render_preview", {}),
        "pipeline_state": get_pipeline_state(),
    }


@router.post("/api/v4/parse-chat-export-excel")
def api_v4_parse_chat_export_excel(
    chat_text: str = Form(""),
    message: str = Form(""),
):
    from app.v4_excel_executor import execute_processed_operations_to_excel
    from app.v4_render_preview import build_render_preview
    from app.v4_render_targets import render_preview_to_html

    logger.info("V4 parse chat and export Excel requested")

    text = str(chat_text or "").strip()
    if not text:
        text = str(message or "").strip()
    if not text:
        return {
            "success": False,
            "stage": "input",
            "error": "chat_text/message 不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    template_path = None
    template_source = ""
    try:
        pipeline_e2e_result = api_v4_parse_chat_run_pipeline({"chat_text": text})
        if not pipeline_e2e_result.get("success"):
            return {
                "success": False,
                "stage": pipeline_e2e_result.get("stage", "parse_chat_run_pipeline"),
                "error": pipeline_e2e_result.get("error", "Chat 到 Pipeline 执行失败"),
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }

        pipeline_result = pipeline_e2e_result.get("pipeline_result", {})
        processed_operations = pipeline_result.get("processed_operations", [])
        if not isinstance(processed_operations, list) or not processed_operations:
            return {
                "success": False,
                "stage": "processed_operations",
                "error": "暂无 processed operations，无法导出 Excel",
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }

        template_path, _, template_source = _resolve_export_template_source()

        export_result = execute_processed_operations_to_excel(template_path, processed_operations)
        if not export_result.get("success"):
            return {
                "success": False,
                "stage": "excel_export",
                "error": export_result.get("error", "Excel 导出失败"),
                "warnings": export_result.get("warnings", []),
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }

        state = merge_mapping_safety(export_result.get("mapping_safety", {}))
        merged_safety = state.get("mapping_safety", {})
        preview = build_render_preview(processed_operations, merged_safety, template_path)
        state = set_render_preview(preview)

        html_result = render_preview_to_html(state.get("render_preview", {}))
        html_preview = ""
        if html_result.get("success"):
            html_preview = html_result.get("html", "")
            state = set_render_targets({"html_preview": html_preview})

        state = set_excel_result(export_result.get("filename"))

        return {
            "success": True,
            "message": "Chat 已解析并导出 Excel",
            "parse_result": pipeline_e2e_result.get("parse_result", {}),
            "pipeline_result": pipeline_result,
            "workspace_fields": pipeline_e2e_result.get("workspace_fields", []),
            "export_result": {
                "filename": export_result.get("filename", ""),
                "download_url": export_result.get("download_url", ""),
                "operations_written": export_result.get("operations_written", 0),
                "warnings": export_result.get("warnings", []),
                "template_source": template_source,
            },
            "render_preview": get_pipeline_state().get("render_preview", {}),
            "html_preview": get_pipeline_state().get("render_targets", {}).get("html_preview", html_preview),
            "mapping_safety": get_pipeline_state().get("mapping_safety", {}),
            "pipeline_state": get_pipeline_state(),
        }
    except ValueError as exc:
        return {
            "success": False,
            "stage": "template_upload",
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 parse chat and export Excel failed")
        return {
            "success": False,
            "stage": "unknown",
            "error": str(exc) or "Chat 到 Excel 导出失败",
            "pipeline_state": get_pipeline_state(),
        }

def _confirmed_cell_is_image_item(item):
    if not isinstance(item, dict):
        return False
    if str(item.get("field_type") or "").strip() == "image":
        return True
    image = item.get("image")
    return isinstance(image, dict) and bool(str(image.get("data_url") or "").strip())


def _split_confirmed_cells_for_excel_export(confirmed_cells):
    text_items = []
    image_items = []
    for item in confirmed_cells if isinstance(confirmed_cells, list) else []:
        if _confirmed_cell_is_image_item(item):
            image_items.append(item)
        else:
            text_items.append(item)
    return text_items, image_items


def _merge_append_after_colon_confirmed_cells_for_export(confirmed_cells, profile=None):
    items = confirmed_cells if isinstance(confirmed_cells, list) else []
    if not items:
        return []

    config_lookup = _confirmed_config_lookup_from_profile(profile if isinstance(profile, dict) else {})
    enriched_items = []
    append_groups = {}

    for index, item in enumerate(items):
        if not isinstance(item, dict):
            enriched_items.append((index, item, "", None))
            continue

        config = _lookup_confirmed_mapping_config(item, config_lookup)
        enriched_item = _confirmed_item_with_mapping_config(item, config)
        write_mode = str(enriched_item.get("write_mode") or "").strip()
        merge_key = _operation_merge_key(enriched_item)
        if write_mode != "append_after_colon" or not merge_key:
            enriched_items.append((index, enriched_item, "", None))
            continue

        value_text = str(enriched_item.get("value") or "").strip()
        group = append_groups.setdefault(
            merge_key,
            {
                "first_index": index,
                "items": [],
            },
        )
        group["items"].append(enriched_item)
        enriched_items.append((index, enriched_item, merge_key, group))

    merged_items = []
    emitted_append_keys = set()
    for _, item, merge_key, group in enriched_items:
        if not merge_key or not group:
            if isinstance(item, dict):
                merged_items.append(item)
            continue

        if merge_key in emitted_append_keys:
            continue
        emitted_append_keys.add(merge_key)

        group_items = [entry for entry in group.get("items", []) if isinstance(entry, dict)]
        if len(group_items) <= 1:
            merged_items.extend(group_items)
            continue

        merged = deepcopy(group_items[0])
        value_lines = []
        for entry in group_items:
            value_text = str(entry.get("value") or "").strip()
            if not value_text:
                continue
            label = str(entry.get("label") or entry.get("field_key") or "").strip()
            if label and label not in value_text:
                value_lines.append(f"{label}: {value_text}")
            else:
                value_lines.append(value_text)

        merged["value"] = "\n".join(value_lines)
        merged["merged_append_after_colon"] = True
        merged["merged_field_keys"] = [
            str(entry.get("field_key") or "").strip()
            for entry in group_items
            if str(entry.get("field_key") or "").strip()
        ]
        merged_items.append(merged)

    return merged_items


def _image_extension_from_mime_type(mime_type):
    mime = str(mime_type or "").strip().lower()
    if mime == "image/png":
        return ".png"
    if mime in {"image/jpeg", "image/jpg"}:
        return ".jpg"
    if mime == "image/webp":
        return ".webp"
    return ""


def _extract_image_file_from_confirmed_item(item, tmp_dir):
    image = item.get("image") if isinstance(item, dict) else {}
    image = image if isinstance(image, dict) else {}

    data_url = str(image.get("data_url") or "").strip()
    mime_type = str(image.get("mime_type") or "").strip().lower()

    if not data_url:
        return None, "图片 data_url 为空"

    if "," not in data_url or not data_url.startswith("data:"):
        return None, "图片 data_url 格式无效"

    header, encoded = data_url.split(",", 1)
    header_mime = header[5:].split(";")[0].strip().lower()
    mime_type = mime_type or header_mime

    ext = _image_extension_from_mime_type(mime_type)
    if not ext:
        return None, f"不支持的图片类型：{mime_type or 'unknown'}"

    try:
        raw = base64.b64decode(encoded)
    except Exception as exc:
        return None, f"图片 base64 解码失败：{exc}"

    if not raw:
        return None, "图片内容为空"

    tmp_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}{ext}"
    image_path = tmp_dir / filename
    image_path.write_bytes(raw)
    return image_path, ""


def _fit_openpyxl_image_contain(openpyxl_image, max_width=220, max_height=160):
    try:
        width = float(openpyxl_image.width or 0)
        height = float(openpyxl_image.height or 0)
        if width <= 0 or height <= 0:
            return
        ratio = min(float(max_width) / width, float(max_height) / height, 1.0)
        openpyxl_image.width = int(width * ratio)
        openpyxl_image.height = int(height * ratio)
    except Exception:
        return


def _target_cell_for_image_item(item):
    image = item.get("image") if isinstance(item, dict) else {}
    image = image if isinstance(image, dict) else {}
    return _cell_key(
        item.get("image_anchor_cell")
        or image.get("image_anchor_cell")
        or item.get("cell")
        or item.get("display_cell")
    )


def _target_sheet_for_image_item(item):
    image = item.get("image") if isinstance(item, dict) else {}
    image = image if isinstance(image, dict) else {}
    return _sheet_key(
        item.get("target_sheet")
        or item.get("sheet_name")
        or item.get("worksheet")
        or item.get("sheet")
        or image.get("target_sheet")
        or image.get("sheet_name")
        or image.get("worksheet")
        or image.get("sheet")
    )


def _insert_confirmed_images_into_excel(exported_file_path, confirmed_cells, excel_feature_flags=None):
    summary = {
        "total": 0,
        "inserted": 0,
        "skipped": 0,
        "warnings": [],
        "disabled": False,
    }

    flags = excel_feature_flags if isinstance(excel_feature_flags, dict) else {}
    if flags.get("image_fields") is not True:
        summary["disabled"] = True
        return summary

    image_items = [
        item for item in confirmed_cells if _confirmed_cell_is_image_item(item)
    ] if isinstance(confirmed_cells, list) else []
    summary["total"] = len(image_items)

    if not image_items:
        return summary

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception as exc:
        summary["skipped"] = len(image_items)
        summary["warnings"].append(f"openpyxl 图片模块不可用：{exc}")
        return summary

    exported_path = Path(exported_file_path)
    if not exported_path.exists():
        summary["skipped"] = len(image_items)
        summary["warnings"].append(f"导出文件不存在，无法插入图片：{exported_path}")
        return summary

    tmp_dir = Path("output") / "_tmp_images"

    try:
        workbook = load_workbook(exported_path)
        sheet = workbook.active
    except Exception as exc:
        summary["skipped"] = len(image_items)
        summary["warnings"].append(f"打开导出 Excel 失败：{exc}")
        return summary

    changed = False
    for item in image_items:
        label = str(item.get("label") or item.get("field_key") or "").strip()
        target_cell = _target_cell_for_image_item(item)

        if not target_cell:
            summary["skipped"] += 1
            summary["warnings"].append(f"图片字段缺少目标单元格：{label or '未命名图片字段'}")
            continue

        image_path, error = _extract_image_file_from_confirmed_item(item, tmp_dir)
        if error:
            summary["skipped"] += 1
            summary["warnings"].append(f"{target_cell} 图片处理失败：{error}")
            continue

        try:
            target_sheet = _target_sheet_for_image_item(item)
            if target_sheet and target_sheet in workbook.sheetnames:
                image_sheet = workbook[target_sheet]
            else:
                image_sheet = sheet

            excel_image = OpenpyxlImage(str(image_path))
            _fit_openpyxl_image_contain(excel_image, max_width=220, max_height=160)
            image_sheet.add_image(excel_image, target_cell)
            summary["inserted"] += 1
            changed = True
        except Exception as exc:
            summary["skipped"] += 1
            summary["warnings"].append(f"{target_cell} 图片插入失败：{exc}")

    if changed:
        try:
            workbook.save(exported_path)
        except Exception as exc:
            summary["warnings"].append(f"保存包含图片的 Excel 失败：{exc}")

    return summary


def _count_workbook_images(workbook_path):
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path)
        try:
            return sum(len(getattr(sheet, "_images", [])) for sheet in workbook.worksheets), ""
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()
    except Exception as exc:
        return 0, str(exc)


def _operation_image_export_summary(template_path, exported_file_path, operations, confirmed_cells):
    image_items = [
        item for item in confirmed_cells if _confirmed_cell_is_image_item(item)
    ] if isinstance(confirmed_cells, list) else []
    image_operations = [
        operation
        for operation in operations
        if isinstance(operation, dict) and operation.get("op_type") == "write_image"
    ] if isinstance(operations, list) else []

    total = len(image_items) or len(image_operations)
    summary = {
        "total": total,
        "inserted": 0,
        "skipped": 0,
        "warnings": [],
        "disabled": False,
        "source": "operation_image_export",
    }
    if not total:
        return summary

    template_image_count, template_error = _count_workbook_images(template_path)
    exported_image_count, exported_error = _count_workbook_images(exported_file_path)
    if exported_error:
        executable_image_ops = [
            operation
            for operation in image_operations
            if str(operation.get("image_path") or "").strip()
        ]
        summary["inserted"] = min(total, len(executable_image_ops))
        summary["skipped"] = max(0, total - summary["inserted"])
        summary["warnings"].append(f"导出图片数量回读失败: {exported_error}")
        return summary

    inserted = exported_image_count
    if not template_error:
        inserted = max(0, exported_image_count - template_image_count)
    else:
        summary["warnings"].append(f"模板图片数量回读失败: {template_error}")

    summary["inserted"] = min(total, inserted)
    summary["skipped"] = max(0, total - summary["inserted"])

    missing_image_path_count = sum(
        1 for operation in image_operations
        if not str(operation.get("image_path") or "").strip()
    )
    if missing_image_path_count:
        summary["warnings"].append(
            f"{missing_image_path_count} 个图片 operation 缺少 image_path"
        )

    return summary


@router.post("/api/v4/export-confirmed-excel")
def api_v4_export_confirmed_excel(
    chat_text: str = Form(""),
    confirmed_cells_json: str = Form("[]"),
):
    from app.v4_excel_executor import execute_processed_operations_to_excel
    from app.v4_render_preview import build_render_preview
    from app.v4_render_targets import render_preview_to_html

    logger.info("V4 confirmed cells export requested")

    text = str(chat_text or "").strip()
    if not text:
        return {
            "success": False,
            "stage": "input",
            "error": "chat_text 不能为空",
            "pipeline_state": get_pipeline_state(),
        }

    try:
        confirmed_cells = json.loads(str(confirmed_cells_json or "[]"))
    except json.JSONDecodeError as exc:
        return {
            "success": False,
            "stage": "confirmed_cells_json",
            "error": f"confirmed_cells_json 解析失败: {exc}",
            "pipeline_state": get_pipeline_state(),
        }
    if not isinstance(confirmed_cells, list):
        return {
            "success": False,
            "stage": "confirmed_cells_json",
            "error": "confirmed_cells_json 必须是 list",
            "pipeline_state": get_pipeline_state(),
        }

    template_path = None
    template_source = ""
    try:
        pipeline_e2e_result = api_v4_parse_chat_run_pipeline({"chat_text": text})
        if not pipeline_e2e_result.get("success"):
            return {
                "success": False,
                "stage": pipeline_e2e_result.get("stage", "parse_chat_run_pipeline"),
                "error": pipeline_e2e_result.get("error", "Chat 到 Pipeline 执行失败"),
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }

        pipeline_result = pipeline_e2e_result.get("pipeline_result", {})
        processed_operations = pipeline_result.get("processed_operations", [])
        if not isinstance(processed_operations, list):
            return {
                "success": False,
                "stage": "processed_operations",
                "error": "暂无 processed operations，无法导出 Excel",
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }
        confirmed_has_items = any(isinstance(item, dict) for item in confirmed_cells)
        if not processed_operations and not confirmed_has_items:
            return {
                "success": False,
                "stage": "processed_operations",
                "error": "鏆傛棤 processed operations 鎴?confirmed cells锛屾棤娉曞鍑?Excel",
                "pipeline_e2e_result": pipeline_e2e_result,
                "pipeline_state": get_pipeline_state(),
            }

        profile = _current_template_profile_for_export()
        runtime_mapping_source = _get_runtime_mapping_source(profile)
        excel_feature_flags = _get_excel_feature_flags(profile)
        template_path, _, template_source = _resolve_export_template_source()

        text_confirmed_cells, image_confirmed_cells = _split_confirmed_cells_for_excel_export(confirmed_cells)

        use_operation_image_export = True
        override_result = _override_operations_with_confirmed_cells(
            processed_operations,
            confirmed_cells,
            profile=profile,
            template_path=template_path,
        )
        overridden_operations = override_result.get("processed_operations", [])
        confirmed_override_count = override_result.get("override_count", 0)
        confirmed_added_count = override_result.get("added_count", 0)
        write_mode_summary = override_result.get("write_mode_summary", _confirmed_export_empty_summary())
        if not overridden_operations:
            return {
                "success": False,
                "stage": "processed_operations",
                "error": "鏆傛棤鍙墽琛岀殑 processed operations 鎴?confirmed cells锛屾棤娉曞鍑?Excel",
                "pipeline_e2e_result": pipeline_e2e_result,
                "confirmed_override_count": confirmed_override_count,
                "confirmed_added_count": confirmed_added_count,
                "write_mode_summary": write_mode_summary,
                "pipeline_state": get_pipeline_state(),
            }

        export_result = execute_processed_operations_to_excel(template_path, overridden_operations)
        if not export_result.get("success"):
            return {
                "success": False,
                "stage": "excel_export",
                "error": export_result.get("error", "Excel 导出失败"),
                "warnings": export_result.get("warnings", []),
                "pipeline_e2e_result": pipeline_e2e_result,
                "confirmed_override_count": confirmed_override_count,
                "confirmed_added_count": confirmed_added_count,
                "write_mode_summary": write_mode_summary,
                "pipeline_state": get_pipeline_state(),
            }

        exported_filename = str(export_result.get("filename") or "").strip()
        exported_file_path = Path(exported_filename)
        if not exported_file_path.is_absolute():
            if exported_file_path.parts and exported_file_path.parts[0] == "output":
                exported_file_path = exported_file_path
            else:
                exported_file_path = Path("output") / exported_file_path
        if image_confirmed_cells and use_operation_image_export:
            image_export_summary = _operation_image_export_summary(
                template_path,
                exported_file_path,
                overridden_operations,
                image_confirmed_cells,
            )
        elif image_confirmed_cells:
            image_export_summary = _insert_confirmed_images_into_excel(
                exported_file_path,
                image_confirmed_cells,
                excel_feature_flags=excel_feature_flags,
            )
        else:
            image_export_summary = {"total": 0, "inserted": 0, "skipped": 0, "warnings": []}
        if image_export_summary.get("warnings"):
            export_result["warnings"] = [
                *(export_result.get("warnings", []) or []),
                *image_export_summary.get("warnings", []),
            ]

        if excel_feature_flags.get("export_readback_check", True):
            export_readback_audit = _build_export_readback_audit(
                exported_file_path,
                text_confirmed_cells,
                profile=profile,
            )
        else:
            export_readback_audit = {
                "success": True,
                "disabled": True,
                "summary": {},
                "items": [],
                "warnings": [],
                "errors": [],
            }

        set_pipeline_result(overridden_operations, pipeline_result.get("stages", []))
        state = merge_mapping_safety(export_result.get("mapping_safety", {}))
        merged_safety = state.get("mapping_safety", {})
        preview = build_render_preview(overridden_operations, merged_safety, template_path)
        state = set_render_preview(preview)

        html_result = render_preview_to_html(state.get("render_preview", {}))
        html_preview = ""
        if html_result.get("success"):
            html_preview = html_result.get("html", "")
            state = set_render_targets({"html_preview": html_preview})

        state = set_excel_result(export_result.get("filename"))
        response_pipeline_result = dict(pipeline_result)
        response_pipeline_result["processed_operations"] = overridden_operations
        response_pipeline_result["confirmed_override_count"] = confirmed_override_count
        response_pipeline_result["confirmed_added_count"] = confirmed_added_count
        response_pipeline_result["write_mode_summary"] = write_mode_summary
        response_pipeline_result["render_preview"] = get_pipeline_state().get("render_preview", {})

        return {
            "success": True,
            "message": "确认值已导出 Excel",
            "confirmed_override_count": confirmed_override_count,
            "confirmed_added_count": confirmed_added_count,
            "write_mode_summary": write_mode_summary,
            "runtime_mapping_source": runtime_mapping_source,
            "excel_feature_flags": excel_feature_flags,
            "image_export_summary": image_export_summary,
            "export_readback_audit": export_readback_audit,
            "parse_result": pipeline_e2e_result.get("parse_result", {}),
            "pipeline_result": response_pipeline_result,
            "export_result": {
                "filename": export_result.get("filename", ""),
                "download_url": export_result.get("download_url", ""),
                "operations_written": export_result.get("operations_written", 0),
                "warnings": export_result.get("warnings", []),
                "template_source": template_source,
                "write_mode_summary": write_mode_summary,
                "excel_feature_flags": excel_feature_flags,
                "image_export_summary": image_export_summary,
                "readback_audit": export_readback_audit,
            },
            "render_preview": get_pipeline_state().get("render_preview", {}),
            "html_preview": get_pipeline_state().get("render_targets", {}).get("html_preview", html_preview),
            "mapping_safety": get_pipeline_state().get("mapping_safety", {}),
            "pipeline_state": get_pipeline_state(),
        }
    except ValueError as exc:
        return {
            "success": False,
            "stage": "template_upload",
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 confirmed cells export failed")
        return {
            "success": False,
            "stage": "unknown",
            "error": str(exc) or "确认值导出 Excel 失败",
            "pipeline_state": get_pipeline_state(),
        }
@router.post("/api/v4/core-pipeline/export-excel")
def api_v4_core_pipeline_export_excel():
    from app.v4_excel_executor import execute_processed_operations_to_excel
    from app.v4_render_preview import build_render_preview
    from app.v4_render_targets import render_preview_to_html

    logger.info("V4 core pipeline export Excel requested")
    template_path = None
    template_source = ""
    try:
        state = get_pipeline_state()
        processed_operations = state.get("pipeline", {}).get("processed_operations", [])
        if not processed_operations:
            run_result = api_v4_core_pipeline_run()
            if not run_result.get("success"):
                return {
                    "success": False,
                    "error": run_result.get("error", "请先执行核心流水线"),
                    "pipeline_state": get_pipeline_state(),
                }
            processed_operations = run_result.get("processed_operations", [])

        template_path, _, template_source = _resolve_export_template_source()
        result = execute_processed_operations_to_excel(template_path, processed_operations)
        if not result.get("success"):
            return {
                "success": False,
                "error": result.get("error", "Excel 导出失败"),
                "warnings": result.get("warnings", []),
                "pipeline_state": get_pipeline_state(),
            }

        state = merge_mapping_safety(result.get("mapping_safety", {}))
        merged_safety = state.get("mapping_safety", {})
        preview = build_render_preview(processed_operations, merged_safety, template_path)
        state = set_render_preview(preview)
        html_result = render_preview_to_html(state.get("render_preview", {}))
        if html_result.get("success"):
            state = set_render_targets({"html_preview": html_result.get("html", "")})
        state = set_excel_result(result.get("filename"))
        return {
            "success": True,
            "filename": result.get("filename", ""),
            "download_url": result.get("download_url", ""),
            "operations_written": result.get("operations_written", 0),
            "warnings": result.get("warnings", []),
            "template_source": template_source,
            "mapping_safety": get_pipeline_state().get("mapping_safety", {}),
            "pipeline_state": state,
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 core pipeline export Excel failed")
        return {
            "success": False,
            "error": str(exc) or "Excel 导出失败",
            "pipeline_state": get_pipeline_state(),
        }
@router.post("/api/v4/render-preview/build")
def api_v4_render_preview_build(payload: Optional[Any] = Body(None)):
    from app.v4_render_preview import build_render_preview
    from app.v4_render_targets import render_preview_to_html

    logger.info("V4 render preview build requested")
    processed_operations = None
    if isinstance(payload, dict):
        processed_operations = payload.get("processed_operations")
    elif isinstance(payload, list):
        processed_operations = payload

    state = get_pipeline_state()
    if processed_operations is None:
        processed_operations = state.get("pipeline", {}).get("processed_operations", [])

    if not isinstance(processed_operations, list) or not processed_operations:
        return {
            "success": False,
            "error": "暂无 processed operations，无法生成 Render Preview",
            "render_preview": get_pipeline_state().get("render_preview", {}),
        }

    preview = build_render_preview(processed_operations, state.get("mapping_safety", {}), state.get("current_template_path"))
    state = set_render_preview(preview)
    saved_preview = state.get("render_preview", {})
    html_result = render_preview_to_html(saved_preview)
    if html_result.get("success"):
        state = set_render_targets({"html_preview": html_result.get("html", "")})
        saved_preview = state.get("render_preview", {})

    return {
        "success": True,
        "render_preview": saved_preview,
        "html_preview": html_result.get("html", ""),
        "warnings": preview.get("warnings", []),
        "pipeline_state": state,
    }


@router.get("/api/v4/render-preview")
def api_v4_render_preview():
    logger.info("V4 render preview requested")
    state = get_pipeline_state()
    return {
        "success": True,
        "render_preview": state.get("render_preview", {}),
        "html_preview": state.get("render_targets", {}).get("html_preview", ""),
    }


@router.post("/api/v4/template-analysis/upload")
def api_v4_template_analysis_upload(template_file: UploadFile = File(...)):
    logger.info("V4 template analysis upload requested: filename=%s", template_file.filename)
    try:
        template_path = _save_v4_uploaded_template(template_file)
        state = set_current_template(str(template_path))
        state = set_template_analysis(
            {
                "labels": [],
                "structured_mapping_preview": [],
                "table_regions": [],
                "block_regions": [],
                "template_structure": {
                    "regions": [],
                    "raw_regions": [],
                    "deduped_regions": [],
                    "recommended_regions": [],
                    "tables": [],
                    "blocks": [],
                    "labels": [],
                },
                "auto_mapping_preview": {
                    "structured": [],
                    "tables": [],
                    "blocks": [],
                    "needs_review": [],
                    "rejected_candidates": [],
                },
                "summary": {},
            }
        )
        return {
            "success": True,
            "message": "Excel 模板已上传，等待分析",
            "template_path": str(template_path),
            "filename": template_path.name,
            "pipeline_state": state,
        }
    except ValueError as exc:
        return {
            "success": False,
            "error": str(exc),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 template analysis upload failed")
        return {
            "success": False,
            "error": str(exc) or "Excel 模板上传失败",
            "pipeline_state": get_pipeline_state(),
        }


@router.get("/api/v4/template-analysis/run")
def api_v4_template_analysis_run():
    from app.v4_template_analysis import analyze_template

    logger.info("V4 template analysis run requested")
    state = get_pipeline_state()
    template_path = state.get("current_template_path")
    if not template_path:
        return {
            "success": False,
            "error": "请先上传 Excel 模板",
            "template_analysis": state.get("template_analysis", {}),
            "pipeline_state": state,
        }

    path = Path(str(template_path))
    if not path.is_file():
        return {
            "success": False,
            "error": "当前 Excel 模板不存在，请重新上传",
            "template_analysis": state.get("template_analysis", {}),
            "pipeline_state": state,
        }

    try:
        analysis = analyze_template(path)
        saved_state = set_template_analysis(analysis)
        return {
            "success": True,
            "template_analysis": saved_state.get("template_analysis", {}),
            "pipeline_state": saved_state,
        }
    except (BadZipFile, InvalidFileException) as exc:
        logger.warning("V4 template analysis invalid Excel: path=%s", path, exc_info=True)
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "template_analysis": get_pipeline_state().get("template_analysis", {}),
            "pipeline_state": get_pipeline_state(),
        }
    except Exception as exc:
        logger.exception("V4 template analysis failed: path=%s", path)
        return {
            "success": False,
            "error": str(exc) or "Template Analysis 执行失败",
            "template_analysis": get_pipeline_state().get("template_analysis", {}),
            "pipeline_state": get_pipeline_state(),
        }


@router.get("/api/v4/template-analysis/result")
def api_v4_template_analysis_result():
    logger.info("V4 template analysis result requested")
    state = get_pipeline_state()
    return {
        "success": True,
        "template_analysis": state.get("template_analysis", {}),
        "pipeline_state": state,
    }


@router.get("/api/v4/template-structure")
def api_v4_template_structure():
    logger.info("V4 template structure requested")
    state = get_pipeline_state()
    analysis = state.get("template_analysis", {})
    structure = analysis.get("template_structure", {}) if isinstance(analysis, dict) else {}
    return {
        "success": True,
        "template_structure": structure
        if isinstance(structure, dict)
        else {
            "regions": [],
            "raw_regions": [],
            "deduped_regions": [],
            "recommended_regions": [],
            "tables": [],
            "blocks": [],
            "labels": [],
        },
    }


@router.get("/api/v4/template-layout")
def api_v4_template_layout():
    logger.info("V4 template layout requested")
    state = get_pipeline_state()
    template_analysis = state.get("template_analysis", {})
    if not isinstance(template_analysis, dict):
        template_analysis = {}

    layout_result = build_layout_sections_from_template_analysis(template_analysis)
    template_analysis_summary = template_analysis.get("summary", {})
    if not isinstance(template_analysis_summary, dict):
        template_analysis_summary = {}
    profile = _current_template_profile_for_export()
    workspace_fields = _build_workspace_fields_from_profile(profile)
    semantic_workspace_schema = _build_semantic_workspace_schema(profile)
    runtime_mapping_source = _get_runtime_mapping_source(profile)
    excel_feature_flags = _get_excel_feature_flags(profile)

    return {
        "success": True,
        "layout_sections": layout_result.get("layout_sections", []),
        "workspace_fields": workspace_fields,
        "semantic_workspace_schema": semantic_workspace_schema,
        "runtime_mapping_source": runtime_mapping_source,
        "excel_feature_flags": excel_feature_flags,
        "summary": layout_result.get("summary", {}),
        "template_analysis_summary": template_analysis_summary,
        "pipeline_state": get_pipeline_state(),
    }


@router.get("/api/v4/auto-mapping-preview")
def api_v4_auto_mapping_preview():
    from app.v4_mapping_generator import generate_auto_mapping

    logger.info("V4 auto mapping preview requested")
    state = get_pipeline_state()
    analysis = state.get("template_analysis", {})
    if not isinstance(analysis, dict):
        analysis = {}
    preview = analysis.get("auto_mapping_preview")
    if not isinstance(preview, dict):
        preview = generate_auto_mapping(analysis)
    return {
        "success": True,
        "auto_mapping_preview": {
            "structured": preview.get("structured", []) if isinstance(preview.get("structured", []), list) else [],
            "tables": preview.get("tables", []) if isinstance(preview.get("tables", []), list) else [],
            "blocks": preview.get("blocks", []) if isinstance(preview.get("blocks", []), list) else [],
            "needs_review": preview.get("needs_review", [])
            if isinstance(preview.get("needs_review", []), list)
            else [],
            "rejected_candidates": preview.get("rejected_candidates", [])
            if isinstance(preview.get("rejected_candidates", []), list)
            else [],
        },
    }


def _recommended_auto_mapping_payload(preview):
    preview = preview if isinstance(preview, dict) else {}

    def recommended_items(key):
        items = preview.get(key, [])
        if not isinstance(items, list):
            return []
        return [
            item
            for item in items
            if isinstance(item, dict) and str(item.get("status") or "") == "recommended"
        ]

    return {
        "structured": recommended_items("structured"),
        "tables": recommended_items("tables"),
        "blocks": recommended_items("blocks"),
    }


@router.post("/api/v4/template-learn/run")
def api_v4_template_learn_run():
    from app.v4_mapping_generator import generate_auto_mapping
    from app.v4_mapping_workbench import save_auto_mapping_candidates
    from app.v4_template_analysis import analyze_template

    logger.info("V4 template learn run requested")
    state = get_pipeline_state()
    template_path = state.get("current_template_path")
    if not template_path:
        learning_state = set_template_learning(
            {
                "success": False,
                "summary": {},
                "needs_review": [],
                "rejected_candidates": [],
            }
        )
        return {
            "success": False,
            "error": "请先上传 Excel 模板",
            "pipeline_state": learning_state,
        }

    path = Path(str(template_path))
    if not path.is_file():
        learning_state = set_template_learning(
            {
                "success": False,
                "summary": {},
                "needs_review": [],
                "rejected_candidates": [],
            }
        )
        return {
            "success": False,
            "error": "当前 Excel 模板不存在，请重新上传",
            "pipeline_state": learning_state,
        }

    try:
        analysis = analyze_template(path)
        analysis["auto_mapping_preview"] = generate_auto_mapping(analysis)
        preview = analysis.get("auto_mapping_preview", {})
        payload = _recommended_auto_mapping_payload(preview)
        candidate_result = save_auto_mapping_candidates(payload)
        if not candidate_result.get("success"):
            raise ValueError(candidate_result.get("error", "自动映射候选保存失败"))

        summary = {
            "structured_candidates": candidate_result.get("result", {}).get("structured_candidates", 0),
            "table_candidates": candidate_result.get("result", {}).get("table_candidates", 0),
            "block_candidates": candidate_result.get("result", {}).get("block_candidates", 0),
            "needs_review": len(preview.get("needs_review", [])) if isinstance(preview.get("needs_review", []), list) else 0,
            "rejected": len(preview.get("rejected_candidates", []))
            if isinstance(preview.get("rejected_candidates", []), list)
            else 0,
        }

        set_template_analysis(analysis)
        learning_state = set_template_learning(
            {
                "success": True,
                "summary": summary,
                "needs_review": preview.get("needs_review", []),
                "rejected_candidates": preview.get("rejected_candidates", []),
            }
        )
        return {
            "success": True,
            "message": "模板学习完成，候选映射已保存，未写入正式 Mapping",
            "summary": summary,
            "auto_mapping_preview": preview,
            "template_analysis": learning_state.get("template_analysis", {}),
            "pipeline_state": learning_state,
        }
    except (BadZipFile, InvalidFileException) as exc:
        logger.warning("V4 template learn invalid Excel: path=%s", path, exc_info=True)
        learning_state = set_template_learning(
            {
                "success": False,
                "summary": {},
                "needs_review": [],
                "rejected_candidates": [],
            }
        )
        return {
            "success": False,
            "error": f"Excel 模板格式无效：{exc}",
            "pipeline_state": learning_state,
        }
    except Exception as exc:
        logger.exception("V4 template learn failed: path=%s", path)
        learning_state = set_template_learning(
            {
                "success": False,
                "summary": {},
                "needs_review": [],
                "rejected_candidates": [],
            }
        )
        return {
            "success": False,
            "error": str(exc) or "模板学习失败",
            "pipeline_state": learning_state,
        }


@router.get("/api/v4/mapping-workbench/preview")
def api_v4_mapping_workbench_preview():
    logger.info("V4 mapping workbench preview requested")
    return api_v4_auto_mapping_preview()


@router.post("/api/v4/mapping-workbench/save-selected")
def api_v4_mapping_workbench_save_selected(payload: Any = Body(None)):
    from app.v4_mapping_workbench import save_selected_mappings

    logger.info("V4 mapping workbench save selected requested")
    payload = payload if isinstance(payload, dict) else {}
    profile = {}
    profile_id = str(payload.get("profile_id") or "").strip()
    if profile_id:
        profile = load_template_profile(profile_id)
        if not profile:
            return {
                "success": False,
                "error": "Template Profile 不存在，无法保存 Mapping Rules",
            }

    result = save_selected_mappings(payload, profile=profile)
    if not result.get("success"):
        return {
            "success": False,
            "error": result.get("error", "映射规则保存失败"),
        }
    return {
        **result,
        "profile_id": profile.get("profile_id") if isinstance(profile, dict) else None,
        "profile_name": profile.get("profile_name") if isinstance(profile, dict) else None,
    }


@router.post("/api/v4/apply-auto-mapping")
def api_v4_apply_auto_mapping(payload: Any = Body(None)):
    from app.v4_mapping_workbench import apply_auto_mapping

    logger.info("V4 apply auto mapping requested")
    result = apply_auto_mapping(payload if isinstance(payload, dict) else {})
    if not result.get("success"):
        return {
            "success": False,
            "error": result.get("error", "自动映射候选保存失败"),
        }
    return result


@router.post("/api/v4/product-schema")
def api_v4_save_product_schema(schema: Any):
    result = save_product_schema(schema)
    if not result.get("success"):
        return {
            "success": False,
            "error": result.get("error", "V4 product schema save failed"),
        }

    return {
        "success": True,
        "data": result.get("data", {}),
    }


@router.get("/api/v4/output/{filename}")
def api_v4_download_output_file(filename: str):
    resolved = _resolve_v4_output_file(filename)
    if not resolved:
        logger.info("V4 output download not found or rejected: filename=%s", filename)
        return {
            "success": False,
            "error": "\u6587\u4ef6\u4e0d\u5b58\u5728",
        }

    output_path, media_type = resolved
    logger.info("V4 output download requested: path=%s", output_path)
    return FileResponse(
        path=str(output_path),
        filename=output_path.name,
        media_type=media_type,
    )


@router.post("/api/v4/template/fingerprint")
def api_v4_template_fingerprint(file: UploadFile = File(...)):
    temp_path = None
    try:
        logger.info("V4 template fingerprint requested: filename=%s", file.filename)
        temp_path = _save_v4_uploaded_template(file)
        fingerprint = build_template_fingerprint(temp_path)
        save_fingerprint(fingerprint)
        return {
            "success": True,
            "data": fingerprint,
            "fingerprint": fingerprint,
            "layout_hash": fingerprint.get("layout_hash", ""),
        }
    except ValueError as exc:
        logger.info("V4 template fingerprint rejected: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except (BadZipFile, InvalidFileException):
        logger.info("V4 template fingerprint rejected as non Excel: filename=%s", file.filename)
        return {
            "success": False,
            "error": "文件不是 Excel",
        }
    except Exception as exc:
        logger.exception("V4 template fingerprint failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"模板指纹生成失败：{exc}",
        }
    finally:
        _remove_v4_uploaded_template(temp_path)


@router.post("/api/v4/template/fingerprint-test")
def api_v4_template_fingerprint_test(file: Optional[UploadFile] = File(None)):
    temp_path = None
    if file is None:
        return {
            "success": False,
            "error": "上传文件为空",
        }

    try:
        logger.info("V4 template fingerprint stability test requested: filename=%s", file.filename)
        temp_path = _save_v4_uploaded_template(file)
        fingerprints = [build_template_fingerprint(temp_path) for _ in range(3)]
        hashes = [fingerprint.get("layout_hash", "") for fingerprint in fingerprints]
        all_equal = len(set(hashes)) == 1
        result = {
            "all_equal": all_equal,
            "hashes": hashes,
            "fingerprints": fingerprints,
        }
        if not all_equal:
            result["warning"] = "同一个 Excel 连续 3 次生成的 layout_hash 不一致"

        return {
            "success": True,
            "data": result,
            **result,
        }
    except ValueError as exc:
        logger.info("V4 template fingerprint stability test rejected: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except (BadZipFile, InvalidFileException):
        logger.info("V4 template fingerprint stability test rejected as non Excel: filename=%s", file.filename)
        return {
            "success": False,
            "error": "文件不是 Excel",
        }
    except Exception as exc:
        logger.exception("V4 template fingerprint stability test failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"模板指纹稳定性测试失败：{exc}",
        }
    finally:
        _remove_v4_uploaded_template(temp_path)


@router.post("/api/v4/template/match")
def api_v4_template_match(file: UploadFile = File(...)):
    temp_path = None
    try:
        logger.info("V4 template match requested: filename=%s", file.filename)
        temp_path = _save_v4_uploaded_template(file)
        result = match_template(temp_path)
        return {
            "success": True,
            "data": result,
            "cache_hit": result.get("cache_hit", False),
            "layout_hash": result.get("layout_hash", ""),
            "fingerprint": result.get("fingerprint"),
            "cached_rules": result.get("cached_rules"),
        }
    except ValueError as exc:
        logger.info("V4 template match rejected: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except (BadZipFile, InvalidFileException):
        logger.info("V4 template match rejected as non Excel: filename=%s", file.filename)
        return {
            "success": False,
            "error": "文件不是 Excel",
        }
    except Exception as exc:
        logger.exception("V4 template match failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"模板缓存检查失败：{exc}",
        }
    finally:
        _remove_v4_uploaded_template(temp_path)


@router.post("/api/v4/template/match-or-parse")
def api_v4_template_match_or_parse(file: Optional[UploadFile] = File(None)):
    temp_path = None
    if file is None:
        return {
            "success": False,
            "error": "上传文件为空",
        }

    try:
        logger.info("V4 template match-or-parse requested: filename=%s", file.filename)
        temp_path = _save_v4_uploaded_template(file)
        result = match_or_parse_template(temp_path)
        if result.get("success") is False:
            return {
                "success": False,
                "error": result.get("error", "AI 解析失败"),
                "cache_hit": result.get("cache_hit", False),
                "source": result.get("source", "ai_template_parser"),
                "layout_hash": result.get("layout_hash", ""),
                "fingerprint": result.get("fingerprint"),
                "rules": result.get("rules", []),
                "warnings": result.get("warnings", []),
                "meta": result.get("meta", {}),
            }

        return {
            "success": True,
            "data": result,
            "cache_hit": result.get("cache_hit", False),
            "source": result.get("source", ""),
            "layout_hash": result.get("layout_hash", ""),
            "fingerprint": result.get("fingerprint"),
            "rules": result.get("rules", []),
            "warnings": result.get("warnings", []),
            "meta": result.get("meta", {}),
        }
    except ValueError as exc:
        logger.info("V4 template match-or-parse rejected: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except (BadZipFile, InvalidFileException):
        logger.info("V4 template match-or-parse rejected as non Excel: filename=%s", file.filename)
        return {
            "success": False,
            "error": "文件不是 Excel",
        }
    except json.JSONDecodeError as exc:
        logger.exception("V4 template cache read failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"缓存读取失败：{exc}",
        }
    except OSError as exc:
        logger.exception("V4 template cache or rules save failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"rules 保存失败：{exc}",
        }
    except Exception as exc:
        logger.exception("V4 template match-or-parse failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"AI 解析失败：{exc}",
        }
    finally:
        _remove_v4_uploaded_template(temp_path)


@router.get("/api/v4/template/cache-list")
def api_v4_template_cache_list():
    try:
        templates = list_cached_templates()
        return {
            "success": True,
            "templates": templates,
        }
    except Exception as exc:
        logger.exception("V4 template cache list failed")
        return {
            "success": False,
            "error": f"已学习模板列表加载失败：{exc}",
            "templates": [],
        }


@router.get("/api/v4/template/cache-detail/{layout_hash}")
def api_v4_template_cache_detail(layout_hash: str):
    try:
        detail = get_cached_template_detail(layout_hash)
        return {
            "success": True,
            "detail": detail,
        }
    except FileNotFoundError:
        logger.info("V4 template cache detail not found: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": "模板缓存不存在",
        }
    except ValueError as exc:
        logger.info("V4 template cache detail rejected: layout_hash=%s error=%s", layout_hash, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except Exception as exc:
        logger.exception("V4 template cache detail failed: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": f"模板详情读取失败：{exc}",
        }


@router.delete("/api/v4/template/cache/{layout_hash}")
def api_v4_template_cache_delete(layout_hash: str):
    try:
        delete_cached_template(layout_hash)
        return {
            "success": True,
            "message": "模板缓存已删除",
        }
    except FileNotFoundError:
        logger.info("V4 template cache delete not found: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": "模板缓存不存在",
        }
    except ValueError as exc:
        logger.info("V4 template cache delete rejected: layout_hash=%s error=%s", layout_hash, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except Exception as exc:
        logger.exception("V4 template cache delete failed: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": f"模板缓存删除失败：{exc}",
        }


@router.post("/api/v4/template/cache-update/{layout_hash}")
def api_v4_template_cache_update(layout_hash: str, payload: Any = Body(None)):
    if not isinstance(payload, dict):
        payload = {}
    try:
        template_name = payload.get("template_name")
        template_note = payload.get("template_note")
        meta = update_template_info(layout_hash, template_name=template_name, template_note=template_note)
        return {
            "success": True,
            "meta": meta,
        }
    except FileNotFoundError:
        logger.info("V4 template cache update not found: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": "模板缓存不存在",
        }
    except ValueError as exc:
        logger.info("V4 template cache update rejected: layout_hash=%s error=%s", layout_hash, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except Exception as exc:
        logger.exception("V4 template cache update failed: layout_hash=%s", layout_hash)
        return {
            "success": False,
            "error": f"模板信息更新失败：{exc}",
        }


def _load_workbench_example_order(example_order_text: str = ""):
    if not isinstance(example_order_text, str):
        example_order_text = ""
    if str(example_order_text or "").strip():
        payload = json.loads(example_order_text)
        if not isinstance(payload, dict):
            raise ValueError("example_order 必须是 JSON object")
        return payload

    example = load_example("soft_capsule_order_example")
    if not example:
        raise ValueError("默认示例订单不存在")
    return example


def _build_workbench_rules_config(rules):
    if not isinstance(rules, list) or not rules:
        raise ValueError("rules 为空")

    template_key = "workbench_template"
    return template_key, {
        "version": "v4.19-workbench",
        "description": "Workbench generated temporary rules config",
        "templates": {
            template_key: {
                "label": "Workbench 智能模板规则",
                "rules": rules,
            }
        },
    }


@router.post("/api/v4/workbench/export")
def api_v4_workbench_export(
    file: Optional[UploadFile] = File(None),
    rules: str = Form(""),
    example_order: str = Form(""),
):
    temp_path = None
    if file is None:
        return {
            "success": False,
            "error": "尚未上传模板",
        }

    try:
        if not str(rules or "").strip():
            return {
                "success": False,
                "error": "尚未完成智能解析",
            }

        parsed_rules = json.loads(rules)
        if not isinstance(parsed_rules, list) or not parsed_rules:
            return {
                "success": False,
                "error": "rules 为空",
            }

        logger.info("V4 workbench export requested: filename=%s rules=%s", file.filename, len(parsed_rules))
        temp_path = _save_v4_uploaded_template(file)
        example = _load_workbench_example_order(example_order)
        template_key, rules_config = _build_workbench_rules_config(parsed_rules)

        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "Excel 生成失败：Renderer 失败"),
            }

        preview_result = build_excel_rule_preview(
            example,
            render_result.get("description_fields", {}),
            rules_config,
            template_key,
        )
        if not preview_result.get("success"):
            return {
                "success": False,
                "error": preview_result.get("error", "Excel 生成失败：规则预览失败"),
            }

        operations = preview_result.get("operations", [])
        if not isinstance(operations, list):
            operations = []

        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_template_name = _safe_output_filename_part(Path(file.filename or "template").stem)
        filename = f"workbench_{safe_template_name}_{timestamp}.xlsx"
        output_path = output_dir / filename

        executor_result = execute_rules_to_template_excel(
            str(temp_path),
            operations,
            str(output_path),
        )
        if not executor_result.get("success"):
            return {
                "success": False,
                "error": executor_result.get("error", "Excel 生成失败"),
                "warnings": preview_result.get("warnings", []) + executor_result.get("warnings", []),
                "operations_count": len(operations),
            }

        download_url = f"/api/v4/output/{filename}"
        if not download_url:
            return {
                "success": False,
                "error": "下载链接生成失败",
            }

        warnings = []
        warnings.extend(render_result.get("warnings", []))
        warnings.extend(preview_result.get("warnings", []))
        warnings.extend(executor_result.get("warnings", []))

        logger.info(
            "V4 workbench export succeeded: filename=%s operations=%s warnings=%s",
            filename,
            len(operations),
            len(warnings),
        )
        return {
            "success": True,
            "filename": filename,
            "download_url": download_url,
            "warnings": warnings,
            "operations_count": len(operations),
            "operations_written": executor_result.get("operations_written", 0),
        }
    except json.JSONDecodeError as exc:
        logger.info("V4 workbench export rejected by invalid JSON: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": f"Excel 生成失败：rules JSON 不合法",
        }
    except ValueError as exc:
        logger.info("V4 workbench export rejected: filename=%s error=%s", file.filename, exc)
        return {
            "success": False,
            "error": str(exc),
        }
    except (BadZipFile, InvalidFileException):
        logger.info("V4 workbench export rejected as non Excel: filename=%s", file.filename)
        return {
            "success": False,
            "error": "文件不是 Excel",
        }
    except Exception as exc:
        logger.exception("V4 workbench export failed: filename=%s", file.filename)
        return {
            "success": False,
            "error": f"Excel 生成失败：{exc}",
        }
    finally:
        _remove_v4_uploaded_template(temp_path)


@router.get("/api/v4/excel-render-rules")
def api_v4_excel_render_rules():
    logger.info("V4 Excel render rules requested")
    return {
        "success": True,
        "data": load_excel_render_rules(),
    }


@router.post("/api/v4/excel-render-rules")
def api_v4_save_excel_render_rules(rules_config: Any):
    logger.info("V4 Excel render rules save requested")
    validation_result = validate_excel_render_rules(rules_config)
    if validation_result.get("errors"):
        logger.info(
            "V4 Excel render rules save rejected by validation: errors=%s warnings=%s",
            len(validation_result.get("errors", [])),
            len(validation_result.get("warnings", [])),
        )
        return {
            "success": False,
            "error": "\u0045\u0078\u0063\u0065\u006c\u6e32\u67d3\u89c4\u5219\u6821\u9a8c\u5931\u8d25",
            "validation": validation_result,
        }

    result = save_excel_render_rules(rules_config)
    if not result.get("success"):
        return {
            "success": False,
            "error": result.get("error", "V4 Excel render rules save failed"),
        }

    return {
        "success": True,
        "data": result.get("data", {}),
        "validation": validation_result,
    }


@router.get("/api/v4/excel-render-rules/validate")
def api_v4_validate_excel_render_rules():
    logger.info("V4 Excel render rules validation requested")
    return {
        "success": True,
        "data": validate_excel_render_rules(load_excel_render_rules()),
    }


@router.get("/api/v4/excel-render-rules/{template_key}")
def api_v4_excel_render_template_rules(template_key: str):
    template_rules = get_template_rules(template_key)
    if not template_rules:
        logger.info("V4 Excel render rules template not found: template_key=%s", template_key)
        return {
            "success": False,
            "error": "\u0045\u0078\u0063\u0065\u006c\u6e32\u67d3\u89c4\u5219\u6a21\u677f\u4e0d\u5b58\u5728",
        }

    logger.info("V4 Excel render rules template requested: template_key=%s", template_key)
    return {
        "success": True,
        "data": template_rules,
    }


@router.get("/api/v4/product-forms")
def api_v4_product_forms():
    logger.info("V4 product forms requested")
    return {
        "success": True,
        "data": get_product_forms(),
    }


@router.get("/api/v4/product-forms/{form_key}")
def api_v4_product_form(form_key: str):
    product_forms = get_product_forms()
    if form_key not in product_forms:
        logger.info("V4 product form not found: form_key=%s", form_key)
        return {
            "success": False,
            "error": "产品形式不存在",
        }

    logger.info("V4 product form requested: form_key=%s", form_key)
    return {
        "success": True,
        "data": get_product_form(form_key),
    }


@router.get("/api/v4/examples")
def api_v4_examples():
    logger.info("V4 examples requested")
    return {
        "success": True,
        "data": list_examples(),
    }


@router.get("/api/v4/examples/{example_name}")
def api_v4_example(example_name: str):
    example = load_example(example_name)
    if not example:
        logger.info("V4 example not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    logger.info("V4 example requested: example_name=%s", example_name)
    return {
        "success": True,
        "data": example,
    }


@router.post("/api/v4/examples/{example_name}")
def api_v4_save_example(example_name: str, data: Any):
    result = save_example(example_name, data)
    if not result.get("success"):
        return {
            "success": False,
            "error": result.get("error", "V4 example save failed"),
        }

    return {
        "success": True,
    }


@router.get("/api/v4/examples/{example_name}/validate")
def api_v4_example_validate(example_name: str):
    example = load_example(example_name)
    if not example:
        logger.info("V4 example validate not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    logger.info("V4 example validate requested: example_name=%s", example_name)
    return {
        "success": True,
        "data": validate_example_order(example, load_product_schema()),
    }


@router.get("/api/v4/examples/{example_name}/render-description")
def api_v4_example_render_description(example_name: str):
    example = load_example(example_name)
    if not example:
        logger.info("V4 example render-description not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    logger.info("V4 example render-description requested: example_name=%s", example_name)
    return {
        "success": True,
        "data": render_example_to_description_fields(example, load_product_schema()),
    }


@router.get("/api/v4/examples/{example_name}/excel-rule-preview")
def api_v4_example_excel_rule_preview(example_name: str, template_key: str = ""):
    if not str(template_key or "").strip():
        return {
            "success": False,
            "error": "template_key \u4e0d\u80fd\u4e3a\u7a7a",
        }

    example = load_example(example_name)
    if not example:
        logger.info("V4 example Excel rule preview not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    try:
        logger.info(
            "V4 example Excel rule preview requested: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        preview_result = build_excel_rule_preview(
            example,
            render_result.get("description_fields", {}),
            load_excel_render_rules(),
            template_key,
        )
        if not preview_result.get("success"):
            return {
                "success": False,
                "error": preview_result.get("error", "V4 Excel rule preview failed"),
            }

        return {
            "success": True,
            "data": preview_result,
        }
    except Exception as exc:
        logger.exception(
            "V4 example Excel rule preview failed: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-rule-excel")
def api_v4_example_export_rule_excel(example_name: str, template_key: str = ""):
    if not str(template_key or "").strip():
        return {
            "success": False,
            "error": "template_key \u4e0d\u80fd\u4e3a\u7a7a",
        }

    example = load_example(example_name)
    if not example:
        logger.info("V4 example export rule Excel not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    try:
        logger.info(
            "V4 example export rule Excel requested: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        preview_result = build_excel_rule_preview(
            example,
            render_result.get("description_fields", {}),
            load_excel_render_rules(),
            template_key,
        )
        if not preview_result.get("success"):
            return {
                "success": False,
                "error": preview_result.get("error", "V4 Excel rule preview failed"),
            }

        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_example_name = _safe_output_filename_part(example_name)
        safe_template_key = _safe_output_filename_part(template_key)
        filename = f"{safe_example_name}_{safe_template_key}_rule_excel_{timestamp}.xlsx"
        output_path = output_dir / filename

        executor_result = execute_excel_rule_preview_to_workbook(
            preview_result.get("operations", []),
            str(output_path),
        )
        if not executor_result.get("success"):
            return {
                "success": False,
                "error": executor_result.get("error", "V4 Excel rule executor failed"),
            }

        warnings = []
        warnings.extend(preview_result.get("warnings", []))
        warnings.extend(executor_result.get("warnings", []))

        logger.info(
            "V4 example export rule Excel succeeded: output_path=%s operations_written=%s warnings=%s",
            output_path,
            executor_result.get("operations_written", 0),
            len(warnings),
        )
        return {
            "success": True,
            "output_path": str(output_path),
            "filename": filename,
            "operations_written": executor_result.get("operations_written", 0),
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception(
            "V4 example export rule Excel failed: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-template-rule-excel")
def api_v4_example_export_template_rule_excel(example_name: str, payload: Any = Body(None), template_key: str = ""):
    if not str(template_key or "").strip():
        return {
            "success": False,
            "error": "template_key \u4e0d\u80fd\u4e3a\u7a7a",
        }
    if not isinstance(payload, dict):
        return {
            "success": False,
            "error": "template_path \u4e0d\u80fd\u4e3a\u7a7a",
        }

    template_path, template_path_error = _resolve_template_path(payload.get("template_path"))
    if template_path_error:
        return {
            "success": False,
            "error": template_path_error,
        }

    example = load_example(example_name)
    if not example:
        logger.info("V4 example export template rule Excel not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    try:
        logger.info(
            "V4 example export template rule Excel requested: example_name=%s template_key=%s template_path=%s",
            example_name,
            template_key,
            template_path,
        )
        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        preview_result = build_excel_rule_preview(
            example,
            render_result.get("description_fields", {}),
            load_excel_render_rules(),
            template_key,
        )
        if not preview_result.get("success"):
            return {
                "success": False,
                "error": preview_result.get("error", "V4 Excel rule preview failed"),
            }
        operations = preview_result.get("operations", [])
        if not isinstance(operations, list):
            operations = []

        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_example_name = _safe_output_filename_part(example_name)
        safe_template_key = _safe_output_filename_part(template_key)
        filename = f"{safe_example_name}_{safe_template_key}_template_rule_{timestamp}.xlsx"
        output_path = output_dir / filename

        executor_result = execute_rules_to_template_excel(
            str(template_path),
            operations,
            str(output_path),
        )
        if not executor_result.get("success"):
            return {
                "success": False,
                "error": executor_result.get("error", "V4 template rule executor failed"),
            }

        warnings = []
        warnings.extend(preview_result.get("warnings", []))
        warnings.extend(executor_result.get("warnings", []))

        logger.info(
            "V4 example export template rule Excel succeeded: output_path=%s operations_written=%s warnings=%s",
            output_path,
            executor_result.get("operations_written", 0),
            len(warnings),
        )
        return {
            "success": True,
            "filename": filename,
            "output_path": str(output_path),
            "operations_written": executor_result.get("operations_written", 0),
            "operations": operations,
            "warnings": warnings,
        }
    except Exception as exc:
        logger.exception(
            "V4 example export template rule Excel failed: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-template-rule-excel-with-preview")
def api_v4_example_export_template_rule_excel_with_preview(example_name: str, payload: Any = Body(None), template_key: str = ""):
    if not str(template_key or "").strip():
        return {
            "success": False,
            "error": "template_key \u4e0d\u80fd\u4e3a\u7a7a",
        }
    if not isinstance(payload, dict):
        return {
            "success": False,
            "error": "template_path \u4e0d\u80fd\u4e3a\u7a7a",
        }

    template_path, template_path_error = _resolve_template_path(payload.get("template_path"))
    if template_path_error:
        return {
            "success": False,
            "error": template_path_error,
        }

    try:
        logger.info(
            "V4 example export template rule Excel with preview requested: example_name=%s template_key=%s template_path=%s",
            example_name,
            template_key,
            template_path,
        )
        result = execute_rules_to_template_excel_with_preview(
            example_name,
            template_key,
            str(template_path),
        )
        if not result.get("success"):
            logger.info(
                "V4 example export template rule Excel with preview failed: example_name=%s template_key=%s error=%s",
                example_name,
                template_key,
                result.get("error", ""),
            )
            return result

        logger.info(
            "V4 example export template rule Excel with preview succeeded: filename=%s operations_written=%s warnings=%s",
            result.get("filename", ""),
            result.get("operations_written", 0),
            len(result.get("warnings", [])),
        )
        return result
    except Exception as exc:
        logger.exception(
            "V4 example export template rule Excel with preview failed: example_name=%s template_key=%s",
            example_name,
            template_key,
        )
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-ai-template-excel")
def api_v4_example_export_ai_template_excel(example_name: str, payload: Any = Body(None)):
    if not isinstance(payload, dict):
        return {
            "success": False,
            "error": "template_path \u4e0d\u80fd\u4e3a\u7a7a",
        }

    template_path, template_path_error = _resolve_template_path(payload.get("template_path"))
    if template_path_error:
        return {
            "success": False,
            "error": template_path_error,
        }

    try:
        logger.info(
            "V4 example export AI template Excel requested: example_name=%s template_path=%s",
            example_name,
            template_path,
        )
        result = execute_ai_template_to_excel(example_name, str(template_path))
        if not result.get("success"):
            logger.info(
                "V4 example export AI template Excel failed: example_name=%s error=%s",
                example_name,
                result.get("error", ""),
            )
            return result

        logger.info(
            "V4 example export AI template Excel succeeded: filename=%s operations_written=%s warnings=%s",
            result.get("filename", ""),
            result.get("operations_written", 0),
            len(result.get("warnings", [])),
        )
        return result
    except Exception as exc:
        logger.exception("V4 example export AI template Excel failed: example_name=%s", example_name)
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-ai-template-excel-cn")
def api_v4_example_export_ai_template_excel_cn(example_name: str, payload: Any = Body(None)):
    if not isinstance(payload, dict):
        return {
            "成功": False,
            "错误": "模板路径不能为空",
        }

    template_path, template_path_error = _resolve_template_path(payload.get("template_path"))
    if template_path_error:
        return {
            "成功": False,
            "错误": template_path_error,
        }

    try:
        logger.info(
            "V4 example export AI template Excel CN requested: example_name=%s template_path=%s",
            example_name,
            template_path,
        )
        result = 执行模板规则并生成Excel(example_name, str(template_path))
        if not result.get("成功"):
            logger.info(
                "V4 example export AI template Excel CN failed: example_name=%s error=%s",
                example_name,
                result.get("错误", ""),
            )
        return result
    except Exception as exc:
        logger.exception("V4 example export AI template Excel CN failed: example_name=%s", example_name)
        return {
            "成功": False,
            "错误": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-batch-ai-excel")
def api_v4_example_export_batch_ai_excel(example_name: str, payload: Any = Body(None)):
    if not isinstance(payload, dict):
        return {
            "成功": False,
            "错误": "模板路径列表不能为空",
        }

    template_paths = payload.get("template_paths")
    if not isinstance(template_paths, list) or not template_paths:
        return {
            "成功": False,
            "错误": "模板路径列表不能为空",
        }

    resolved_template_paths = []
    for template_path in template_paths:
        resolved_path, template_path_error = _resolve_template_path(template_path)
        if template_path_error:
            return {
                "成功": False,
                "错误": f"模板缺失：{template_path or '未命名模板'}；{template_path_error}",
            }
        resolved_template_paths.append(str(resolved_path))

    try:
        logger.info(
            "V4 example export batch AI Excel requested: example_name=%s templates=%s",
            example_name,
            len(resolved_template_paths),
        )
        result = execute_batch_template_to_excel(example_name, resolved_template_paths)
        if not result.get("成功"):
            logger.info(
                "V4 example export batch AI Excel finished with issues: example_name=%s error=%s",
                example_name,
                result.get("错误", ""),
            )
        return result
    except Exception as exc:
        logger.exception("V4 example export batch AI Excel failed: example_name=%s", example_name)
        return {
            "成功": False,
            "错误": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-render-json")
def api_v4_example_export_render_json(example_name: str):
    example = load_example(example_name)
    if not example:
        logger.info("V4 example export render JSON not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    try:
        logger.info("V4 example export render JSON requested: example_name=%s", example_name)
        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = str(example_name or "").strip()
        if safe_name.endswith(".json"):
            safe_name = safe_name[:-5]
        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{safe_name}_render_{timestamp}.json"
        output_path = output_dir / filename

        payload = {
            "example_name": example_name,
            "generated_at": generated_at,
            "description_fields": render_result.get("description_fields", {}),
            "warnings": render_result.get("warnings", []),
        }
        with output_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
            f.write("\n")

        logger.info("V4 example export render JSON succeeded: path=%s", output_path)
        return {
            "success": True,
            "output_path": str(output_path),
            "filename": filename,
        }
    except Exception as exc:
        logger.exception("V4 example export render JSON failed: example_name=%s", example_name)
        return {
            "success": False,
            "error": str(exc),
        }


@router.post("/api/v4/examples/{example_name}/export-debug-excel")
def api_v4_example_export_debug_excel(example_name: str):
    example = load_example(example_name)
    if not example:
        logger.info("V4 example export debug Excel not found: example_name=%s", example_name)
        return {
            "success": False,
            "error": "\u793a\u4f8b\u8ba2\u5355\u4e0d\u5b58\u5728",
        }

    try:
        logger.info("V4 example export debug Excel requested: example_name=%s", example_name)
        render_result = render_example_to_description_fields(example, load_product_schema())
        if not render_result.get("success"):
            return {
                "success": False,
                "error": render_result.get("error", "V4 renderer failed"),
            }

        export_result = export_description_fields_to_debug_excel(
            example_name,
            render_result.get("description_fields", {}),
        )
        if not export_result.get("success"):
            return {
                "success": False,
                "error": export_result.get("error", "V4 debug Excel export failed"),
            }

        logger.info(
            "V4 example export debug Excel succeeded: example_name=%s, filename=%s",
            example_name,
            export_result.get("filename"),
        )
        return {
            "success": True,
            "output_path": export_result.get("output_path", ""),
            "filename": export_result.get("filename", ""),
        }
    except Exception as exc:
        logger.exception("V4 example export debug Excel failed: example_name=%s", example_name)
        return {
            "success": False,
            "error": str(exc),
        }

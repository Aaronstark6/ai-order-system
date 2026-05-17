from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from app.logger import get_logger


logger = get_logger(__name__)


def _normalize_direction(value):
    return "below" if str(value or "").strip() == "below" else "right"


def _normalize_mappings(mappings):
    if not isinstance(mappings, list):
        return []

    normalized = []
    seen = set()
    for item in mappings:
        if isinstance(item, dict):
            label = str(item.get("label") or "").strip()
            direction = _normalize_direction(item.get("target_direction"))
        else:
            label = str(item or "").strip()
            direction = "right"

        if not label:
            continue

        key = (label, direction)
        if key not in seen:
            normalized.append({
                "label": label,
                "target_direction": direction,
            })
            seen.add(key)
    return normalized


def _find_merged_range(ws, cell_ref):
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range
    return None


def _resolve_merged_cell(ws, cell_ref):
    merged_range = _find_merged_range(ws, cell_ref)
    return merged_range.start_cell.coordinate if merged_range else cell_ref


def _target_cell_ref(ws, cell, direction):
    label_range = _find_merged_range(ws, cell.coordinate)
    if label_range and direction == "below":
        target_ref = f"{get_column_letter(label_range.min_col)}{label_range.max_row + 1}"
    elif label_range:
        target_ref = f"{get_column_letter(label_range.max_col + 1)}{label_range.min_row}"
    elif direction == "below":
        target_ref = f"{get_column_letter(cell.column)}{cell.row + 1}"
    else:
        target_ref = f"{get_column_letter(cell.column + 1)}{cell.row}"

    return _resolve_merged_cell(ws, target_ref)


def scan_labels_in_excel(template_path, mappings):
    warnings = []
    mapping_list = _normalize_mappings(mappings)
    if not mapping_list:
        return {
            "success": False,
            "matches": [],
            "unmatched_labels": [],
            "warnings": ["没有可扫描的字段标签"],
        }

    label_list = [item["label"] for item in mapping_list]
    template = Path(template_path) if template_path else None
    if not template or not template.is_file():
        return {
            "success": False,
            "matches": [],
            "unmatched_labels": label_list,
            "warnings": ["请先上传 Excel 模板"],
        }

    try:
        wb = load_workbook(template, data_only=True)
    except (InvalidFileException, OSError, ValueError) as exc:
        logger.exception("[LabelLocator] Template scan failed: path=%s", template)
        return {
            "success": False,
            "matches": [],
            "unmatched_labels": label_list,
            "warnings": [str(exc) or "模板扫描失败"],
        }

    direction_by_label = {}
    for item in mapping_list:
        direction_by_label.setdefault(item["label"], item["target_direction"])

    label_set = set(direction_by_label.keys())
    matches_by_label = {}
    duplicate_count = {}

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue

                value = str(cell.value).strip()
                if value not in label_set:
                    continue

                target_direction = direction_by_label.get(value, "right")
                target_cell = _target_cell_ref(ws, cell, target_direction)
                match = {
                    "label": value,
                    "target_direction": target_direction,
                    "sheet": ws.title,
                    "label_cell": cell.coordinate,
                    "target_cell": target_cell,
                    "value": value,
                }
                if value not in matches_by_label:
                    matches_by_label[value] = match
                else:
                    duplicate_count[value] = duplicate_count.get(value, 1) + 1

    for label, count in duplicate_count.items():
        if count > 1:
            warnings.append(f"标签“{label}”命中多个位置，已使用第一个")

    ordered_labels = []
    seen_labels = set()
    for label in label_list:
        if label not in seen_labels:
            ordered_labels.append(label)
            seen_labels.add(label)

    matches = [matches_by_label[label] for label in ordered_labels if label in matches_by_label]
    unmatched_labels = [label for label in ordered_labels if label not in matches_by_label]

    logger.info(
        "[LabelLocator] Template labels scanned: labels=%s matches=%s unmatched=%s warnings=%s",
        len(label_list),
        len(matches),
        len(unmatched_labels),
        len(warnings),
    )
    return {
        "success": True,
        "matches": matches,
        "unmatched_labels": unmatched_labels,
        "warnings": warnings,
    }

import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.v4_template_intelligence import infer_structured_mapping_from_labels
from app.v4_mapping_generator import generate_auto_mapping
from app.v4_template_scanner import scan_excel_labels


HEADER_FIELDS = {
    "原料名": "name",
    "原料": "name",
    "项目名称": "name",
    "项目": "name",
    "含量": "amount",
    "结果": "result",
    "百分比": "percentage",
    "规格": "spec",
    "标准": "standard",
    "营养成分": "nutrient",
    "每份含量": "amount",
    "NRV%": "nrv",
    "产品名称": "product_name",
    "产品": "product_name",
    "品名": "product_name",
    "数量": "quantity",
    "订购数量": "quantity",
    "采购数量": "quantity",
    "包装": "packaging",
    "包装方式": "packaging",
    "单价": "unit_price",
    "价格": "unit_price",
    "交期": "delivery_time",
    "交货期": "delivery_time",
}

BLOCK_KEYWORDS = [
    "产品说明",
    "产品特点",
    "包装要求",
    "其他要求",
    "备注",
    "说明",
]

SCORING_KEYWORDS = [
    "产品描述",
    "产品详细要求",
    "包装要求",
    "配方要求",
    "容器要求",
    "装量要求",
    "容器内填充物",
    "瓶口密封",
    "盖子密封",
    "标签要求",
    "客户设计制作",
    "我司设计制作",
    "生产日期",
    "批号",
    "客户名称",
    "客户性质",
    "负责人",
    "文档编号",
    "日期",
    "原料名",
    "含量",
    "百分比",
    "项目名称",
    "标准",
    "结果",
]

SEMANTIC_RULES = [
    (
        "product_detail",
        [
            "产品详细",
            "产品描述",
            "软胶囊",
            "片剂",
            "固体饮料",
            "软糖",
        ],
    ),
    (
        "packaging",
        [
            "包装要求",
            "容器要求",
            "装量要求",
            "容器内填充物",
            "瓶口密封",
            "盖子密封",
        ],
    ),
    (
        "formula",
        [
            "配方要求",
            "原料名",
            "含量",
            "百分比",
        ],
    ),
    (
        "label_requirement",
        [
            "标签要求",
            "客户设计制作",
            "我司设计制作",
            "不贴标签",
        ],
    ),
    (
        "production_batch",
        [
            "生产日期",
            "批号",
        ],
    ),
    (
        "customer_info",
        [
            "客户名称",
            "客户性质",
            "数量",
            "负责人",
        ],
    ),
    (
        "document_info",
        [
            "文档编号",
            "日期",
        ],
    ),
]

SEMANTIC_NAMES = {
    "product_detail": "产品详细要求区域",
    "packaging": "包装要求区域",
    "formula": "配方要求区域",
    "label_requirement": "标签要求区域",
    "production_batch": "生产日期批号区域",
    "customer_info": "客户信息区域",
    "document_info": "文档信息区域",
    "unknown": "智能识别区域",
}


def _clean_text(value):
    text = str(value or "").strip()
    return text.rstrip(":：") if text else ""


def _cell_value(cell):
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def _is_empty_cell(cell):
    return _cell_value(cell) == ""


def _load_workbook(template_path):
    return load_workbook(template_path, data_only=True)


def _row_text_cells(sheet, row_number):
    cells = []
    for cell in sheet[row_number]:
        value = _clean_text(cell.value)
        if value:
            cells.append(
                {
                    "sheet": sheet.title,
                    "row": cell.row,
                    "col": cell.column,
                    "cell": cell.coordinate,
                    "value": value,
                }
            )
    return cells


def _contiguous_runs(cells):
    if not cells:
        return []

    runs = []
    current = [cells[0]]
    for cell in cells[1:]:
        if cell["col"] == current[-1]["col"] + 1:
            current.append(cell)
        else:
            runs.append(current)
            current = [cell]
    runs.append(current)
    return runs


def _has_data_below(sheet, header_row, start_col, end_col):
    for row_number in range(header_row + 1, min(sheet.max_row, header_row + 8) + 1):
        values = [
            _cell_value(sheet.cell(row=row_number, column=col))
            for col in range(start_col, end_col + 1)
        ]
        if any(values):
            return True
    return False


def _infer_table_name(columns):
    labels = {column.get("label", "") for column in columns}
    if {"原料名", "含量"} & labels and ("百分比" in labels or "规格" in labels):
        return "配方表"
    if "项目名称" in labels and ("标准" in labels or "结果" in labels):
        return "检测项目表"
    if "营养成分" in labels or "NRV%" in labels:
        return "营养成分表"
    if "原料" in labels or "原料名" in labels:
        return "原料表"
    return "智能识别表格"


def _column_for_header(cell_info):
    label = cell_info.get("value", "")
    return {
        "label": label,
        "field": HEADER_FIELDS.get(label, label),
        "target_col": get_column_letter(cell_info.get("col", 1)),
        "header_cell": cell_info.get("cell", ""),
    }


def infer_region_bounds(items, default_end_row=None):
    normalized_items = [item for item in items if isinstance(item, dict)]
    if not normalized_items:
        return {
            "start_row": None,
            "end_row": None,
            "start_col": None,
            "end_col": None,
        }

    rows = [int(item["row"]) for item in normalized_items if item.get("row") is not None]
    cols = [int(item["col"]) for item in normalized_items if item.get("col") is not None]
    if not rows or not cols:
        return {
            "start_row": None,
            "end_row": None,
            "start_col": None,
            "end_col": None,
        }

    return {
        "start_row": min(rows),
        "end_row": default_end_row if default_end_row is not None else max(rows),
        "start_col": min(cols),
        "end_col": max(cols),
    }


def _bounds_numbers(region):
    bounds = region.get("bounds", {}) if isinstance(region, dict) else {}
    try:
        start_row = int(bounds.get("start_row"))
        end_row = int(bounds.get("end_row"))
        start_col = int(bounds.get("start_col"))
        end_col = int(bounds.get("end_col"))
    except (TypeError, ValueError):
        return None

    if end_row < start_row or end_col < start_col:
        return None

    return start_row, end_row, start_col, end_col


def _region_area(region):
    numbers = _bounds_numbers(region)
    if not numbers:
        return 0
    start_row, end_row, start_col, end_col = numbers
    return (end_row - start_row + 1) * (end_col - start_col + 1)


def _region_rows_count(region):
    numbers = _bounds_numbers(region)
    if not numbers:
        return 0
    start_row, end_row, _, _ = numbers
    return end_row - start_row + 1


def _region_cols_count(region):
    numbers = _bounds_numbers(region)
    if not numbers:
        return 0
    _, _, start_col, end_col = numbers
    return end_col - start_col + 1


def _region_text(region):
    values = [
        str(region.get("name", "")),
        str(region.get("type", "")),
    ]
    labels = region.get("label_names")
    if isinstance(labels, list):
        values.extend(str(item) for item in labels)
    return " ".join(values)


def _candidate_type(score):
    if score >= 80:
        return "strong"
    if score >= 50:
        return "candidate"
    return "weak"


def _confidence(score):
    if score >= 80:
        return "high"
    if score >= 50:
        return "medium"
    return "low"


def classify_region_semantics(region):
    text = _region_text(region)
    for semantic_type, keywords in SEMANTIC_RULES:
        matched_keywords = [keyword for keyword in keywords if keyword in text]
        if matched_keywords:
            return {
                "semantic_type": semantic_type,
                "matched_keywords": matched_keywords,
            }
    return {
        "semantic_type": "unknown",
        "matched_keywords": [],
    }


def suggest_region_name(region):
    semantic_type = str(region.get("semantic_type") or "unknown")
    return SEMANTIC_NAMES.get(semantic_type, SEMANTIC_NAMES["unknown"])


def score_region(region):
    scored = dict(region) if isinstance(region, dict) else {}
    bounds = dict(scored.get("bounds", {})) if isinstance(scored.get("bounds", {}), dict) else {}
    scored["bounds"] = bounds

    score = 35
    reasons = []
    area = _region_area(scored)
    rows_count = _region_rows_count(scored)
    cols_count = _region_cols_count(scored)
    text = _region_text(scored)
    semantic = classify_region_semantics(scored)
    scored["semantic_type"] = semantic.get("semantic_type", "unknown")
    scored["matched_keywords"] = semantic.get("matched_keywords", [])
    scored["suggested_name"] = suggest_region_name(scored)

    if area <= 1:
        score -= 25
        reasons.append("区域过小")
    elif 4 <= area <= 60:
        score += 18
        reasons.append("区域面积合理")
    elif area > 160:
        score -= 30
        reasons.append("区域过大")
    elif area > 90:
        score -= 12
        reasons.append("区域偏大")

    matched_keywords = [keyword for keyword in SCORING_KEYWORDS if keyword in text]
    if matched_keywords:
        score += min(35, len(matched_keywords) * 8)
        reasons.append("命中高质量标题")

    if scored.get("semantic_type") != "unknown":
        score += 10
        reasons.append("语义分类明确")

    columns_count = int(scored.get("columns_count") or cols_count or 0)
    if columns_count >= 2:
        score += 12
        reasons.append("列结构有效")
    if rows_count >= 2:
        score += 10
        reasons.append("包含数据行")

    label_names = scored.get("label_names")
    if isinstance(label_names, list) and len(label_names) > 8:
        score -= 18
        reasons.append("覆盖标签过多")
    if scored.get("name") in {"智能识别表格", ""}:
        score -= 5
        reasons.append("名称较泛")

    score = max(0, min(100, score))
    confidence = _confidence(score)
    if scored.get("suggested_name") != SEMANTIC_NAMES["unknown"] and confidence == "low":
        confidence = "medium"
    scored["score"] = score
    scored["confidence"] = confidence
    scored["candidate_type"] = _candidate_type(score)
    scored["reason"] = "；".join(reasons) if reasons else "基础结构候选"
    return scored


def _intersection_area(first, second):
    first_bounds = _bounds_numbers(first)
    second_bounds = _bounds_numbers(second)
    if not first_bounds or not second_bounds:
        return 0

    first_start_row, first_end_row, first_start_col, first_end_col = first_bounds
    second_start_row, second_end_row, second_start_col, second_end_col = second_bounds
    row_overlap = min(first_end_row, second_end_row) - max(first_start_row, second_start_row) + 1
    col_overlap = min(first_end_col, second_end_col) - max(first_start_col, second_start_col) + 1
    if row_overlap <= 0 or col_overlap <= 0:
        return 0
    return row_overlap * col_overlap


def _overlap_ratio(first, second):
    intersection = _intersection_area(first, second)
    if intersection <= 0:
        return 0
    smaller_area = min(_region_area(first), _region_area(second))
    if smaller_area <= 0:
        return 0
    return intersection / smaller_area


def _region_key(region):
    bounds = region.get("bounds", {}) if isinstance(region, dict) else {}
    return (
        region.get("type", ""),
        region.get("name", ""),
        region.get("sheet", ""),
        bounds.get("start_row"),
        bounds.get("end_row"),
        bounds.get("start_col"),
        bounds.get("end_col"),
    )


def deduplicate_regions(regions):
    unique = {}
    for region in regions if isinstance(regions, list) else []:
        key = _region_key(region)
        current = unique.get(key)
        if current is None or region.get("score", 0) > current.get("score", 0):
            unique[key] = region

    sorted_regions = sorted(
        unique.values(),
        key=lambda item: (
            item.get("score", 0),
            -_region_area(item),
            item.get("type", ""),
        ),
        reverse=True,
    )

    kept = []
    for region in sorted_regions:
        should_keep = True
        for kept_region in kept:
            if region.get("sheet") != kept_region.get("sheet"):
                continue
            if region.get("type") != kept_region.get("type"):
                continue
            if _overlap_ratio(region, kept_region) > 0.7:
                should_keep = False
                break
        if should_keep:
            kept.append(region)

    return sorted(
        kept,
        key=lambda item: (
            item.get("bounds", {}).get("start_row") or 0,
            item.get("bounds", {}).get("start_col") or 0,
            -item.get("score", 0),
        ),
    )


def _infer_table_end_row(sheet, header_row, start_col, end_col):
    end_row = header_row
    blank_streak = 0
    for row_number in range(header_row + 1, sheet.max_row + 1):
        row_has_value = any(
            _cell_value(sheet.cell(row=row_number, column=col))
            for col in range(start_col, end_col + 1)
        )
        if row_has_value:
            end_row = row_number
            blank_streak = 0
            continue
        blank_streak += 1
        if blank_streak >= 2:
            break
    return end_row


def _infer_table_semantic_type(header_values):
    values = {str(value or "").strip() for value in header_values if str(value or "").strip()}
    order_item_keywords = {
        "产品名称",
        "产品",
        "品名",
        "数量",
        "订购数量",
        "采购数量",
        "规格",
        "产品规格",
        "包装",
        "包装方式",
        "单价",
        "价格",
        "交期",
        "交货期",
    }
    formula_keywords = {
        "原料名",
        "原料",
        "含量",
        "百分比",
        "规格",
    }
    if len(values & order_item_keywords) >= 2:
        return "order_items"
    if len(values & formula_keywords) >= 2:
        return "formula"
    return "table"


def infer_header_row(sheet, row_number):
    row_cells = _row_text_cells(sheet, row_number)
    candidates = []
    for run in _contiguous_runs(row_cells):
        if len(run) < 2:
            continue
        keyword_count = sum(1 for cell in run if cell.get("value") in HEADER_FIELDS)
        start_col = run[0]["col"]
        end_col = run[-1]["col"]
        has_data_below = _has_data_below(sheet, row_number, start_col, end_col)
        if keyword_count == 0 and (len(run) < 3 or not has_data_below):
            continue

        columns = [_column_for_header(cell) for cell in run]
        candidates.append(
            {
                "header_row": row_number,
                "columns": columns,
                "start_col": start_col,
                "end_col": end_col,
                "score": keyword_count * 2 + len(run) + (1 if has_data_below else 0),
                "has_data_below": has_data_below,
            }
        )

    if not candidates:
        return None
    return sorted(candidates, key=lambda item: item["score"], reverse=True)[0]


def scan_table_regions(template_path):
    workbook = _load_workbook(template_path)
    try:
        table_regions = []
        for sheet in workbook.worksheets:
            for row_number in range(1, sheet.max_row + 1):
                header = infer_header_row(sheet, row_number)
                if not header:
                    continue

                end_row = _infer_table_end_row(
                    sheet,
                    header["header_row"],
                    header["start_col"],
                    header["end_col"],
                )
                bounds = infer_region_bounds(
                    [
                        {"row": header["header_row"], "col": header["start_col"]},
                        {"row": end_row, "col": header["end_col"]},
                    ],
                    default_end_row=end_row,
                )
                columns = header.get("columns", [])
                header_values = []
                for column in columns:
                    if isinstance(column, dict):
                        header_values.append(
                            column.get("label")
                            or column.get("header")
                            or column.get("value")
                            or ""
                        )
                semantic_type = _infer_table_semantic_type(header_values)
                table_regions.append(
                    {
                        "type": "table",
                        "semantic_type": semantic_type,
                        "sheet": sheet.title,
                        "table_key": f"smart_table_{len(table_regions) + 1}",
                        "table_name": _infer_table_name(columns),
                        "header_row": header["header_row"],
                        "start_cell": columns[0]["header_cell"] if columns else "",
                        "columns": columns,
                        "bounds": bounds,
                        "detection": {
                            "method": "smart_header_row",
                            "score": header.get("score", 0),
                            "has_data_below": header.get("has_data_below", False),
                        },
                    }
                )
        return table_regions
    finally:
        workbook.close()


def _infer_block_end_row(sheet, start_row, start_col):
    end_row = start_row
    blank_streak = 0
    for row_number in range(start_row + 1, min(sheet.max_row, start_row + 12) + 1):
        row_texts = [
            _cell_value(sheet.cell(row=row_number, column=col))
            for col in range(start_col, min(sheet.max_column, start_col + 3) + 1)
        ]
        if any(row_texts):
            end_row = row_number
            blank_streak = 0
            continue
        blank_streak += 1
        if blank_streak >= 2:
            break
    return end_row


def scan_block_regions(template_path):
    workbook = _load_workbook(template_path)
    try:
        block_regions = []
        for sheet in workbook.worksheets:
            block_labels = []
            for row_number in range(1, sheet.max_row + 1):
                for cell in sheet[row_number]:
                    value = _clean_text(cell.value)
                    if value in BLOCK_KEYWORDS:
                        block_labels.append(
                            {
                                "sheet": sheet.title,
                                "row": cell.row,
                                "col": cell.column,
                                "cell": cell.coordinate,
                                "value": value,
                            }
                        )

            clusters = []
            current = []
            for label in block_labels:
                if not current:
                    current = [label]
                    continue
                same_area = label["col"] == current[-1]["col"] and label["row"] <= current[-1]["row"] + 4
                if same_area:
                    current.append(label)
                else:
                    clusters.append(current)
                    current = [label]
            if current:
                clusters.append(current)

            for cluster in clusters:
                start = cluster[0]
                end_row = max(_infer_block_end_row(sheet, item["row"], item["col"]) for item in cluster)
                bounds = infer_region_bounds(
                    [
                        {"row": start["row"], "col": start["col"]},
                        {"row": end_row, "col": min(sheet.max_column, start["col"] + 3)},
                    ],
                    default_end_row=end_row,
                )
                block_regions.append(
                    {
                        "sheet": sheet.title,
                        "block_name": " / ".join(item["value"] for item in cluster),
                        "source_cell": start["cell"],
                        "target_cell": start["cell"],
                        "labels": [item["value"] for item in cluster],
                        "bounds": bounds,
                        "detection": {
                            "method": "smart_text_cluster",
                            "label_count": len(cluster),
                        },
                    }
                )
        return block_regions
    finally:
        workbook.close()


def build_template_structure(labels, table_regions, block_regions):
    label_regions = []
    for label in labels:
        cell = str(label.get("cell", ""))
        row_digits = "".join(ch for ch in cell if ch.isdigit())
        col_letters = "".join(ch for ch in cell if ch.isalpha())
        col_number = column_index_from_string(col_letters) if col_letters else None
        label_regions.append(
            {
                "type": "label",
                "name": label.get("value", ""),
                "sheet": label.get("sheet", ""),
                "cell": cell,
                "bounds": {
                    "start_row": int(row_digits) if row_digits else None,
                    "end_row": int(row_digits) if row_digits else None,
                    "start_col": col_number,
                    "end_col": col_number,
                },
            }
        )

    table_nodes = [
        {
            "type": "table",
            "name": item.get("table_name", ""),
            "sheet": item.get("sheet", ""),
            "bounds": item.get("bounds", {}),
            "columns_count": len(item.get("columns", [])) if isinstance(item.get("columns"), list) else 0,
            "rows_count": _region_rows_count(item),
            "label_names": [
                column.get("label", "")
                for column in item.get("columns", [])
                if isinstance(column, dict) and column.get("label")
            ]
        }
        for item in table_regions
    ]
    block_nodes = [
        {
            "type": "block",
            "name": item.get("block_name", ""),
            "sheet": item.get("sheet", ""),
            "bounds": item.get("bounds", {}),
            "labels_count": len(item.get("labels", [])) if isinstance(item.get("labels"), list) else 0,
            "rows_count": _region_rows_count(item),
            "label_names": item.get("labels", []) if isinstance(item.get("labels", []), list) else [],
        }
        for item in block_regions
    ]
    raw_regions = [score_region(region) for region in table_nodes + block_nodes]
    deduped_regions = deduplicate_regions(raw_regions)
    recommended_regions = [
        region
        for region in sorted(deduped_regions, key=lambda item: item.get("score", 0), reverse=True)
        if region.get("candidate_type") in {"strong", "candidate"}
    ][:10]
    recommended_regions = sorted(
        recommended_regions,
        key=lambda item: (
            item.get("bounds", {}).get("start_row") or 0,
            item.get("bounds", {}).get("start_col") or 0,
            -item.get("score", 0),
        ),
    )

    return {
        "regions": recommended_regions,
        "raw_regions": raw_regions,
        "deduped_regions": deduped_regions,
        "recommended_regions": recommended_regions,
        "tables": [region for region in recommended_regions if region.get("type") == "table"],
        "blocks": [region for region in recommended_regions if region.get("type") == "block"],
        "labels": label_regions,
    }


def _cell_ref(row, col):
    return f"{get_column_letter(col)}{row}"


def _cell_point(cell_ref):
    match = re.match(r"^([A-Z]+)([0-9]+)$", str(cell_ref or "").strip().upper())
    if not match:
        return None
    return int(match.group(2)), column_index_from_string(match.group(1))


def _build_semantic_context(template_path):
    workbook = load_workbook(template_path, data_only=True)
    try:
        context = {"sheets": {}, "cells": {}, "rows": {}, "max_row": 0, "max_col": 0}
        for sheet in workbook.worksheets:
            sheet_cells = {}
            sheet_rows = {}
            merged_lookup = {}
            for merged_range in sheet.merged_cells.ranges:
                start_cell = merged_range.start_cell.coordinate
                range_text = str(merged_range)
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        merged_lookup[_cell_ref(row, col)] = {
                            "start_cell": start_cell,
                            "range": range_text,
                            "is_start": _cell_ref(row, col) == start_cell,
                        }

            max_row = min(int(sheet.max_row or 1), 200)
            max_col = min(int(sheet.max_column or 1), 80)
            context["max_row"] = max(context["max_row"], max_row)
            context["max_col"] = max(context["max_col"], max_col)
            for row in range(1, max_row + 1):
                row_values = []
                row_cells = []
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    coord = cell.coordinate
                    value = _cell_value(cell)
                    if value:
                        row_values.append(value)
                        row_cells.append(coord)
                    fill = getattr(cell.fill, "fgColor", None)
                    fill_text = str(getattr(fill, "rgb", "") or getattr(fill, "indexed", "") or "")
                    if fill_text in {"00000000", "0", "None"}:
                        fill_text = ""
                    merge = merged_lookup.get(coord, {})
                    info = {
                        "sheet": sheet.title,
                        "cell": coord,
                        "row": row,
                        "col": col,
                        "value": value,
                        "bold": bool(cell.font.bold),
                        "font_size": float(cell.font.sz or 0),
                        "align": str(cell.alignment.horizontal or "").lower(),
                        "fill_color": fill_text,
                        "merged_range": merge.get("range", ""),
                        "merged_start": merge.get("start_cell", ""),
                        "is_merged": bool(merge),
                    }
                    sheet_cells[coord] = info
                    context["cells"][coord] = info
                sheet_rows[row] = {"values": row_values, "cells": row_cells}
            context["sheets"][sheet.title] = {"cells": sheet_cells, "rows": sheet_rows, "max_row": max_row, "max_col": max_col}
            if not context["rows"]:
                context["rows"] = sheet_rows
        return context
    finally:
        workbook.close()


def _semantic_cell_info(context, cell_ref):
    cells = context.get("cells") if isinstance(context, dict) else {}
    return cells.get(str(cell_ref or "").strip().upper(), {}) if isinstance(cells, dict) else {}


def _semantic_cell_text(context, row, col):
    return str(_semantic_cell_info(context, _cell_ref(row, col)).get("value") or "").strip()


def _semantic_blank_cell(context, row, col):
    if row < 1 or col < 1:
        return False
    max_row = int(context.get("max_row") or row)
    max_col = int(context.get("max_col") or col)
    if row > max_row or col > max_col:
        return False
    return not _semantic_cell_text(context, row, col)


def _semantic_target_cell(context, row, col):
    ref = _cell_ref(row, col)
    info = _semantic_cell_info(context, ref)
    return str(info.get("merged_start") or ref).strip().upper()


def _semantic_coordinates(source_cell="", target_cell="", cells=None, row=None, col=None, page=None, bbox=None):
    normalized_cells = []
    if isinstance(cells, list):
        normalized_cells = [
            str(cell).strip().upper()
            for cell in cells
            if str(cell).strip()
        ]

    return {
        "source_cell": str(source_cell or "").strip().upper(),
        "target_cell": str(target_cell or "").strip().upper(),
        "cells": normalized_cells,
        "row": row,
        "col": col,
        "page": page,
        "bbox": bbox if isinstance(bbox, dict) else {},
    }


def _semantic_right_target(context, row, col):
    for offset in range(1, 5):
        target_col = col + offset
        if _semantic_blank_cell(context, row, target_col):
            return _semantic_target_cell(context, row, target_col)
    return ""


def _semantic_below_target(context, row, col):
    for offset in range(1, 4):
        target_row = row + offset
        if _semantic_blank_cell(context, target_row, col):
            return _semantic_target_cell(context, target_row, col)
    return ""


def _semantic_row_short_texts(context, row):
    rows = context.get("rows") if isinstance(context, dict) else {}
    values = rows.get(row, {}).get("values", []) if isinstance(rows, dict) else []
    return [str(value).strip() for value in values if 0 < len(str(value).strip()) <= 12]


def _semantic_region(region_id, region_type, label, source_cell, target_cell, cells, row, col, confidence, reason, write_mode, intent_type):
    return {
        "region_id": region_id,
        "type": region_type,
        "label": label,
        "coordinates": _semantic_coordinates(
            source_cell=source_cell,
            target_cell=target_cell,
            cells=cells,
            row=row,
            col=col,
        ),
        "confidence": round(float(confidence), 2),
        "reason": reason,
        "write_mode": write_mode,
        "intent_type": intent_type,
    }


def _semantic_summary(regions):
    by_type = {}
    for region in regions:
        region_type = str(region.get("type") or "unknown")
        by_type[region_type] = by_type.get(region_type, 0) + 1
    return {
        "total_regions": len(regions),
        "by_type": by_type,
    }


def build_semantic_regions(template_analysis):
    analysis = template_analysis if isinstance(template_analysis, dict) else {}
    labels = analysis.get("labels") if isinstance(analysis.get("labels"), list) else []
    context = analysis.get("_semantic_context") if isinstance(analysis.get("_semantic_context"), dict) else {}
    explicit_section_keywords = [
        "产品详细要求",
        "包装要求",
        "标签要求",
        "其他要求",
        "配方要求",
        "检测项目",
    ]
    regions = []
    used_region_keys = set()

    def add(
        region_type,
        label,
        source_cell,
        target_cell,
        cells,
        row,
        col,
        confidence,
        reason,
        write_mode,
        intent_type,
        extra=None,
    ):
        key = (region_type, source_cell, target_cell, label)
        if key in used_region_keys:
            return
        used_region_keys.add(key)
        region = _semantic_region(
            f"semantic_{len(regions) + 1:03d}",
            region_type,
            label,
            source_cell,
            target_cell,
            cells,
            row,
            col,
            confidence,
            reason,
            write_mode,
            intent_type,
        )
        if isinstance(extra, dict):
            region.update(extra)
        regions.append(region)

    for table in analysis.get("table_regions", []) if isinstance(analysis.get("table_regions"), list) else []:
        columns = table.get("columns") if isinstance(table.get("columns"), list) else []
        header_cells = [str(column.get("header_cell") or "").strip().upper() for column in columns if column.get("header_cell")]
        if header_cells:
            first_cell = header_cells[0]
            point = _cell_point(first_cell) or (int(table.get("header_row") or 0), 1)
            add(
                "table_region",
                table.get("table_name") or "表格区域",
                first_cell,
                "",
                header_cells,
                point[0],
                point[1],
                0.86,
                "检测到连续表头字段",
                "write_table_column",
                "table_column_header",
                extra={
                    "table_key": table.get("table_key"),
                    "table_name": table.get("table_name"),
                    "semantic_type": table.get("semantic_type"),
                    "columns": columns,
                    "header_cells": header_cells,
                    "header_row": table.get("header_row"),
                    "start_cell": table.get("start_cell"),
                    "bounds": table.get("bounds"),
                    "detection": table.get("detection"),
                },
            )
            for column in columns:
                cell = str(column.get("header_cell") or "").strip().upper()
                row, col = _cell_point(cell) or (int(table.get("header_row") or 0), 0)
                add("table_header", column.get("label") or "", cell, "", [cell], row, col, 0.82, "表格首行字段", "write_table_column", "table_column_header")

    row_option_cells = {}
    for label_item in labels:
        if not isinstance(label_item, dict):
            continue
        text = str(label_item.get("value") or label_item.get("name") or "").strip()
        cell = str(label_item.get("cell") or "").strip().upper()
        point = _cell_point(cell)
        if not text or not point:
            continue
        row, col = point
        info = _semantic_cell_info(context, cell)
        normalized = text.rstrip(":：").strip()
        compact_normalized = re.sub(r"\s+", "", normalized)
        lower_text = text.lower()
        has_colon = "：" in text or ":" in text
        ends_colon = text.endswith(("：", ":"))
        right_target = _semantic_right_target(context, row, col)
        below_target = _semantic_below_target(context, row, col)
        short_row = _semantic_row_short_texts(context, row)
        merged_or_styled = bool(info.get("is_merged")) or bool(info.get("bold")) or bool(info.get("fill_color"))

        if any(word in lower_text for word in ["图片", "照片", "附件", "上传", "附图", "image", "photo", "attachment", "upload"]):
            add("image_attachment_area", normalized, cell, "", [cell], row, col, 0.82, "包含图片/附件关键词", "skip", "image_area")
            continue
        if len(text) >= 22 or any(word in text for word in ["说明", "备注", "注意", "命名原则", "要求如下", "附图片"]):
            add("note_instruction", normalized, cell, "", [cell], row, col, 0.76, "文本较长或包含说明/备注关键词", "skip", "note_instruction")
            continue
        if row <= 3 and (info.get("is_merged") or info.get("font_size", 0) >= 14 or info.get("bold") or info.get("align") == "center") and len(normalized) <= 24:
            add("title", normalized, cell, "", [cell], row, col, 0.84, "顶部区域且具备标题样式", "skip", "title")
            continue
        is_explicit_section_header = any(word in compact_normalized for word in explicit_section_keywords)
        if not is_explicit_section_header and right_target:
            add("field_label", normalized, cell, right_target, [cell, right_target], row, col, 0.84, "存在右侧可填写目标单元格，优先识别为字段", "write_right_cell", "label_fill_right")
            continue
        if not is_explicit_section_header and below_target:
            add("field_label", normalized, cell, below_target, [cell, below_target], row, col, 0.78, "存在下方可填写目标单元格，优先识别为字段", "write_below_cell", "label_fill_below")
            continue
        if is_explicit_section_header or (merged_or_styled and 3 < row <= 80 and len(normalized) <= 18 and not has_colon):
            add("section_header", normalized, cell, "", [cell], row, col, 0.76, "具备分组标题样式或关键词", "skip", "section_header")
            continue
        if any(mark in text for mark in ["□", "☐", "☑", "√", "✔", "[ ]", "( )", "（ ）"]):
            option_value = re.sub(r"[□☐☑√✔\[\]\(\)（）\s]+", "", normalized).strip()
            row_option_cells.setdefault(row, []).append(cell)
            add("option_item", option_value or normalized, cell, cell, [cell], row, col, 0.84, "包含 checkbox-like 符号", "check_option", "option_checkbox")
            continue
        if len(short_row) >= 3 and len(normalized) <= 12 and not has_colon:
            row_option_cells.setdefault(row, []).append(cell)
            add("option_item", normalized, cell, cell, [cell], row, col, 0.68, "同一行存在多个短文本选项", "select_option_text", "option_text_choice")
            continue
        if ends_colon and right_target:
            add("field_label", normalized, cell, right_target, [cell, right_target], row, col, 0.86, "以冒号结尾，右侧存在空白单元格", "write_right_cell", "label_fill_right")
            continue
        if ends_colon and below_target:
            add("field_label", normalized, cell, below_target, [cell, below_target], row, col, 0.8, "以冒号结尾，下方存在空白单元格", "write_below_cell", "label_fill_below")
            continue
        if has_colon:
            add("inline_field", normalized, cell, cell, [cell], row, col, 0.74, "文本包含冒号，适合在原单元格冒号后补值", "append_after_colon", "inline_fill_after_colon")
            continue

        add("unknown", normalized, cell, "", [cell], row, col, 0.35, "暂未匹配到明确语义规则", "skip", "unknown")

    for row, cells in row_option_cells.items():
        if len(cells) < 2:
            continue
        first_cell = cells[0]
        labels_in_row = [_semantic_cell_text(context, row, (_cell_point(cell) or (row, 0))[1]) or cell for cell in cells]
        add("option_group", " / ".join(labels_in_row[:6]), first_cell, "", cells, row, (_cell_point(first_cell) or (row, 0))[1], 0.72, "同一行多个选项项合并为选项组", "select_option_text", "option_text_choice")

    return regions


def analyze_template(template_path):
    labels = scan_excel_labels(template_path)
    structured_mapping_preview = infer_structured_mapping_from_labels(labels)
    table_regions = scan_table_regions(template_path)
    block_regions = scan_block_regions(template_path)
    template_structure = build_template_structure(labels, table_regions, block_regions)
    semantic_context = _build_semantic_context(template_path)
    analysis = {
        "success": True,
        "labels": labels,
        "structured_mapping_preview": structured_mapping_preview,
        "table_regions": table_regions,
        "block_regions": block_regions,
        "template_structure": template_structure,
        "summary": {
            "labels_count": len(labels),
            "structured_mapping_count": len(structured_mapping_preview),
            "table_regions_count": len(table_regions),
            "block_regions_count": len(block_regions),
            "structure_regions_count": len(template_structure.get("regions", [])),
            "raw_regions_count": len(template_structure.get("raw_regions", [])),
            "deduped_regions_count": len(template_structure.get("deduped_regions", [])),
            "recommended_regions_count": len(template_structure.get("recommended_regions", [])),
        },
    }
    analysis["_semantic_context"] = semantic_context
    semantic_regions = build_semantic_regions(analysis)
    analysis.pop("_semantic_context", None)
    analysis["semantic_regions"] = semantic_regions
    analysis["semantic_summary"] = _semantic_summary(semantic_regions)
    analysis["auto_mapping_preview"] = generate_auto_mapping(analysis)

    return analysis

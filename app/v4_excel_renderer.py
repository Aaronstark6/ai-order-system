from datetime import datetime
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from app.excel_writer import safe_write_cell
from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


def _safe_filename_part(value):
    text = str(value or "").strip()
    text = re.sub(r"[^\w.-]+", "_", text, flags=re.UNICODE)
    return text.strip("._") or "example"


def export_description_fields_to_debug_excel(example_name: str, description_fields: dict) -> dict:
    logger.info("V4 debug Excel export started: example_name=%s", example_name)

    try:
        if not isinstance(description_fields, dict):
            raise ValueError("description_fields must be a dict")

        output_dir = get_base_dir() / "v4" / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = _safe_filename_part(example_name)
        filename = f"{safe_name}_debug_excel_{timestamp}.xlsx"
        output_path = output_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "V4 Renderer"

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 80

        safe_write_cell(ws, "A1", "V4 Renderer Debug Export")
        ws["A1"].font = Font(bold=True, size=14)

        safe_write_cell(ws, "A3", "字段名")
        safe_write_cell(ws, "B3", "字段内容")
        ws["A3"].font = Font(bold=True)
        ws["B3"].font = Font(bold=True)

        field_order = ["产品形式", "产品要求", "配方要求", "包装要求"]
        ordered_keys = field_order + [key for key in description_fields.keys() if key not in field_order]

        row = 4
        for field_name in ordered_keys:
            value = description_fields.get(field_name)
            safe_write_cell(ws, f"A{row}", field_name)
            safe_write_cell(ws, f"B{row}", "" if value is None else str(value))
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 48 if "\n" in str(value or "") else 24
            row += 1

        wb.save(output_path)

        logger.info("V4 debug Excel export succeeded: output_path=%s", output_path)
        return {
            "success": True,
            "output_path": str(output_path),
            "filename": filename,
        }
    except Exception as exc:
        logger.exception("V4 debug Excel export failed")
        return {
            "success": False,
            "error": str(exc),
        }

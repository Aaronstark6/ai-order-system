from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook

from app.runtime_paths import get_base_dir


SUPPORTED_CELL_OP_TYPES = {
    "write_text",
    "write_number",
    "write_multiline",
    "write_block",
    "write_table_cell",
}


def _safe_filename_part(value):
    text = re.sub(r"[^\w.-]+", "_", str(value or "").strip(), flags=re.UNICODE)
    return text.strip("._") or "template"


def _operation_worksheet(workbook, operation, warnings):
    sheet_name = str(operation.get("sheet_name") or "").strip()
    if not sheet_name:
        return workbook.active
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    warnings.append(f"指定工作表不存在：{sheet_name}，已使用 active worksheet")
    return workbook.active


def execute_operations_to_excel(template_file, operations):
    template_path = Path(template_file)
    if not template_path.is_file():
        raise FileNotFoundError(f"Excel 模板不存在：{template_path}")
    if not isinstance(operations, list):
        raise TypeError("operations must be a list")

    workbook = load_workbook(template_path)
    warnings = []
    operations_written = 0

    for operation in operations:
        if not isinstance(operation, dict):
            warnings.append("operation 格式无效")
            continue

        op_type = str(operation.get("op_type") or operation.get("type") or "").strip()
        if op_type == "write_image":
            warnings.append("暂不支持图片写入")
            continue
        if op_type not in SUPPORTED_CELL_OP_TYPES:
            warnings.append(f"暂不支持 operation 类型：{op_type or 'unknown'}")
            continue

        target_cell = str(operation.get("target_cell") or "").strip()
        if not target_cell:
            warnings.append("operation 缺少 target_cell")
            continue

        worksheet = _operation_worksheet(workbook, operation, warnings)
        cell = worksheet[target_cell]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            warnings.append(f"公式保护：{worksheet.title}!{target_cell}")
            continue

        value = operation.get("value")
        cell.value = "" if value is None else value
        operations_written += 1

    output_dir = get_base_dir() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"v4_stage2_{_safe_filename_part(template_path.stem)}_{timestamp}.xlsx"
    output_path = output_dir / filename
    workbook.save(output_path)

    return {
        "success": True,
        "filename": filename,
        "output_path": str(output_path),
        "download_url": f"/api/download/{filename}",
        "operations_written": operations_written,
        "warnings": warnings,
    }

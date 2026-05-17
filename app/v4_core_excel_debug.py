from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


DESCRIPTION_FIELD_KEYS = [
    "product_form",
    "product_requirements",
    "formula_requirements",
    "packaging_requirements",
]


def export_description_fields_to_debug_excel(description_fields):
    if not isinstance(description_fields, dict):
        return {
            "success": False,
            "error": "description_fields 必须是对象",
        }

    output_dir = get_base_dir() / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"v4_core_description_fields_debug_{timestamp}.xlsx"
    output_path = output_dir / filename

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "description_fields"

        ws["A1"] = "字段名"
        ws["B1"] = "内容"
        header_fill = PatternFill("solid", fgColor="DDEBF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        for row_index, key in enumerate(DESCRIPTION_FIELD_KEYS, start=2):
            ws.cell(row=row_index, column=1, value=key)
            ws.cell(row=row_index, column=2, value=str(description_fields.get(key) or ""))
            ws.cell(row=row_index, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row=row_index, column=2).alignment = Alignment(vertical="top", wrap_text=True)

            line_count = max(1, str(description_fields.get(key) or "").count("\n") + 1)
            ws.row_dimensions[row_index].height = min(120, max(28, line_count * 18))

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 80
        ws.freeze_panes = "A2"

        wb.save(output_path)
        logger.info("[CoreExcelDebug] Debug Excel generated: path=%s", output_path)
        return {
            "success": True,
            "filename": filename,
            "download_url": f"/api/download/{filename}",
            "output_path": str(output_path),
        }
    except Exception as exc:
        logger.exception("[CoreExcelDebug] Debug Excel generation failed")
        return {
            "success": False,
            "error": str(exc) or "Excel 生成失败",
        }

import json
from json import JSONDecodeError

from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, column_index_from_string

from app.logger import get_logger
from app.runtime_paths import get_base_dir


logger = get_logger(__name__)


def _get_table_mapping_path():
    return get_base_dir() / "v4" / "rules" / "table_mapping.json"


def load_table_mapping():
    mapping_path = _get_table_mapping_path()
    try:
        with mapping_path.open("r", encoding="utf-8") as f:
            mapping = json.load(f)
    except JSONDecodeError as exc:
        logger.error("[TableRenderer] Table mapping JSON parse failed: path=%s error=%s", mapping_path, exc)
        return {}
    except OSError as exc:
        logger.error("[TableRenderer] Table mapping read failed: path=%s error=%s", mapping_path, exc)
        return {}

    return normalize_table_mapping(mapping)


def _infer_table_key(source_path, table_name=""):
    parts = [part for part in str(source_path or "").split(".") if part]
    if parts:
        return parts[-1]
    return str(table_name or "").strip()


def _normalize_column_rule(column):
    if not isinstance(column, dict):
        return None

    label = str(column.get("label") or "").strip()
    field = str(column.get("field") or "").strip()
    target_col = str(column.get("target_col") or "").strip().upper()
    if not label and not field and not target_col:
        return None

    return {
        "label": label or field or target_col,
        "field": field,
        "target_col": target_col,
    }


def _normalize_table_rule(table):
    if not isinstance(table, dict):
        return None

    table_name = str(table.get("table_name") or "").strip()
    source_path = str(table.get("source_path") or "").strip()
    table_key = str(table.get("table_key") or "").strip() or _infer_table_key(source_path, table_name)
    start_cell = str(table.get("start_cell") or "").strip().upper()
    raw_columns = table.get("columns", [])
    if not isinstance(raw_columns, list):
        raw_columns = []

    columns = []
    for column in raw_columns:
        normalized_column = _normalize_column_rule(column)
        if normalized_column:
            columns.append(normalized_column)

    if not table_key and not table_name and not source_path and not start_cell and not columns:
        return None

    return {
        "table_key": table_key,
        "table_name": table_name or table_key or "未命名表格",
        "source_path": source_path,
        "start_cell": start_cell,
        "columns": columns,
    }


def normalize_table_mapping(mapping):
    source = mapping if isinstance(mapping, dict) else {}
    raw_tables = source.get("tables", [])
    if not isinstance(raw_tables, list):
        raw_tables = []

    tables = []
    for table in raw_tables:
        normalized_table = _normalize_table_rule(table)
        if normalized_table:
            tables.append(normalized_table)

    return {
        "version": "V4-Core.18",
        "tables": tables,
    }


def save_table_mapping(mapping):
    mapping_path = _get_table_mapping_path()
    normalized = normalize_table_mapping(mapping)
    try:
        mapping_path.parent.mkdir(parents=True, exist_ok=True)
        with mapping_path.open("w", encoding="utf-8") as f:
            json.dump(normalized, f, ensure_ascii=False, indent=2)
            f.write("\n")
    except OSError as exc:
        logger.exception("[TableRenderer] Table mapping save failed: path=%s", mapping_path)
        return {
            "success": False,
            "error": str(exc) or "动态表格规则保存失败",
        }

    logger.info(
        "[TableRenderer] Table mapping saved: path=%s tables=%s",
        mapping_path,
        len(normalized.get("tables", [])),
    )
    return {
        "success": True,
        "data": normalized,
    }


def _resolve_source_path(data, source_path):
    current = data
    for part in str(source_path or "").split("."):
        if not part:
            return None, False
        if not isinstance(current, dict) or part not in current:
            return None, False
        current = current.get(part)
    return current, True


def _normalize_tables(table_mapping):
    normalized = normalize_table_mapping(table_mapping)
    return normalized.get("tables", [])


def _normalize_columns(table):
    columns = table.get("columns", []) if isinstance(table, dict) else []
    if not isinstance(columns, list):
        return []

    normalized = []
    for column in columns:
        normalized_column = _normalize_column_rule(column)
        if not normalized_column:
            continue
        if normalized_column["field"] and normalized_column["target_col"]:
            normalized.append(normalized_column)
    return normalized


def _start_row_from_cell(start_cell):
    try:
        row, _ = coordinate_to_tuple(str(start_cell or "").strip())
        return row
    except ValueError:
        return None


def _validate_target_col(target_col):
    try:
        column_index_from_string(str(target_col or "").strip().upper())
        return True
    except ValueError:
        return False


def _build_table_range(start_cell, rows_count, columns):
    if not rows_count or not columns:
        return ""
    try:
        start_row, _ = coordinate_to_tuple(str(start_cell or "").strip())
        target_col_indexes = [
            column_index_from_string(column["target_col"])
            for column in columns
            if _validate_target_col(column.get("target_col"))
        ]
    except ValueError:
        return ""

    if not target_col_indexes:
        return ""

    start_col = get_column_letter(min(target_col_indexes))
    end_col = get_column_letter(max(target_col_indexes))
    end_row = start_row + rows_count - 1
    return f"{start_col}{start_row}:{end_col}{end_row}"


def order_object_to_table_operations(order_object, table_mapping):
    warnings = []
    operations = []
    table_summaries = []

    if not isinstance(order_object, dict):
        return {
            "success": False,
            "operations": [],
            "tables": [],
            "warnings": ["Order Object 必须是对象"],
        }

    for table in _normalize_tables(table_mapping):
        if not isinstance(table, dict):
            continue

        table_key = str(table.get("table_key") or "").strip()
        table_name = str(table.get("table_name") or table_key or "").strip() or "未命名表格"
        source_path = str(table.get("source_path") or "").strip()
        start_cell = str(table.get("start_cell") or "").strip().upper()
        start_row = _start_row_from_cell(start_cell)
        columns = _normalize_columns(table)

        if not source_path:
            warnings.append(f"{table_name} 缺少 source_path，已跳过。")
            continue
        if start_row is None:
            warnings.append(f"{table_name} 的 start_cell 无效，已跳过。")
            continue
        if not columns:
            warnings.append(f"{table_name} 缺少 columns，已跳过。")
            continue

        invalid_columns = [column["target_col"] for column in columns if not _validate_target_col(column["target_col"])]
        if invalid_columns:
            warnings.append(f"{table_name} 存在无效目标列：{', '.join(invalid_columns)}，已跳过。")
            continue

        rows, found = _resolve_source_path(order_object, source_path)
        if not found:
            warnings.append(f"未找到表格数据：{table_name}")
            continue
        if not isinstance(rows, list):
            warnings.append(f"表格数据不是数组：{table_name}")
            continue

        rows_count = 0
        for row_index, row_data in enumerate(rows):
            row_number = row_index + 1
            if not isinstance(row_data, dict):
                warnings.append(f"{table_name} 第 {row_number} 行不是对象，已跳过。")
                continue

            rows_count += 1
            excel_row = start_row + row_index
            for column in columns:
                field = column["field"]
                label = column["label"]
                if field not in row_data:
                    warnings.append(f"{table_name} 第 {row_number} 行缺少字段 {field}。")
                value = row_data.get(field, "")
                operations.append({
                    "operation": "write_text",
                    "target_cell": f"{column['target_col']}{excel_row}",
                    "value": "" if value is None else str(value),
                    "table_key": table_key,
                    "table_name": table_name,
                    "row_index": row_index,
                    "row_number": row_number,
                    "field": field,
                    "label": label,
                })

        if rows_count:
            table_summaries.append({
                "table_key": table_key,
                "table_name": table_name,
                "source_path": source_path,
                "start_cell": start_cell,
                "range": _build_table_range(start_cell, rows_count, columns),
                "rows_count": rows_count,
                "columns_count": len(columns),
                "fields": [column["field"] for column in columns],
            })

    logger.info(
        "[TableRenderer] Table operations generated: operations=%s tables=%s warnings=%s",
        len(operations),
        len(table_summaries),
        len(warnings),
    )
    return {
        "success": True,
        "operations": operations,
        "tables": table_summaries,
        "warnings": warnings,
    }
